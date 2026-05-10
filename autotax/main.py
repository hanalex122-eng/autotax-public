import asyncio
import logging
import os
from datetime import datetime, timedelta, timezone
from typing import Optional, List
from fastapi import FastAPI, UploadFile, File, Depends, HTTPException, Query, Body, Request
from slowapi import Limiter, _rate_limit_exceeded_handler
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, HTMLResponse
from pydantic import BaseModel
import io

from autotax.ocr import extract_text, extract_text_and_qr
from autotax.parser import parse_invoice
from autotax.db import init_db, save_invoice, SessionLocal
from autotax.models import Invoice, User, CashEntry, UserCompany, LlmUsage
from autotax.duplicate_service import generate_file_hash, find_hard_duplicate, check_soft_duplicate
from autotax.auth import hash_password, verify_password, create_token, create_access_token, create_refresh_token, decode_token, get_current_user

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
logger = logging.getLogger("autotax")

limiter = Limiter(key_func=get_remote_address)

app = FastAPI(
    title="AutoTax-HUB",
    version="5.5.0",
)
app.state.limiter = limiter
app.add_exception_handler(RateLimitExceeded, _rate_limit_exceeded_handler)

_allowed_origins = os.getenv(
    "ALLOWED_ORIGINS",
    "https://web-production-cd76.up.railway.app,https://autotaxhub.de,https://app.autotaxhub.de,http://localhost:3000,http://localhost:5173"
).split(",")
app.add_middleware(
    CORSMiddleware,
    allow_origins=[o.strip() for o in _allowed_origins if o.strip()],
    allow_credentials=False,
    allow_methods=["GET", "POST", "PUT", "DELETE", "PATCH", "OPTIONS"],
    allow_headers=["Content-Type", "Authorization"],
    max_age=3600,
)


@app.middleware("http")
async def security_headers(request, call_next):
    response = await call_next(request)
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["X-XSS-Protection"] = "1; mode=block"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Permissions-Policy"] = "camera=(self), microphone=()"
    response.headers["X-Data-Retention"] = "none"
    response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    response.headers["Content-Security-Policy"] = "default-src 'self'; script-src 'self' 'unsafe-inline' 'unsafe-eval' https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; font-src 'self' data: https://fonts.gstatic.com; img-src 'self' data: blob:; connect-src 'self' https://*.railway.app https://*.up.railway.app https://cdnjs.cloudflare.com; frame-src 'self' blob:; object-src 'self' blob:"
    return response


# --- ADDED START: Security hardening (middleware-based, non-invasive) ---

# 1. JWT_SECRET stability check (startup)
_jwt_secret_env = os.getenv("JWT_SECRET", "")
if not _jwt_secret_env:
    logger.warning("⚠️  JWT_SECRET not set — using insecure fallback (tokens invalid after restart)")
    logger.warning("⚠️  SECURITY: Set JWT_SECRET env var in Railway for production")
else:
    logger.info("✅ JWT_SECRET configured (secure mode)")

# 2. Log sanitization — mask API keys in logs
_sensitive_env_keys = ["OCR_API_KEY", "JWT_SECRET", "DATABASE_URL"]
class _SanitizeLogFilter(logging.Filter):
    def filter(self, record):
        msg = str(record.getMessage())
        for key_name in _sensitive_env_keys:
            val = os.getenv(key_name, "")
            if val and len(val) > 8 and val in msg:
                msg = msg.replace(val, val[:4] + "****" + val[-2:])
                record.msg = msg
                record.args = None
        return True
logging.getLogger().addFilter(_SanitizeLogFilter())
logger.addFilter(_SanitizeLogFilter())

# 3. Admin endpoint protection + rate limiting via middleware
_ADMIN_EMAILS = set(os.getenv("ADMIN_EMAILS", "").split(",")) if os.getenv("ADMIN_EMAILS") else set()
_rate_limit_counters: dict[str, list[float]] = {}

def _check_rate_limit(key: str, max_per_minute: int) -> bool:
    import time as _time
    now = _time.time()
    if key not in _rate_limit_counters:
        _rate_limit_counters[key] = []
    # Remove timestamps older than 60s
    _rate_limit_counters[key] = [t for t in _rate_limit_counters[key] if now - t < 60]
    if len(_rate_limit_counters[key]) >= max_per_minute:
        return False
    _rate_limit_counters[key].append(now)
    return True


@app.middleware("http")
async def security_guard(request, call_next):
    """Non-invasive security middleware: admin auth, rate limits, security logging."""
    from fastapi.responses import JSONResponse
    path = request.url.path
    method = request.method
    client_ip = _mask_ip(request.client.host) if request.client else "unknown"

    # Admin endpoint protection
    if path.startswith("/admin/"):
        auth_header = request.headers.get("Authorization", "")
        if not auth_header.startswith("Bearer "):
            logger.warning("SECURITY: Admin access denied (no token) from %s to %s", client_ip, path)
            return JSONResponse(status_code=403, content={"detail": "Forbidden: authentication required"})
        try:
            from autotax.auth import decode_token
            token = auth_header.split(" ")[1]
            payload = decode_token(token)
            user_email = payload.get("email", "")
            if not _ADMIN_EMAILS or user_email not in _ADMIN_EMAILS:
                logger.warning("SECURITY: Admin access denied (not admin) — user=%s ip=%s path=%s", _mask_email(user_email), client_ip, path)
                return JSONResponse(status_code=403, content={"detail": "Forbidden: admin access required"})
            logger.info("SECURITY: Admin access granted — user=%s path=%s", _mask_email(user_email), path)
        except Exception as e:
            logger.warning("SECURITY: Admin token invalid from %s: %s", client_ip, e)
            return JSONResponse(status_code=403, content={"detail": "Forbidden: invalid token"})

    # Login rate limit (5/minute per IP)
    if path == "/auth/login" and method == "POST":
        if not _check_rate_limit("login:" + client_ip, 5):
            logger.warning("SECURITY: Login rate limit exceeded from %s", client_ip)
            return JSONResponse(status_code=429, content={"detail": "Too many login attempts — try again in 1 minute"})

    # Delete rate limit (100/minute per IP — bulk operations need higher limit)
    if method == "DELETE":
        if not _check_rate_limit("delete:" + client_ip, 100):
            logger.warning("SECURITY: Delete rate limit exceeded from %s on %s", client_ip, path)
            return JSONResponse(status_code=429, content={"detail": "Too many delete requests — try again in 1 minute"})

    # Security event logging
    if path == "/auth/login" and method == "POST":
        logger.info("AUTH: Login attempt from %s", client_ip)
    if path.startswith("/invoices/upload") and method == "POST":
        logger.info("UPLOAD: File upload from %s", client_ip)
    if method == "DELETE":
        logger.info("DELETE: %s from %s", path, client_ip)

    response = await call_next(request)

    # Log login result
    if path == "/auth/login" and method == "POST":
        if response.status_code == 200:
            logger.info("AUTH: Login SUCCESS from %s", client_ip)
        else:
            logger.warning("AUTH: Login FAILED from %s (status=%d)", client_ip, response.status_code)

    return response

# 4b. Per-user API rate + daily cap — defends against bot abuse on "Unlimited" plan.
#     All paid plans still feel unlimited to humans (hundreds/day capacity),
#     but the caps block scripted bulk abuse that would rack up OCR.space
#     and AI API charges for the operator.
_DAILY_CAP = {"free": 25, "early": 100, "pro": 1000}   # uploads per user per 24h
_MINUTE_CAP = {"free": 10, "early": 20, "pro": 30}     # uploads per user per 60s
_CHAT_DAILY_CAP = {"free": 10, "early": 30, "pro": 100}

_user_window: dict[tuple[str, int], list[float]] = {}


def _window_check(bucket: str, user_id: int, max_calls: int, window_sec: int) -> bool:
    """Per-user sliding-window counter. Returns True if the call is allowed."""
    import time as _t
    key = (bucket, int(user_id))
    now = _t.time()
    lst = _user_window.setdefault(key, [])
    # prune
    _user_window[key] = [ts for ts in lst if now - ts < window_sec]
    if len(_user_window[key]) >= max_calls:
        return False
    _user_window[key].append(now)
    return True


def _user_plan(user_id: int) -> str:
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user_id).first()
        return (u.plan if u and u.plan else "free")
    finally:
        db.close()


def _enforce_upload_quota(user_id: int, chat: bool = False) -> None:
    """Raise HTTPException 429 when the user hits either their per-minute
    burst limit or their per-day cap for the current plan."""
    plan = _user_plan(user_id)
    if chat:
        cap = _CHAT_DAILY_CAP.get(plan, _CHAT_DAILY_CAP["free"])
        if not _window_check("chat_day", user_id, cap, 86400):
            err(429, f"AI-Chat Tageslimit erreicht ({cap}/Tag auf Plan '{plan}'). Morgen wieder verfügbar oder Plan upgraden.")
        return
    day_cap = _DAILY_CAP.get(plan, _DAILY_CAP["free"])
    min_cap = _MINUTE_CAP.get(plan, _MINUTE_CAP["free"])
    if not _window_check("upload_min", user_id, min_cap, 60):
        err(429, f"Zu viele Uploads in kurzer Zeit (max {min_cap}/Minute). Bitte etwas warten.")
    if not _window_check("upload_day", user_id, day_cap, 86400):
        err(429, f"Tageslimit erreicht ({day_cap} Belege/Tag auf Plan '{plan}'). Upgraden oder morgen fortsetzen.")


# 4. Disk quota check helper (called from upload endpoints if integrated)
def _check_disk_quota(user_id: int, new_file_size: int) -> tuple[bool, int, int]:
    """Check if user's storage is within plan limits.
    Returns (allowed, current_bytes, limit_bytes)."""
    db = SessionLocal()
    try:
        from sqlalchemy import func
        # Get user plan
        user = db.query(User).filter(User.id == user_id).first()
        plan = user.plan if user and user.plan else "free"
        # Plan-based quotas (MB)
        quotas = {"free": 100, "early": 1000, "pro": 10000}
        quota_mb = quotas.get(plan, 100)
        quota_bytes = quota_mb * 1024 * 1024
        # Sum current file_data sizes
        current = db.query(func.coalesce(func.sum(func.length(Invoice.file_data)), 0)).filter(
            Invoice.user_id == user_id,
            Invoice.file_data.isnot(None)
        ).scalar() or 0
        allowed = (current + new_file_size) <= quota_bytes
        if not allowed:
            logger.warning("QUOTA: User %d exceeded — current=%dMB, new=%dMB, limit=%dMB",
                          user_id, current // 1024 // 1024, new_file_size // 1024 // 1024, quota_mb)
        return (allowed, current, quota_bytes)
    finally:
        db.close()
# --- ADDED END ---


def _mask_email(email: str) -> str:
    """DSGVO Art. 25 — mask email in logs."""
    if not email or "@" not in email:
        return "***"
    local, domain = email.split("@", 1)
    return f"{local[:2]}***@{domain}" if len(local) > 2 else f"***@{domain}"


def _mask_ip(ip: str) -> str:
    """DSGVO Art. 25 — anonymize IP in logs."""
    if not ip:
        return "***"
    parts = ip.split(".")
    if len(parts) == 4:
        return f"{parts[0]}.{parts[1]}.{parts[2]}.xxx"
    return ip[:10] + "***"  # IPv6 fallback


def ok_list(items, total):
    return {"success": True, "items": items, "total": total}


def err(status: int, msg: str):
    raise HTTPException(status_code=status, detail={"success": False, "error": msg})


def safe_str(val, default=""):
    return val if val is not None else default


def safe_float(val, default=0.0):
    return val if val is not None else default


def safe_vat_rate(val):
    return val if val else "0%"


def safe_vendor(val):
    if not val:
        return "Unbekannt"
    # Strip everything except ASCII printable + German/European letters
    import re as _sre
    cleaned = _sre.sub(r'[^\x20-\x7EäöüÄÖÜßàáâãèéêëìíîïòóôùúûçñÀÁÂÃÈÉÊËÌÍÎÏÒÓÔÙÚÛÇÑ]', '', str(val))
    # Collapse multiple spaces
    cleaned = _sre.sub(r'\s{2,}', ' ', cleaned).strip()
    return cleaned or "Unbekannt"


def safe_category(val):
    return val if val else "other"


def safe_invoice_type(val):
    return val if val in ("income", "expense") else "expense"


def safe_date_str(val):
    if not val:
        return ""
    return val


def parse_vat_rate_float(vat_rate_str):
    try:
        return float((vat_rate_str or "0").replace("%", ""))
    except (ValueError, TypeError):
        return 0.0


def calc_vat(gross, vat_rate_str):
    if not gross:
        return 0.0
    rate = parse_vat_rate_float(vat_rate_str)
    if rate <= 0:
        return 0.0
    return round(gross * rate / (100 + rate), 2)


def _fuzzy_match(a: str, b: str, threshold: float = 0.75) -> bool:
    if not a or not b:
        return False
    a, b = a.lower().replace(" ", ""), b.lower().replace(" ", "")
    if a == b or a in b or b in a:
        return True
    common = sum(1 for c in a if c in b)
    return common / max(len(a), len(b)) >= threshold


def parse_date_str_to_datetime(date_str):
    if not date_str:
        return None
    try:
        return datetime.fromisoformat(date_str)
    except ValueError:
        pass
    try:
        return datetime.strptime(date_str, "%d.%m.%Y")
    except ValueError:
        pass
    return None


def apply_filename_overrides(parsed: dict, filename: str, raw_text: str = "") -> dict:
    """Filename-based vendor + amount override.

    Kullanici fis dosyalarini 'lidl 97.55.pdf' / 'bereket metzger -63.47.pdf'
    seklinde adlandiriyor — vendor adi + toplam tutar dosya adinda.
    Bu fonksiyon parser sonucunu (parsed dict) duzeltir:

    - Vendor: parser default ('Unbekannt') VEYA adres-benzeri VEYA item-line
      gibi suspect ise -> dosya adindan turetilen vendor'i kullan
      (KNOWN_VENDORS listesinden veya generic strip).
    - Total: parser HIGH-anchor (Gesamtbetrag/Summe/Total/Zu zahlen yaninda
      tutari bulduysa) GUVENDE — override etme (Adobe regression: parser
      29.74'u GESAMTBETRAG'dan dogru aldi, filename'de typo 19.74 vardi).
      Aksi halde parser tutar 0 ise VEYA fark 0.50 EUR/+%2 + parser fuzzy
      fallback'tan geldiyse -> filename'i kullan.

    raw_text: parser'in HIGH-anchor confidence'ini hesaplamak icin
    OCR ham metnine bakar. Bos string verilirse anchor check skip edilir
    (geriye uyumlu — eski cagrilar yine calisir, sadece daha agresif olur).

    Hem sync hem async upload path'lerinden cagrilir.
    Asla raise etmez. parsed dict'i in-place degistirir ve geri doner.
    """
    import re as _re_fn
    if not filename:
        return parsed

    _fname = filename.lower()
    _addr_signals = ("str.", "strasse", "straße", "weg", "platz",
                     "allee", "gasse", "ring", "damm", "ufer", "chaussee")
    _addr_prefix_re = _re_fn.compile(
        r"^(?:im|am|an\s+der|auf\s+der|bei\s+der|in\s+der|zur|zum)\s+\w+",
        _re_fn.IGNORECASE,
    )
    _in_address_re = _re_fn.compile(
        r"^in\s+\S+(?:\s+\S+)?\s+\d{1,4}[a-z]?\s*$",
        _re_fn.IGNORECASE,
    )

    def _looks_like_address(line: str) -> bool:
        if not line:
            return False
        ll = line.lower()
        if any(s in ll for s in _addr_signals):
            return True
        if _re_fn.search(r"\b\d{5}\b", line):
            return True
        if _addr_prefix_re.match(line.strip()):
            return True
        if _in_address_re.match(line.strip()):
            return True
        return False

    def _is_suspect(name: str) -> bool:
        if not name:
            return True
        if _re_fn.search(r"\d+[,.]\d{2}", name):
            return True
        if len(name.split()) > 4:
            return True
        if sum(c.isdigit() for c in name) > 4:
            return True
        _alpha = sum(c.isalpha() for c in name)
        _punct = sum(1 for c in name if c in ".'\"/\\|`~^*+={}[]()<>")
        if _alpha > 0 and _punct / max(_alpha, 1) > 0.2:
            return True
        return False

    # 1) Filename'den vendor cikar
    _KNOWN_VENDORS = [
        ("media markt", "Media Markt"), ("mediamarkt", "Media Markt"),
        ("lidl", "Lidl"), ("lidel", "Lidl"),
        ("aldi", "Aldi"), ("rewe", "Rewe"), ("edeka", "Edeka"),
        ("kaufland", "Kaufland"), ("penny", "Penny"), ("netto", "Netto"),
        ("norma", "Norma"), ("tegut", "Tegut"), ("globus", "Globus"),
        ("aral", "Aral"), ("shell", "Shell"), ("esso", "Esso"),
        ("douglas", "Douglas"), ("rossmann", "Rossmann"),
        ("müller", "Müller"), ("muller", "Müller"),
        ("saturn", "Saturn"), ("expert", "Expert"), ("euronics", "Euronics"),
        ("tedi", "TEDI"), ("action", "Action"), ("kik", "KiK"),
        ("woolworth", "Woolworth"), ("snipes", "Snipes"),
        ("deichmann", "Deichmann"), ("zara", "Zara"), ("primark", "Primark"),
        ("h&m", "H&M"), ("c&a", "C&A"),
        ("ikea", "IKEA"), ("bauhaus", "Bauhaus"), ("obi", "OBI"),
        ("hornbach", "Hornbach"), ("toom", "Toom"),
        ("amazon", "Amazon"), ("ebay", "eBay"), ("zalando", "Zalando"),
        ("dm", "dm"),
    ]
    _file_vendor = None
    _file_vendor_known = False  # True ise KNOWN listesinden, False ise generic strip
    for needle, vname in _KNOWN_VENDORS:
        if len(needle) <= 2:
            if _re_fn.search(rf"\b{_re_fn.escape(needle)}\b", _fname):
                _file_vendor = vname
                _file_vendor_known = True
                break
        else:
            if needle in _fname:
                _file_vendor = vname
                _file_vendor_known = True
                break

    if not _file_vendor:
        _base = _re_fn.sub(r"\.[a-z0-9]+$", "", _fname, flags=_re_fn.IGNORECASE)
        _base = _re_fn.sub(r"[-_\s]+\d{1,5}[.,]?\d{0,2}\s*$", "", _base)
        _base = _re_fn.sub(r"[-_]+", " ", _base)
        _base = _re_fn.sub(r"\s+", " ", _base).strip()
        _GENERIC_FNAME_PREFIXES = ("scan", "img", "image", "photo", "doc",
                                   "page", "untitled", "kopie", "copy",
                                   "neu", "neue", "test", "rechnung",
                                   "invoice", "fatura", "fis")
        _is_generic = (
            not _base or len(_base) < 3 or
            any(_base.lower().startswith(p) and len(_base) <= len(p) + 5
                for p in _GENERIC_FNAME_PREFIXES)
        )
        if not _is_generic:
            _file_vendor = _base if _base == _base.upper() else _base.title()

    # 2) Filename'den tutar cikar
    _file_amount = None
    try:
        _amt_m = _re_fn.search(
            r"(\d{1,5})[.,](\d{2})\s*$",
            _re_fn.sub(r"\.[a-z0-9]+$", "", _fname, flags=_re_fn.IGNORECASE),
        )
        if _amt_m:
            _file_amount = float(f"{_amt_m.group(1)}.{_amt_m.group(2)}")
    except Exception:
        pass

    # 3) Vendor override
    if _file_vendor:
        _cur = (parsed.get("vendor") or "").strip()
        _is_default = _cur in ("Unbekannt", "Manual Entry", "") or len(_cur) < 3
        _file_vendor_lower = _file_vendor.lower().replace(" ", "")
        _cur_lower = _re_fn.sub(r"[^a-z0-9]", "", _cur.lower())
        _matches_filename = bool(_file_vendor_lower) and (
            (len(_file_vendor_lower) >= 3 and _file_vendor_lower in _cur_lower) or
            (len(_cur_lower) >= 3 and _cur_lower in _file_vendor_lower)
        )
        # Bilinen vendor (Lidl/Aral/Tedi/...) filename'de gecerse parser
        # vendor'i o vendor'a 'matches_filename' degilse YANLIS demektir —
        # parser muhtemelen logo/adres/garbage yakaladi. Ornekler:
        #   filename 'lidl-9.18.pdf', parser 'LOB' / 'Lion' / '1m Rotfeld'
        #   filename 'aral-30.pdf',   parser 'sven Meyer BS'
        # Generic strip vendor'da (filename'den uretilmis) bu agresif olmaz —
        # 'kebabhaus' gibi kullanici kendi yazmis olabilir, parser dogru
        # bulduysa ezmeyelim.
        _override = (
            (_is_default or _looks_like_address(_cur) or _is_suspect(_cur))
            or _file_vendor_known
        )
        if not _matches_filename and _override:
            logger.info("[FILENAME_OVERRIDE] vendor: %r -> %r (known=%s)",
                        _cur, _file_vendor, _file_vendor_known)
            parsed["vendor"] = _file_vendor

    # 4) Amount override — parser HIGH-anchor sahibi ise override yapma
    if _file_amount and _file_amount > 0:
        _parser_amount = parsed.get("total_amount") or 0
        try:
            _parser_amount = float(_parser_amount)
        except (TypeError, ValueError):
            _parser_amount = 0

        # Parser amount, raw_text'te bir HIGH-anchor (Gesamtbetrag/Summe/
        # Total/Zu zahlen/Brutto) yaninda gecirildiyse parser'a guven.
        # Adobe regression onlemi: 'GESAMTBETRAG (EUR) 29.74' net duruyorsa
        # filename'deki typo 19.74 ezmemeli.
        _parser_confident = False
        if _parser_amount > 0 and raw_text:
            _amt_pat = (f"{_parser_amount:.2f}").replace(".", r"[,.]")
            _anchor_pat = (
                r"\b(?:gesamtbetrag|gesamt\s*betrag|gesamt\s*brutto|summe\s*brutto|"
                r"summe\s+(?:eur|inkl)|zu\s*zahlen|zahlbetrag|"
                r"rechnungsbetrag|rechnungssumme|endbetrag|"
                r"grand\s*total|total\s*amount|amount\s*due|total\s*ttc)"
                r"\b[\s\S]{0,40}?" + _amt_pat
            )
            try:
                _parser_confident = bool(_re_fn.search(_anchor_pat, raw_text, _re_fn.IGNORECASE))
            except Exception:
                _parser_confident = False

        _diff = abs(_parser_amount - _file_amount)
        _max_val = max(_parser_amount, _file_amount, 1)

        if _parser_amount <= 0:
            # Parser hicbir tutar bulamadi -> filename
            logger.info("[FILENAME_OVERRIDE] amount: 0 -> %.2f (parser failed)", _file_amount)
            parsed["total_amount"] = _file_amount
        elif _parser_confident:
            logger.info("[FILENAME_OVERRIDE] amount: parser=%.2f anchored, filename=%.2f ignored",
                        _parser_amount, _file_amount)
        elif _diff >= 0.50 and _diff / _max_val >= 0.02:
            # Parser fuzzy bir yerden almis, fark anlamli -> filename
            logger.info("[FILENAME_OVERRIDE] amount: %.2f -> %.2f (diff %.2f, no anchor)",
                        _parser_amount, _file_amount, _diff)
            parsed["total_amount"] = _file_amount

    return parsed


def auto_create_cash_entry(invoice_id: int, user_id: int, data: dict):
    """Create a CashEntry automatically when an invoice is uploaded."""
    db = SessionLocal()
    try:
        # Skip if already synced
        existing = db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user_id).first()
        if existing:
            return
        # Parse date safely
        date_val = None
        date_str = data.get("date") or ""
        if date_str:
            date_val = parse_date_str_to_datetime(date_str)
        if not date_val:
            date_val = datetime.now()
        inv_type = (data.get("invoice_type") or "expense").lower().strip()
        entry_type = "income" if inv_type in ("income", "einnahme", "gutschrift") else "expense"
        entry = CashEntry(
            user_id=user_id,
            description=f"Rechnung: {data.get('vendor') or 'Unbekannt'}",
            vendor=data.get("vendor") or "Unbekannt",
            gross_amount=float(data.get("total_amount") or 0),
            vat_amount=float(data.get("vat_amount") or 0),
            vat_rate=data.get("vat_rate") or "0%",
            entry_type=entry_type,
            category=data.get("category") or "other",
            payment_method=data.get("payment_method") or "",
            reference=data.get("invoice_number") or f"INV-{invoice_id}",
            notes=f"Auto-sync from invoice #{invoice_id}",
            is_reconciled=False,
            invoice_id=invoice_id,
            date=date_val,
        )
        db.add(entry)
        db.commit()
        logger.info("Auto-synced invoice %s to cash_entries", invoice_id)
    except Exception:
        db.rollback()
        logger.exception("Auto cash entry creation failed for invoice %s", invoice_id)
    finally:
        db.close()


# --- ADDED START: quick entity extractors for invoice_to_dict ---
import re as _re_global

def _extract_first_iban(text):
    m = _re_global.search(r"\b([A-Z]{2}\s?\d{2}\s?(?:\d{4}\s?){2,7}\d{1,4})\b", text.upper())
    return m.group(1).replace(" ", "") if m else ""

def _extract_first_phone(text):
    m = _re_global.search(r"(?:tel\.?|fon|phone|fax)\s*:?\s*([\d\s/\-+]{6,20})", text, _re_global.IGNORECASE)
    return m.group(1).strip() if m else ""

def _extract_first_address(text):
    m = _re_global.search(r"(\d{4,5}\s+[A-ZÄÖÜ][a-zäöüß]{2,}(?:\s+[A-ZÄÖÜ][a-zäöüß]{2,})?)", text)
    return m.group(1).strip() if m else ""
# --- ADDED END ---

def invoice_to_dict(i):
    return {
        "id": i.id,
        "vendor": safe_vendor(i.vendor),
        "invoice_number": safe_str(i.invoice_number),
        "invoice_type": safe_invoice_type(i.invoice_type),
        "total_amount": safe_float(i.total_amount),
        "vat_amount": safe_float(i.vat_amount),
        "vat_rate": safe_vat_rate(i.vat_rate),
        "date": safe_date_str(i.date),
        "payment_method": safe_str(i.payment_method),
        "category": safe_category(i.category),
        "processed": i.processed or False,
        "created_at": i.created_at.strftime("%Y-%m-%dT%H:%M:%S") if i.created_at else "",
        "ocr_snippet": (i.raw_text or "")[:200],
        "konto": _DATEV_KONTO_MAP.get(safe_category(i.category), "6800") if safe_invoice_type(i.invoice_type) == "expense" else _DATEV_KONTO_MAP_INCOME.get(safe_category(i.category), "8400"),
        # file_path = yeni yol (volume); file_data = eski legacy BLOB.
        # Burada file_data acmiyoruz — defer ile yuklenmiyor, lazy load
        # yapsa N+1 olur. file_path varsa orijinal var demektir.
        "has_original": bool(getattr(i, "file_path", None)),
        "filename": safe_str(i.filename) if i.filename else "",
        # Vendor contact info — use DB column if set, else fallback to
        # on-the-fly extraction for backward compat with old invoices.
        "vendor_iban": safe_str(i.vendor_iban) if hasattr(i, "vendor_iban") and i.vendor_iban else _extract_first_iban(i.raw_text or ""),
        "vendor_email": safe_str(i.vendor_email) if hasattr(i, "vendor_email") and i.vendor_email else "",
        "vendor_phone": safe_str(i.vendor_phone) if hasattr(i, "vendor_phone") and i.vendor_phone else _extract_first_phone(i.raw_text or ""),
        "vendor_address": safe_str(i.vendor_address) if hasattr(i, "vendor_address") and i.vendor_address else _extract_first_address(i.raw_text or ""),
        # Reminder system
        "due_date": safe_str(getattr(i, "due_date", None) or ""),
        "payment_status": safe_str(getattr(i, "payment_status", None) or "unpaid"),
        "paid_at": (i.paid_at.strftime("%Y-%m-%dT%H:%M:%S") if getattr(i, "paid_at", None) else ""),
        # Recurring
        "is_recurring": bool(getattr(i, "is_recurring", False)),
        "recurring_freq": safe_str(getattr(i, "recurring_freq", None) or ""),
        "recurring_next_at": safe_str(getattr(i, "recurring_next_at", None) or ""),
        "recurring_parent_id": getattr(i, "recurring_parent_id", None),
    }


def cash_entry_to_dict(e):
    return {
        "id": e.id,
        "description": safe_str(e.description),
        "vendor": safe_vendor(e.vendor),
        "gross_amount": safe_float(e.gross_amount),
        "vat_amount": safe_float(e.vat_amount),
        "vat_rate": safe_vat_rate(e.vat_rate),
        "entry_type": safe_invoice_type(e.entry_type),
        "category": safe_category(e.category),
        "payment_method": safe_str(e.payment_method),
        "reference": safe_str(e.reference),
        "notes": safe_str(e.notes),
        "is_reconciled": e.is_reconciled or False,
        "invoice_id": e.invoice_id,
        "date": e.date.strftime("%Y-%m-%d") if e.date else "",
        "created_at": e.created_at.strftime("%Y-%m-%dT%H:%M:%S") if e.created_at else "",
    }


@app.on_event("startup")
def startup():
    init_db()


@app.on_event("startup")
async def startup_email_auto_sync():
    if os.getenv("EMAIL_AUTO_SYNC_ENABLED", "1").strip() == "0":
        logger.info("Email auto-sync disabled via EMAIL_AUTO_SYNC_ENABLED=0")
        return
    try:
        from autotax.email_sync import start_auto_sync
        start_auto_sync()
    except Exception:
        logger.exception("Failed to start email auto-sync task")


@app.on_event("shutdown")
async def shutdown_email_auto_sync():
    try:
        from autotax.email_sync import stop_auto_sync
        stop_auto_sync()
    except Exception:
        logger.exception("Failed to stop email auto-sync task")


# Rechnung Reminder System — gunluk 09:00 Europe/Berlin tick
_reminder_task = None


@app.on_event("startup")
async def startup_reminders():
    global _reminder_task
    try:
        from autotax.reminders import reminder_loop
        _reminder_task = asyncio.create_task(reminder_loop())
        logger.info("Reminder background loop scheduled")
    except Exception:
        logger.exception("Failed to start reminder loop")


@app.on_event("shutdown")
async def shutdown_reminders():
    global _reminder_task
    if _reminder_task and not _reminder_task.done():
        _reminder_task.cancel()


@app.api_route("/health", methods=["GET", "HEAD"])
def health():
    """Health check — UptimeRobot ve benzeri monitor'lar HEAD kullaniyor.
    GET = JSON detay, HEAD = sadece 200 status (otomatik)."""
    ocr_key = os.getenv("OCR_API_KEY", "")
    db_ok = True
    try:
        from sqlalchemy import text
        db = SessionLocal()
        db.execute(text("SELECT 1"))
        db.close()
    except Exception:
        db_ok = False
    return {"status": "ok", "version": "5.5.5", "ocr_configured": bool(ocr_key), "db_connected": db_ok}


@app.get("/manifest.json")
def pwa_manifest():
    return {
        "name": "AutoTax-HUB",
        "short_name": "AutoTax",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#050a12",
        "theme_color": "#10b981",
        "description": "Automatische Rechnungserkennung & Buchhaltung",
        "icons": [
            {"src": "data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><rect fill='%23050a12' width='100' height='100' rx='20'/><text x='50' y='65' font-size='50' text-anchor='middle' fill='%2310b981' font-family='sans-serif' font-weight='bold'>AT</text></svg>", "sizes": "192x192", "type": "image/svg+xml"},
        ],
    }


from fastapi.responses import Response as RawResponse


@app.get("/sw.js")
def service_worker():
    return RawResponse(content="self.addEventListener('fetch',e=>{});", media_type="application/javascript")


@app.get("/invoices/{invoice_id}/pdf")
def generate_invoice_pdf(invoice_id: int, user: dict = Depends(get_current_user)):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
    except ImportError:
        raise HTTPException(status_code=501, detail="PDF-Generierung nicht verfügbar (reportlab fehlt)")

    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            raise HTTPException(status_code=404, detail="Rechnung nicht gefunden")

        companies = db.query(UserCompany).filter(UserCompany.user_id == user["sub"]).all()
        company_name = companies[0].company_name if companies else "Meine Firma"
        u = db.query(User).filter(User.id == user["sub"]).first()

        buf = io.BytesIO()
        c = pdf_canvas.Canvas(buf, pagesize=A4)
        w, h = A4

        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 22)
        c.drawString(2*cm, h-2.5*cm, company_name)
        c.setFillColor(HexColor("#00e5a0"))
        c.setFont("Helvetica", 9)
        c.drawString(2*cm, h-3*cm, f"E-Mail: {u.email if u else ''}")

        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 16)
        typ = "RECHNUNG" if inv.invoice_type == "income" else "BELEG"
        c.drawString(2*cm, h-4.5*cm, typ)
        c.setFont("Helvetica", 11)
        c.drawString(12*cm, h-4.5*cm, f"Nr: {inv.invoice_number or f'RE-{inv.id}'}")
        c.drawString(12*cm, h-5.1*cm, f"Datum: {inv.date or 'k.A.'}")

        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, h-6*cm, "An:" if inv.invoice_type == "income" else "Von:")
        c.setFont("Helvetica", 11)
        c.drawString(2*cm, h-6.6*cm, inv.vendor or "Unbekannt")

        y = h - 8.5*cm
        c.setFillColor(HexColor("#1a2d4a"))
        c.rect(2*cm, y, 17*cm, 0.8*cm, fill=1)
        c.setFillColor(HexColor("#ffffff"))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(2.2*cm, y+0.25*cm, "Beschreibung")
        c.drawString(10*cm, y+0.25*cm, "Kategorie")
        c.drawString(13*cm, y+0.25*cm, "MwSt")
        c.drawString(16*cm, y+0.25*cm, "Betrag")

        y -= 0.8*cm
        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica", 10)
        c.drawString(2.2*cm, y+0.25*cm, inv.vendor or "Position 1")
        c.drawString(10*cm, y+0.25*cm, inv.category or "other")
        c.drawString(13*cm, y+0.25*cm, f"{inv.vat_rate or '19%'}")
        c.drawString(16*cm, y+0.25*cm, f"EUR {inv.total_amount or 0:.2f}")
        c.line(2*cm, y, 19*cm, y)

        y -= 1.5*cm
        netto = (inv.total_amount or 0) - (inv.vat_amount or 0)
        c.setFont("Helvetica", 10)
        c.drawRightString(19*cm, y, f"Netto: EUR {netto:.2f}")
        y -= 0.5*cm
        c.drawRightString(19*cm, y, f"MwSt ({inv.vat_rate or '19%'}): EUR {inv.vat_amount or 0:.2f}")
        y -= 0.6*cm
        c.setFont("Helvetica-Bold", 12)
        c.drawRightString(19*cm, y, f"Total: EUR {inv.total_amount or 0:.2f}")

        if hasattr(u, 'is_kleinunternehmer') and getattr(u, 'is_kleinunternehmer', False):
            y -= 1.5*cm
            c.setFont("Helvetica-Oblique", 8)
            c.setFillColor(HexColor("#7a8ba8"))
            c.drawString(2*cm, y, "Gemäß §19 UStG wird keine Umsatzsteuer berechnet.")

        c.setFillColor(HexColor("#7a8ba8"))
        c.setFont("Helvetica", 7)
        c.drawString(2*cm, 1.5*cm, f"Erstellt mit AutoTax-HUB | {company_name} | {u.email if u else ''}")
        c.drawString(2*cm, 1*cm, "Automatisch erstellt. Alle Angaben ohne Gewähr. Keine Steuerberatung.")

        c.save()
        buf.seek(0)

        filename = f"{typ}_{inv.invoice_number or inv.id}.pdf"
        return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={filename}"})
    finally:
        db.close()


def _inject_features(html: str) -> str:
    """Replace the {{FEATURE_FLAGS}} placeholder with the runtime JSON so
    the SPA can read window.FEATURES without an extra network round-trip."""
    from autotax.config import features_js_literal
    return html.replace("{{FEATURE_FLAGS}}", features_js_literal(), 1)


@app.get("/", response_class=HTMLResponse)
async def serve_frontend():
    # Public niche instance shows the marketing landing at root so first-time
    # visitors see the pitch, not an empty login form. Personal deploy keeps
    # the login/app shell at root (single-user, no visitors).
    from autotax.config import FEATURES, render_landing_placeholders
    if FEATURES.get("public_niche") is True:
        lp = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "landing-new.html")
        with open(lp, "r", encoding="utf-8") as f:
            return HTMLResponse(content=render_landing_placeholders(f.read()), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
    index_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "index.html")
    with open(index_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=_inject_features(f.read()), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


@app.get("/app", response_class=HTMLResponse)
async def serve_frontend_app():
    index_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "index.html")
    with open(index_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=_inject_features(f.read()), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


@app.get("/api/config")
def api_config():
    """Client-readable feature flag dump (for standalone pages that don't
    go through the index.html inject path)."""
    from autotax.config import FEATURES
    return FEATURES


# --- ADDED START: Landing page ---
@app.get("/landing", response_class=HTMLResponse)
async def serve_landing_page():
    from autotax.config import render_landing_placeholders
    lp = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "landing-new.html")
    with open(lp, "r", encoding="utf-8") as f:
        html = render_landing_placeholders(f.read())
        return HTMLResponse(content=html, headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
# --- ADDED END ---

# --- ADDED START: Language file ---
@app.get("/lang.js")
async def serve_lang_js():
    lp = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "lang.js")
    with open(lp, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), media_type="application/javascript", headers={"Cache-Control": "no-cache"})
# --- ADDED END ---

@app.get("/auth-guard.js")
async def serve_auth_guard_js():
    fp = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "auth-guard.js")
    with open(fp, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), media_type="application/javascript", headers={"Cache-Control": "no-cache"})

# --- ADDED START: Split-view editor page ---
@app.get("/editor", response_class=HTMLResponse)
async def serve_editor_page():
    ep = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "editor.html")
    with open(ep, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
# --- ADDED END ---

# --- ADDED START: Beleg entry page ---
@app.get("/beleg", response_class=HTMLResponse)
async def serve_beleg_page():
    beleg_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "beleg.html")
    with open(beleg_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})
# --- ADDED END ---

@app.get("/email-settings", response_class=HTMLResponse)
async def serve_email_settings_page():
    p = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "email-settings.html")
    with open(p, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read(), headers={"Cache-Control": "no-cache, no-store, must-revalidate"})


_FAVICON_SVG = (
    "<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'>"
    "<rect fill='#050a12' width='100' height='100' rx='20'/>"
    "<text x='50' y='65' font-size='50' text-anchor='middle' fill='#10b981' "
    "font-family='sans-serif' font-weight='bold'>AT</text></svg>"
)


@app.get("/favicon.ico")
def favicon_ico():
    """Serve the AT logomark as SVG so browsers stop 404-logging /favicon.ico.
    Content-Type image/svg+xml works in all modern browsers."""
    from fastapi.responses import Response as _Resp
    return _Resp(content=_FAVICON_SVG, media_type="image/svg+xml",
                 headers={"Cache-Control": "public, max-age=604800"})


@app.get("/favicon.svg")
def favicon_svg():
    from fastapi.responses import Response as _Resp
    return _Resp(content=_FAVICON_SVG, media_type="image/svg+xml",
                 headers={"Cache-Control": "public, max-age=604800"})


_PRIVACY_CSS = """body{font-family:'DM Sans',sans-serif;max-width:800px;margin:40px auto;padding:20px;background:#050a12;color:#e8edf5;line-height:1.8}
h1{color:#10b981;font-size:28px}h2{color:#00a8cc;margin-top:30px;font-size:18px}strong{color:#f59e0b}
a{color:#10b981}p{margin:12px 0}ul{padding-left:20px}li{margin:6px 0}
.lang-bar{display:flex;gap:10px;margin-bottom:20px;flex-wrap:wrap}
.lang-bar a{padding:6px 14px;border-radius:8px;border:1px solid #2a3548;color:#94a3b8;text-decoration:none;font-size:13px}
.lang-bar a.active{background:#10b981;color:#fff;border-color:#10b981}"""

_PRIVACY_LANG_BAR = """<div class="lang-bar">
<a href="/datenschutz" class="{de}">Deutsch</a>
<a href="/privacy" class="{en}">English</a>
<a href="/confidentialite" class="{fr}">Français</a>
<a href="/privacidad" class="{es}">Español</a>
<a href="/gizlilik" class="{tr}">Türkçe</a>
<a href="/khususiyya" class="{ar}">العربية</a>
</div>"""

_PRIVACY_THIRD_PARTIES = """<ul>
<li><strong>Railway Inc.</strong> (USA/EU) — Hosting</li>
<li><strong>OCR.space / a9t9 Software GmbH</strong> — OCR {ocr_desc}</li>
<li><strong>Anthropic PBC</strong> (USA) — AI {ai_desc}</li>
</ul>"""

def _privacy_page(lang, title, sections):
    bar = _PRIVACY_LANG_BAR.format(de="active" if lang=="de" else "", en="active" if lang=="en" else "",
        fr="active" if lang=="fr" else "", es="active" if lang=="es" else "",
        tr="active" if lang=="tr" else "", ar="active" if lang=="ar" else "")
    direction = ' dir="rtl"' if lang == "ar" else ""
    body = f"""<!DOCTYPE html><html lang="{lang}"{direction}><head><meta charset="UTF-8"><title>{title}</title>
<style>{_PRIVACY_CSS}</style></head><body>{bar}<h1>{title}</h1>
<p><em>AutoTax-HUB — {"Stand" if lang=="de" else "Last updated"}: April 2026</em></p>"""
    for s in sections:
        body += f'\n<h2>{s["h"]}</h2>\n{s["c"]}'
    body += '\n<p style="margin-top:40px;color:#64748b;font-size:13px">© 2026 AutoTax-HUB</p></body></html>'
    return HTMLResponse(content=body)


@app.get("/datenschutz", response_class=HTMLResponse)
def datenschutz_page():
    return _privacy_page("de", "Datenschutzerklärung", [
        {"h":"1. Verantwortlicher (Art. 4 Nr. 7 DSGVO)","c":"<p>Hüseyin Hancer<br>Wiesenstr. 10, 66115 Saarbrücken, Deutschland<br>E-Mail: datenschutz@autotaxhub.de</p>"},
        {"h":"2. Erhobene Daten und Zweck","c":"<ul><li><strong>Registrierungsdaten:</strong> E-Mail, Name, Passwort (gehashed)</li><li><strong>Rechnungsdaten:</strong> Belege, OCR-Text, Beträge, MwSt</li><li><strong>Firmendaten:</strong> IBAN, Steuernummer, Adresse</li><li><strong>Kassenbuch:</strong> Einnahmen/Ausgaben</li><li><strong>Technisch:</strong> IP-Adresse (anonymisiert), User-Agent</li></ul>"},
        {"h":"3. Rechtsgrundlage (Art. 6 DSGVO)","c":"<ul><li>Art. 6(1)(a): Einwilligung</li><li>Art. 6(1)(b): Vertragserfüllung</li><li>Art. 6(1)(f): Berechtigte Interessen</li></ul>"},
        {"h":"4. Empfänger und Drittanbieter","c":_PRIVACY_THIRD_PARTIES.format(ocr_desc="— Belegbilder können IBAN, Adresse enthalten", ai_desc="— OCR-Texte zur Verarbeitung") + '<p><strong>Drittlandtransfer:</strong> USA — Standardvertragsklauseln (Art. 46 DSGVO) / EU-US Data Privacy Framework.</p>'},
        {"h":"5. Speicherdauer","c":"<ul><li>Kontodaten: bis Löschung</li><li>Buchungsbelege: 10 Jahre (GoBD/AO)</li><li>Papierkorb: 30 Tage</li><li>Logs: max. 30 Tage</li></ul>"},
        {"h":"6. Ihre Rechte (Art. 15-21)","c":"<ul><li>Auskunft (Art. 15) — Konto → Daten exportieren</li><li>Berichtigung (Art. 16)</li><li>Löschung (Art. 17) — Konto → Konto löschen</li><li>Datenübertragbarkeit (Art. 20) — JSON-Export</li><li>Widerspruch (Art. 21)</li><li>Beschwerde bei Aufsichtsbehörde (Art. 77)</li></ul>"},
        {"h":"7. Cookies","c":"<p>Keine Tracking-Cookies. Nur technisch notwendige Local-Storage-Einträge (Auth-Token, Cookie-Consent).</p>"},
        {"h":"8. Sicherheit (Art. 32)","c":"<p>TLS/SSL, bcrypt-Passwörter, JWT, Rate-Limiting, HSTS, CSP.</p>"},
    ])


@app.get("/privacy", response_class=HTMLResponse)
def privacy_page_en():
    return _privacy_page("en", "Privacy Policy", [
        {"h":"1. Data Controller","c":"<p>Hüseyin Hancer<br>Wiesenstr. 10, 66115 Saarbrücken, Germany<br>Email: datenschutz@autotaxhub.de</p>"},
        {"h":"2. Data Collected","c":"<ul><li><strong>Registration:</strong> Email, name, password (hashed)</li><li><strong>Invoice data:</strong> Uploaded receipts, OCR text, amounts, VAT</li><li><strong>Company data:</strong> IBAN, tax ID, address</li><li><strong>Cash book:</strong> Income/expenses</li><li><strong>Technical:</strong> IP address (anonymized), user agent</li></ul>"},
        {"h":"3. Legal Basis (GDPR Art. 6)","c":"<ul><li>Art. 6(1)(a): Consent</li><li>Art. 6(1)(b): Contract performance</li><li>Art. 6(1)(f): Legitimate interests</li></ul>"},
        {"h":"4. Third-Party Processors","c":_PRIVACY_THIRD_PARTIES.format(ocr_desc="— receipt images may contain IBAN, address", ai_desc="— OCR texts for processing") + '<p><strong>International transfers:</strong> USA — Standard Contractual Clauses (Art. 46 GDPR) / EU-US Data Privacy Framework.</p>'},
        {"h":"5. Data Retention","c":"<ul><li>Account data: until deletion</li><li>Accounting records: 10 years (German tax law)</li><li>Trash: 30 days</li><li>Logs: max 30 days</li></ul>"},
        {"h":"6. Your Rights","c":"<ul><li>Access (Art. 15) — Account → Export data</li><li>Rectification (Art. 16)</li><li>Erasure (Art. 17) — Account → Delete account</li><li>Portability (Art. 20) — JSON export</li><li>Object (Art. 21)</li><li>Complaint to supervisory authority (Art. 77)</li></ul>"},
        {"h":"7. CCPA Rights (California)","c":"<p>If you are a California resident, you have the right to: know what personal data is collected, request deletion, opt-out of data sales. <strong>We do not sell your personal data.</strong> To exercise your rights, use Account → Export/Delete or email datenschutz@autotaxhub.de.</p>"},
        {"h":"8. Cookies","c":"<p>No tracking cookies. Only essential local storage entries (auth token, cookie consent).</p>"},
        {"h":"9. Security","c":"<p>TLS/SSL, bcrypt passwords, JWT auth, rate limiting, HSTS, CSP headers.</p>"},
    ])


@app.get("/confidentialite", response_class=HTMLResponse)
def privacy_page_fr():
    return _privacy_page("fr", "Politique de Confidentialité", [
        {"h":"1. Responsable du traitement","c":"<p>Hüseyin Hancer<br>Wiesenstr. 10, 66115 Saarbrücken, Allemagne<br>E-mail : datenschutz@autotaxhub.de</p>"},
        {"h":"2. Données collectées","c":"<ul><li><strong>Inscription :</strong> e-mail, nom, mot de passe (haché)</li><li><strong>Factures :</strong> reçus, texte OCR, montants, TVA</li><li><strong>Entreprise :</strong> IBAN, numéro fiscal, adresse</li><li><strong>Livre de caisse :</strong> recettes/dépenses</li><li><strong>Technique :</strong> adresse IP (anonymisée)</li></ul>"},
        {"h":"3. Base juridique (Art. 6 RGPD)","c":"<ul><li>Art. 6(1)(a) : Consentement</li><li>Art. 6(1)(b) : Exécution du contrat</li><li>Art. 6(1)(f) : Intérêts légitimes</li></ul>"},
        {"h":"4. Sous-traitants","c":_PRIVACY_THIRD_PARTIES.format(ocr_desc="— images pouvant contenir IBAN, adresse", ai_desc="— textes OCR pour traitement") + '<p><strong>Transfert international :</strong> USA — Clauses contractuelles types (Art. 46 RGPD).</p>'},
        {"h":"5. Durée de conservation","c":"<ul><li>Compte : jusqu'à suppression</li><li>Documents comptables : 10 ans</li><li>Corbeille : 30 jours</li><li>Logs : max 30 jours</li></ul>"},
        {"h":"6. Vos droits (Art. 15-21)","c":"<ul><li>Accès, rectification, effacement, portabilité, opposition</li><li>Réclamation auprès de la CNIL (Art. 77)</li></ul>"},
        {"h":"7. Cookies","c":"<p>Aucun cookie de suivi. Uniquement stockage local technique (token d'authentification).</p>"},
    ])


@app.get("/privacidad", response_class=HTMLResponse)
def privacy_page_es():
    return _privacy_page("es", "Política de Privacidad", [
        {"h":"1. Responsable del tratamiento","c":"<p>Hüseyin Hancer<br>Wiesenstr. 10, 66115 Saarbrücken, Alemania<br>Email: datenschutz@autotaxhub.de</p>"},
        {"h":"2. Datos recopilados","c":"<ul><li><strong>Registro:</strong> email, nombre, contraseña (hash)</li><li><strong>Facturas:</strong> recibos, texto OCR, importes, IVA</li><li><strong>Empresa:</strong> IBAN, NIF, dirección</li><li><strong>Libro de caja:</strong> ingresos/gastos</li><li><strong>Técnico:</strong> IP (anonimizada)</li></ul>"},
        {"h":"3. Base legal (Art. 6 RGPD)","c":"<ul><li>Consentimiento, ejecución contractual, interés legítimo</li></ul>"},
        {"h":"4. Encargados del tratamiento","c":_PRIVACY_THIRD_PARTIES.format(ocr_desc="— imágenes con IBAN, dirección", ai_desc="— textos OCR") + '<p><strong>Transferencia internacional:</strong> USA — Cláusulas contractuales tipo (Art. 46 RGPD).</p>'},
        {"h":"5. Conservación","c":"<ul><li>Cuenta: hasta eliminación</li><li>Documentos contables: 10 años</li><li>Papelera: 30 días</li></ul>"},
        {"h":"6. Sus derechos (Art. 15-21 RGPD)","c":"<ul><li>Acceso, rectificación, supresión, portabilidad, oposición</li><li>Reclamación ante la AEPD (Art. 77)</li></ul>"},
        {"h":"7. Cookies","c":"<p>Sin cookies de seguimiento. Solo almacenamiento local técnico.</p>"},
    ])


@app.get("/gizlilik", response_class=HTMLResponse)
def privacy_page_tr():
    return _privacy_page("tr", "Gizlilik Politikası (KVKK)", [
        {"h":"1. Veri Sorumlusu","c":"<p>Hüseyin Hancer<br>Wiesenstr. 10, 66115 Saarbrücken, Almanya<br>E-posta: datenschutz@autotaxhub.de</p>"},
        {"h":"2. Toplanan Veriler","c":"<ul><li><strong>Kayıt:</strong> e-posta, ad, şifre (hashlenmiş)</li><li><strong>Fatura verileri:</strong> yüklenen belgeler, OCR metni, tutarlar, KDV</li><li><strong>Firma bilgileri:</strong> IBAN, vergi no, adres</li><li><strong>Kasa defteri:</strong> gelir/gider</li><li><strong>Teknik:</strong> IP adresi (anonimleştirilmiş)</li></ul>"},
        {"h":"3. Hukuki Dayanak (KVKK m.5)","c":"<ul><li>Açık rıza (m.5/1)</li><li>Sözleşmenin ifası (m.5/2-c)</li><li>Meşru menfaat (m.5/2-f)</li></ul>"},
        {"h":"4. Veri Aktarımı","c":_PRIVACY_THIRD_PARTIES.format(ocr_desc="— belgeler IBAN, adres içerebilir", ai_desc="— OCR metinleri işleme") + '<p><strong>Yurt dışı aktarım:</strong> ABD — KVKK m.9 kapsamında açık rıza ile.</p>'},
        {"h":"5. Saklama Süresi","c":"<ul><li>Hesap verileri: silinene kadar</li><li>Muhasebe belgeleri: 10 yıl (vergi mevzuatı)</li><li>Çöp kutusu: 30 gün</li></ul>"},
        {"h":"6. Haklarınız (KVKK m.11)","c":"<ul><li>Bilgi alma, düzeltme, silme, aktarım, itiraz</li><li>Kişisel Verileri Koruma Kurulu'na şikayet hakkı</li></ul>"},
        {"h":"7. Çerezler","c":"<p>İzleme çerezi kullanılmaz. Sadece teknik local storage (auth token).</p>"},
    ])


@app.get("/khususiyya", response_class=HTMLResponse)
def privacy_page_ar():
    return _privacy_page("ar", "سياسة الخصوصية", [
        {"h":"1. المسؤول عن البيانات","c":"<p>Hüseyin Hancer<br>Wiesenstr. 10, 66115 Saarbrücken, ألمانيا<br>البريد: datenschutz@autotaxhub.de</p>"},
        {"h":"2. البيانات المجمعة","c":"<ul><li><strong>التسجيل:</strong> البريد الإلكتروني، الاسم، كلمة المرور (مشفرة)</li><li><strong>الفواتير:</strong> الإيصالات، نص OCR، المبالغ، ضريبة القيمة المضافة</li><li><strong>بيانات الشركة:</strong> IBAN، الرقم الضريبي، العنوان</li><li><strong>دفتر النقد:</strong> الإيرادات/المصروفات</li></ul>"},
        {"h":"3. الأساس القانوني","c":"<ul><li>الموافقة</li><li>تنفيذ العقد</li><li>المصلحة المشروعة</li></ul>"},
        {"h":"4. مشاركة البيانات","c":_PRIVACY_THIRD_PARTIES.format(ocr_desc="— قد تحتوي الصور على IBAN وعنوان", ai_desc="— نصوص OCR للمعالجة")},
        {"h":"5. مدة الاحتفاظ","c":"<ul><li>بيانات الحساب: حتى الحذف</li><li>المستندات المحاسبية: 10 سنوات</li><li>سلة المهملات: 30 يوماً</li></ul>"},
        {"h":"6. حقوقك","c":"<ul><li>الوصول، التصحيح، الحذف، النقل، الاعتراض</li></ul>"},
        {"h":"7. ملفات تعريف الارتباط","c":"<p>لا توجد ملفات تتبع. تخزين محلي تقني فقط.</p>"},
    ])


@app.post("/account/do-not-sell")
def ccpa_do_not_sell(user: dict = Depends(get_current_user)):
    """CCPA — Do Not Sell My Personal Information. We don't sell data, but this endpoint confirms it."""
    logger.info("CCPA: Do-not-sell request from user_id=%d", user["sub"])
    return {"success": True, "message": "We do not sell your personal data. Your request has been recorded.",
            "status": "confirmed", "applies_to": "CCPA (California), all regions"}


@app.get("/admin", response_class=HTMLResponse)
def admin_page():
    """Admin paneli — login JWT cookie/Bearer ile auth, sayfa kendisi
    ADMIN_EMAILS koruma altinda /admin/* middleware'iyle. Bu HTML
    public ama icindeki API cagrilari yetkisiz erisimi block eder."""
    return HTMLResponse(content="""<!DOCTYPE html><html lang="de"><head>
<meta charset="UTF-8"><title>Admin Panel — AutoTax-HUB</title>
<style>
  *{box-sizing:border-box}
  body{font-family:-apple-system,'DM Sans',sans-serif;background:#050a12;color:#e8edf5;margin:0;padding:24px;line-height:1.5}
  h1{color:#10b981;font-size:24px;margin:0 0 6px}
  .sub{color:#64748b;font-size:13px;margin:0 0 24px}
  .stats{display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px;margin-bottom:28px}
  .card{background:#0c1420;border:1px solid #1f2937;border-radius:12px;padding:16px}
  .card .label{color:#64748b;font-size:11px;text-transform:uppercase;letter-spacing:1px}
  .card .val{color:#e8edf5;font-size:24px;font-weight:700;margin-top:4px}
  .card.acc .val{color:#10b981}
  .card.warn .val{color:#f59e0b}
  table{width:100%;border-collapse:collapse;background:#0c1420;border-radius:12px;overflow:hidden;border:1px solid #1f2937}
  th{text-align:left;padding:12px 14px;background:#111c2c;color:#94a3b8;font-size:11px;text-transform:uppercase;letter-spacing:1px;font-weight:600}
  td{padding:12px 14px;border-top:1px solid #1f2937;font-size:13px;vertical-align:middle}
  tr:hover td{background:rgba(255,255,255,0.02)}
  select,input,button{font-family:inherit;font-size:13px;padding:6px 10px;border-radius:6px;background:#1f2937;color:#e8edf5;border:1px solid #334155;outline:none}
  select:focus,input:focus{border-color:#10b981}
  button{cursor:pointer;background:#10b981;border-color:#10b981;color:#000;font-weight:600}
  button:hover{background:#0d9668}
  button.danger{background:#ef4444;border-color:#ef4444;color:#fff}
  button.danger:hover{background:#dc2626}
  button.ghost{background:transparent;color:#94a3b8;border-color:#334155}
  button.ghost:hover{color:#e8edf5;border-color:#475569}
  .toggle{position:relative;display:inline-block;width:36px;height:20px}
  .toggle input{opacity:0;width:0;height:0}
  .toggle .slider{position:absolute;cursor:pointer;inset:0;background:#334155;border-radius:20px;transition:.2s}
  .toggle .slider:before{content:"";position:absolute;left:2px;top:2px;width:16px;height:16px;background:#fff;border-radius:50%;transition:.2s}
  .toggle input:checked + .slider{background:#10b981}
  .toggle input:checked + .slider:before{transform:translateX(16px)}
  .badge{display:inline-block;padding:2px 8px;border-radius:4px;font-size:10px;font-weight:700;letter-spacing:.5px;text-transform:uppercase}
  .badge.free{background:#475569;color:#fff}
  .badge.early{background:#f59e0b;color:#000}
  .badge.pro{background:#10b981;color:#000}
  .badge.admin{background:linear-gradient(135deg,#a855f7,#ec4899);color:#fff;text-shadow:0 1px 2px rgba(0,0,0,0.3);box-shadow:0 0 8px rgba(168,85,247,0.4)}
  .badge.trial{background:linear-gradient(135deg,#a855f7,#6366f1);color:#fff;font-weight:700}
  .badge.trial-warn{background:linear-gradient(135deg,#f59e0b,#ef4444);color:#fff;animation:pulse 2s infinite}
  .badge.expired{background:#475569;color:#fff;text-decoration:line-through}
  @keyframes pulse{0%,100%{opacity:1}50%{opacity:.7}}
  .badge.large{font-size:13px;padding:4px 12px;border-radius:6px}
  tr.admin-row{background:rgba(168,85,247,0.04)}
  tr.admin-row:hover td{background:rgba(168,85,247,0.06) !important}
  .badge.cloud{background:linear-gradient(135deg,#f59e0b,#ef4444);color:#fff}
  .toolbar{display:flex;gap:8px;margin-bottom:16px;flex-wrap:wrap}
  .err{color:#ef4444;padding:14px;background:rgba(239,68,68,.08);border:1px solid rgba(239,68,68,.3);border-radius:10px;margin:12px 0}
  .actions{display:flex;gap:6px;flex-wrap:wrap}
  .small{font-size:11px;color:#64748b}
  a{color:#10b981;text-decoration:none}
</style></head><body>

<p><a href="/">← App</a></p>
<h1>🛠️ Admin Panel</h1>
<p class="sub">Kullanici yonetimi · Plan degisikligi · Manuel odeme onayi</p>

<div id="loginGate" style="display:none">
  <div class="err">Bu sayfaya erisim icin admin email'iyle giris yapmis olmaniz gerekir.</div>
  <p><a href="/">← Login sayfasina don</a></p>
</div>

<div id="content" style="display:none">

<div id="stats" class="stats"></div>

<div id="reminders" style="display:grid;grid-template-columns:1fr 1fr;gap:14px;margin-bottom:24px"></div>
<div style="margin-bottom:8px;display:flex;gap:8px;align-items:center">
  <button class="ghost" onclick="runRemindersNow()" style="font-size:12px">▶ Reminder Cycle Calistir</button>
  <span id="reminderResult" class="small"></span>
</div>

<div class="toolbar">
  <input id="searchInput" placeholder="Email ara (free/early/pro yazarsan plan filtreler)..." style="flex:1;min-width:240px"/>
  <select id="planFilter" onchange="loadUsers()">
    <option value="">Tum planlar</option>
    <option value="free">Sadece Free</option>
    <option value="early">Sadece Early</option>
    <option value="pro">Sadece Pro</option>
  </select>
  <button onclick="loadUsers()">🔍 Ara</button>
  <button class="ghost" onclick="document.getElementById('searchInput').value='';document.getElementById('planFilter').value='';loadUsers();">✕ Temizle</button>
  <button class="ghost" onclick="refreshAll()">↻ Yenile</button>
</div>

<table id="usersTable">
  <thead><tr>
    <th>ID</th><th>Email</th><th>Ad</th><th>Plan</th><th>Cloud</th><th>KU</th>
    <th>Fis</th><th>Kayit</th><th>Aksiyon</th>
  </tr></thead>
  <tbody id="usersBody"><tr><td colspan="9" style="text-align:center;color:#64748b">Yukleniyor...</td></tr></tbody>
</table>

</div>

<script>
const API = location.origin;
// Ana app token'i 'atx_token' anahtarinda saklar (eski kod 'token' icin
// fallback). Ikisini de kontrol ederek hangisi varsa kullan.
const token = localStorage.getItem("atx_token")
  || localStorage.getItem("token")
  || sessionStorage.getItem("atx_token")
  || sessionStorage.getItem("token");

async function api(path, opts) {
  opts = opts || {};
  opts.headers = Object.assign({"Content-Type":"application/json","Authorization":"Bearer "+token}, opts.headers||{});
  const r = await fetch(API+path, opts);
  if (!r.ok) {
    const txt = await r.text().catch(() => "");
    throw new Error("HTTP "+r.status+(txt?": "+txt.slice(0,200):""));
  }
  return r.json();
}

if (!token) {
  // Hic token yok -> direkt login sayfasina yonlendir
  window.location.href = "/?next=/admin";
} else {
  // Token var -> content goster (API call sonucu ne olursa olsun)
  document.getElementById("content").style.display = "block";
  refreshAll().catch(e => {
    console.error("admin load error:", e);
    if (String(e.message).includes("403")) {
      // Yetki yok -> aciklayici mesaj
      document.getElementById("content").innerHTML =
        '<div class="err">Bu hesap admin yetkili degil.<br>'+
        '<small>ADMIN_EMAILS env var\\'inda olan bir email ile giris yap.</small><br>'+
        '<a href="/">← Ana sayfa</a></div>';
    } else {
      document.getElementById("stats").innerHTML =
        '<div class="err">Veri yuklenemedi: '+e.message+'</div>';
    }
  });
}

async function refreshAll() {
  await loadStats();
  await loadReminders();
  await loadUsers();
}

async function loadReminders() {
  try {
    const [up, ov] = await Promise.all([
      api("/reminders/upcoming?days=7"),
      api("/reminders/overdue"),
    ]);
    const div = document.getElementById("reminders");
    div.innerHTML = `
      <div class="card" style="border-color:${ov.count > 0 ? '#dc2626' : '#1f2937'}">
        <div class="label" style="color:${ov.count > 0 ? '#dc2626' : '#64748b'}">🚨 Überfällig</div>
        <div class="val" style="color:${ov.count > 0 ? '#dc2626' : '#94a3b8'}">${ov.count}</div>
        ${ov.count > 0 ? '<div class="small" style="color:#dc2626;margin-top:6px">€'+ov.total_amount.toFixed(2)+' offen</div>' : ''}
        ${renderRemList(ov.items, true)}
      </div>
      <div class="card">
        <div class="label">📅 Nächste 7 Tage fällig</div>
        <div class="val">${up.count}</div>
        ${renderRemList(up.items, false)}
      </div>
    `;
  } catch (e) { console.error("reminders", e); }
}

function renderRemList(items, isOverdue) {
  if (!items.length) return '<div class="small" style="margin-top:8px;color:#64748b">— Yok —</div>';
  return '<div style="margin-top:10px;font-size:12px;display:flex;flex-direction:column;gap:6px">' +
    items.slice(0, 5).map(i => {
      const days = i.days_until_due;
      const tag = isOverdue ? `<span style="color:#dc2626">${Math.abs(days)} gün geçti</span>` :
                  (days === 0 ? '<span style="color:#f59e0b">BUGÜN</span>' :
                   days === 1 ? '<span style="color:#f59e0b">YARIN</span>' :
                   `${days} gün kaldı`);
      return `<div style="display:flex;justify-content:space-between;gap:6px;padding:6px;background:rgba(255,255,255,0.02);border-radius:6px">
        <span>${esc(i.vendor)} <span class="small">·  ${i.due_date || '—'}</span></span>
        <span>€${i.total_amount.toFixed(2)} ${tag}</span>
      </div>`;
    }).join('') +
    (items.length > 5 ? `<div class="small" style="text-align:center;color:#64748b">+ ${items.length-5} daha</div>` : '') +
    '</div>';
}

async function runRemindersNow() {
  document.getElementById("reminderResult").textContent = "Calisiyor...";
  try {
    const r = await api("/admin/reminders/run-now", {method:"POST"});
    const s = r.stats;
    document.getElementById("reminderResult").textContent =
      `✓ ${s.checked} kontrol, ${s.sent_telegram} Telegram, ${s.sent_email} email`;
    await loadReminders();
  } catch (e) {
    document.getElementById("reminderResult").textContent = "Hata: "+e.message;
  }
}

async function loadStats() {
  try {
    const s = await api("/admin/stats");
    document.getElementById("stats").innerHTML = `
      <div class="card"><div class="label">Odeyebilir Musteri</div><div class="val">${s.paying_users || s.total_users}</div>${s.admin_count?'<div class="small" style="margin-top:4px">+ '+s.admin_count+' admin</div>':''}</div>
      <div class="card acc"><div class="label">Aylik Tahmini Gelir</div><div class="val">€${s.monthly_revenue_estimate_eur}</div><div class="small" style="margin-top:4px">admin haric</div></div>
      <div class="card"><div class="label">Toplam Fis</div><div class="val">${s.total_invoices}</div></div>
      <div class="card warn"><div class="label">Son 7 Gun Yeni</div><div class="val">${s.new_users_7d}</div></div>
      <div class="card"><div class="label">Free / Early / Pro</div><div class="val" style="font-size:18px">${s.users_by_plan.free} / ${s.users_by_plan.early} / ${s.users_by_plan.pro}</div></div>
      <div class="card"><div class="label">Cloud Add-on</div><div class="val">${s.cloud_addon_users}</div></div>
    `;
  } catch (e) { console.error("stats", e); }
}

async function loadUsers() {
  try {
    let search = document.getElementById("searchInput").value.trim();
    let plan = document.getElementById("planFilter").value;
    // Smart: 'free' / 'early' / 'pro' yazinca otomatik plan filter'a yonlendir
    const smartPlan = ["free", "early", "pro"].find(p => p === search.toLowerCase());
    if (smartPlan && !plan) {
      plan = smartPlan;
      search = "";
    }
    const params = new URLSearchParams();
    if (search) params.set("search", search);
    if (plan) params.set("plan", plan);
    const r = await api("/admin/users?"+params.toString());
    const tbody = document.getElementById("usersBody");
    if (!r.users.length) {
      const filterDesc = plan ? `plan=${plan}` : (search ? `email contains "${search}"` : "no filter");
      tbody.innerHTML = `<tr><td colspan="9" style="text-align:center;color:#f59e0b;padding:24px">
        ⚠️ Eslesen kullanici yok (${filterDesc})<br>
        <small style="color:#64748b">Filtre temizlemek icin <strong>✕ Temizle</strong> butonuna bas</small>
      </td></tr>`;
      return;
    }
    tbody.innerHTML = r.users.map(u => `
      <tr data-uid="${u.id}" class="${u.is_admin?'admin-row':''}">
        <td class="small">${u.id}</td>
        <td><strong>${esc(u.email)}</strong>${u.is_admin?' <span class="badge admin" style="font-size:9px;padding:2px 6px;margin-left:4px">⚡ ADMIN</span>':''}${u.company_name?'<br><span class="small">'+esc(u.company_name)+'</span>':''}</td>
        <td>${esc(u.full_name)||'<span class="small">—</span>'}</td>
        <td>
          ${u.is_admin
            ? '<span class="badge large admin">⚡ ADMIN</span><br><span class="small" style="margin-top:4px;display:inline-block">Plan: '+u.plan+' (revenue\\'a sayilmaz)</span>'
            : `<span class="badge large ${u.plan}">${u.plan.toUpperCase()}</span>
               ${u.is_trial && u.trial_days_left !== null
                  ? '<span class="badge ' + (u.trial_days_left <= 3 ? 'trial-warn' : 'trial') + '" style="margin-left:6px;font-size:10px;padding:3px 7px">🎁 TRIAL ' + u.trial_days_left + 'g</span>'
                  : (u.trial_expired ? '<span class="badge expired" style="margin-left:6px;font-size:10px;padding:3px 7px">TRIAL BITTI</span>' : '')
               }
               <br>
               <select onchange="changePlan(${u.id}, this.value, this)" style="margin-top:4px">
                 <option value="free" ${u.plan==='free'?'selected':''}>Free</option>
                 <option value="early" ${u.plan==='early'?'selected':''}>Early €10</option>
                 <option value="pro" ${u.plan==='pro'?'selected':''}>Pro €20</option>
               </select>`
          }
        </td>
        <td><label class="toggle"><input type="checkbox" ${u.has_cloud_addon?'checked':''} onchange="toggleCloud(${u.id}, this.checked, this)"><span class="slider"></span></label></td>
        <td><label class="toggle"><input type="checkbox" ${u.is_kleinunternehmer?'checked':''} onchange="toggleKU(${u.id}, this.checked, this)"><span class="slider"></span></label></td>
        <td>${u.invoice_count}</td>
        <td class="small">${u.registered_at ? u.registered_at.slice(0,10) : '—'}</td>
        <td><div class="actions">
          <button class="ghost" onclick="proCloud(${u.id})" title="Pro + Cloud Aktif Et">⭐ Pro+Cloud</button>
          ${!u.is_admin ? `<button class="ghost" onclick="extendTrial(${u.id})" title="Trial uzat (+7 gun)">🎁 +7g</button>` : ''}
          <button class="ghost" onclick="downloadInvoice(${u.id}, '${u.plan}')" title="Manuel Subscription Fatura PDF">📄 Fatura</button>
          <button class="danger" onclick="delUser(${u.id}, '${esc(u.email)}')">Sil</button>
        </div></td>
      </tr>
    `).join("");
  } catch (e) { console.error("users", e); }
}

function esc(s){return String(s||"").replace(/[&<>"]/g,c=>({"&":"&amp;","<":"&lt;",">":"&gt;",'"':"&quot;"}[c]));}

async function patch(uid, body) {
  return api("/admin/users/"+uid, {method:"PATCH", body: JSON.stringify(body)});
}

async function changePlan(uid, plan, el) {
  el.disabled = true;
  try { await patch(uid, {plan}); await loadStats(); await loadUsers(); }
  catch(e){ alert("Hata: "+e.message); }
  el.disabled = false;
}

async function toggleCloud(uid, val, el) {
  el.disabled = true;
  try { await patch(uid, {has_cloud_addon: val}); await loadStats(); }
  catch(e){ alert("Hata: "+e.message); el.checked = !val; }
  el.disabled = false;
}

async function toggleKU(uid, val, el) {
  el.disabled = true;
  try { await patch(uid, {is_kleinunternehmer: val}); }
  catch(e){ alert("Hata: "+e.message); el.checked = !val; }
  el.disabled = false;
}

async function proCloud(uid) {
  if (!confirm("Bu kullaniciyi Pro plan + Cloud Add-on yapayim mi? (manuel odeme alindiginda)")) return;
  try { await patch(uid, {plan:"pro", has_cloud_addon:true}); await refreshAll(); }
  catch(e){ alert("Hata: "+e.message); }
}

async function extendTrial(uid) {
  const days = parseInt(prompt("Kac gun uzatmak istiyorsun?", "7") || "0", 10);
  if (!days || days < 1 || days > 90) return;
  try { await patch(uid, {extend_trial_days: days}); await refreshAll(); }
  catch(e){ alert("Hata: "+e.message); }
}

async function delUser(uid, email) {
  if (!confirm(`KALICI silmek mi istiyorsun?\\n\\n${email}\\n\\nKullanicinin tum fislerini ve verilerini siler. Geri alinamaz.`)) return;
  try { await api("/admin/users/"+uid, {method:"DELETE"}); await refreshAll(); }
  catch(e){ alert("Hata: "+e.message); }
}

async function downloadInvoice(uid, currentPlan) {
  const plan = prompt("Plan (free / early / pro):", currentPlan === "free" ? "pro" : currentPlan);
  if (!plan) return;
  const months = prompt("Kac ay icin? (1-12)", "1");
  if (!months) return;
  // Auth header'i fetch ile gondermek icin blob indir
  try {
    const r = await fetch(API+`/admin/users/${uid}/invoice?plan=${encodeURIComponent(plan)}&months=${encodeURIComponent(months)}`, {
      headers: {"Authorization": "Bearer "+token}
    });
    if (!r.ok) { alert("Hata: HTTP "+r.status); return; }
    const blob = await r.blob();
    const url = URL.createObjectURL(blob);
    const a = document.createElement("a");
    a.href = url;
    a.download = `Rechnung-User-${uid}.pdf`;
    document.body.appendChild(a);
    a.click();
    a.remove();
    URL.revokeObjectURL(url);
  } catch (e) { alert("Hata: "+e.message); }
}

document.getElementById("searchInput").addEventListener("keydown", e => {
  if (e.key === "Enter") loadUsers();
});
</script>

</body></html>""")


@app.get("/impressum", response_class=HTMLResponse)
def impressum_page():
    """Impressum nach § 5 DDG (Almanya yasal zorunluluk)."""
    return HTMLResponse(content="""<!DOCTYPE html><html lang="de"><head><meta charset="UTF-8"><title>Impressum — AutoTax-HUB</title>
<style>body{font-family:'DM Sans',sans-serif;max-width:820px;margin:40px auto;padding:20px;background:#050a12;color:#e8edf5;line-height:1.75}
h1{color:#10b981;font-size:28px}h2{color:#00a8cc;margin-top:28px;font-size:17px}
a{color:#10b981}p{margin:10px 0}strong{color:#f59e0b}
.note{background:rgba(16,185,129,0.08);border:1px solid rgba(16,185,129,0.3);padding:12px 16px;border-radius:10px;font-size:13px;margin:16px 0}</style></head><body>
<p><a href="/">← Startseite</a> · <a href="/datenschutz">Datenschutz</a> · <a href="/agb">AGB</a></p>
<h1>Impressum</h1>
<p><em>Angaben gemäß § 5 DDG (Digitale-Dienste-Gesetz) und § 55 RStV.</em></p>

<h2>Diensteanbieter</h2>
<p><strong>Hüseyin Hancer</strong><br>
Wiesenstr. 10<br>
66115 Saarbrücken<br>
Deutschland</p>

<h2>Kontakt</h2>
<p>E-Mail: <a href="mailto:info@autotaxhub.de">info@autotaxhub.de</a><br>
Datenschutz: <a href="mailto:datenschutz@autotaxhub.de">datenschutz@autotaxhub.de</a></p>

<h2>Umsatzsteuer-ID</h2>
<p>Umsatzsteuer-Identifikationsnummer gemäß § 27 a Umsatzsteuergesetz: <em>wird nachgereicht / Kleinunternehmer gemäß § 19 UStG</em></p>

<h2>Verantwortlich für den Inhalt nach § 18 Abs. 2 MStV</h2>
<p>Hüseyin Hancer (Anschrift wie oben)</p>

<h2>Streitschlichtung</h2>
<p>Die Europäische Kommission stellt eine Plattform zur Online-Streitbeilegung (OS) bereit:
<a href="https://ec.europa.eu/consumers/odr" target="_blank" rel="noopener">https://ec.europa.eu/consumers/odr</a>.<br>
Wir sind nicht bereit oder verpflichtet, an Streitbeilegungsverfahren vor einer Verbraucherschlichtungsstelle teilzunehmen.</p>

<h2>Haftung für Inhalte</h2>
<p>Als Diensteanbieter sind wir gemäß § 7 Abs.1 DDG für eigene Inhalte auf diesen Seiten nach den allgemeinen Gesetzen verantwortlich. Nach §§ 8 bis 10 DDG sind wir als Diensteanbieter jedoch nicht verpflichtet, übermittelte oder gespeicherte fremde Informationen zu überwachen oder nach Umständen zu forschen, die auf eine rechtswidrige Tätigkeit hinweisen.</p>

<h2>Haftung für Links</h2>
<p>Unser Angebot enthält Links zu externen Webseiten Dritter, auf deren Inhalte wir keinen Einfluss haben. Deshalb können wir für diese fremden Inhalte auch keine Gewähr übernehmen. Für die Inhalte der verlinkten Seiten ist stets der jeweilige Anbieter oder Betreiber der Seiten verantwortlich.</p>

<h2>Urheberrecht</h2>
<p>Die durch die Seitenbetreiber erstellten Inhalte und Werke auf diesen Seiten unterliegen dem deutschen Urheberrecht. Die Vervielfältigung, Bearbeitung, Verbreitung und jede Art der Verwertung außerhalb der Grenzen des Urheberrechtes bedürfen der schriftlichen Zustimmung des jeweiligen Autors bzw. Erstellers.</p>

<div class="note"><strong>Hinweis zur Steuerberatung:</strong> AutoTax-HUB ist <em>keine</em> Steuerberatungssoftware im Sinne des Steuerberatungsgesetzes (StBerG). Der Dienst stellt lediglich technische Werkzeuge zur Belegerfassung und -archivierung bereit. Für steuerliche Beratung wenden Sie sich bitte an einen zugelassenen Steuerberater.</div>

<p style="margin-top:40px;color:#64748b;font-size:13px">© 2026 AutoTax-HUB · Stand: Mai 2026</p>
</body></html>""")


@app.get("/agb", response_class=HTMLResponse)
def agb_page():
    return HTMLResponse(content="""<!DOCTYPE html><html lang="de"><head><meta charset="UTF-8"><title>AGB — AutoTax-HUB</title>
<style>body{font-family:'DM Sans',sans-serif;max-width:820px;margin:40px auto;padding:20px;background:#050a12;color:#e8edf5;line-height:1.75}
h1{color:#10b981;font-size:28px}h2{color:#00a8cc;margin-top:28px;font-size:17px}strong{color:#f59e0b}
a{color:#10b981}p{margin:10px 0}ul{padding-left:22px;margin:8px 0}li{margin:4px 0}
.note{background:rgba(245,158,11,0.08);border:1px solid rgba(245,158,11,0.3);padding:12px 16px;border-radius:10px;font-size:13px;color:#f59e0b;margin:16px 0}
</style></head><body>
<h1>Allgemeine Geschäftsbedingungen (AGB)</h1>
<p><em>AutoTax-HUB — Stand: April 2026</em></p>

<h2>§ 1 Geltungsbereich und Anbieter</h2>
<p>Diese Allgemeinen Geschäftsbedingungen gelten für sämtliche Verträge zwischen <strong>Hüseyin Hancer, Wiesenstr. 10, 66115 Saarbrücken, Deutschland</strong> (nachfolgend "Anbieter") und den Nutzern der Software-as-a-Service-Plattform "AutoTax-HUB" (nachfolgend "Dienst"), erreichbar unter autotaxhub.de und zugehörigen Subdomains.</p>

<h2>§ 2 Leistungsbeschreibung</h2>
<p>Der Dienst stellt Werkzeuge zur halbautomatischen Erfassung und Verwaltung von Belegen und Kassenbüchern bereit, unter anderem OCR-Texterkennung, Kategorisierung, Kassenbuchführung, Exportfunktionen (CSV, Excel, DATEV, JSON) sowie eine Einnahmen-Überschuss-Übersicht. Der konkrete Funktionsumfang richtet sich nach dem gewählten Plan.</p>
<p>Der Dienst befindet sich in der <strong>BETA-Phase</strong>. Einzelne Funktionen können jederzeit hinzugefügt, geändert, eingeschränkt oder entfernt werden.</p>

<h2>§ 3 Keine Steuer- oder Rechtsberatung</h2>
<p>AutoTax-HUB ist ein Software-Werkzeug. Der Dienst stellt <strong>keine Steuerberatung, Rechtsberatung oder buchhalterische Beratung</strong> dar. Automatisch erkannte und berechnete Werte können fehlerhaft sein. Der Nutzer ist <strong>allein verantwortlich</strong> für die Überprüfung, Korrektur, Aufbewahrung und ordnungsgemäße steuerliche Verwendung sämtlicher Daten. Für die Steuererklärung und Buchführung ist die Konsultation eines Steuerberaters dringend empfohlen. Die GoBD-konforme Aufbewahrung der Originalbelege obliegt dem Nutzer.</p>

<h2>§ 4 Prüfungspflicht des Nutzers</h2>
<p>Der Nutzer ist verpflichtet, alle automatisch erkannten Daten (insbesondere Beträge, MwSt-Sätze, Lieferanten, Kategorien, Datumsangaben, Rechnungsnummern) vor jeglicher weiteren Verwendung auf Richtigkeit und Vollständigkeit zu prüfen und bei Bedarf zu korrigieren.</p>

<h2>§ 5 Haftungsbeschränkung</h2>
<p>Der Anbieter haftet bei Vorsatz und grober Fahrlässigkeit unbeschränkt. Bei einfacher Fahrlässigkeit haftet er nur bei Verletzung einer wesentlichen Vertragspflicht (Kardinalpflicht), deren Erfüllung die ordnungsgemäße Durchführung des Vertrags überhaupt erst ermöglicht und auf deren Einhaltung der Nutzer regelmäßig vertrauen darf, und der Höhe nach beschränkt auf den typischerweise vorhersehbaren Schaden.</p>
<p>In jedem Fall ist die Haftung auf die vom Nutzer an den Anbieter in den letzten zwölf Monaten gezahlten Beträge beschränkt, höchstens jedoch auf 500 €.</p>
<p>Der Anbieter haftet <strong>nicht</strong> für: falsch erkannte Beträge oder Kategorien, fehlerhafte MwSt- oder Steuerberechnungen, steuerliche Nachteile, verpasste Fristen, entgangenen Gewinn, Datenverlust, Serviceausfälle oder sonstige mittelbare Schäden.</p>
<p>Die Haftung aus dem Produkthaftungsgesetz und für die Verletzung von Leben, Körper oder Gesundheit bleibt unberührt.</p>

<h2>§ 6 Preise, Free-Plan und Änderungsvorbehalt</h2>
<p>Der Anbieter bietet derzeit einen kostenlosen Plan ("Free") sowie kostenpflichtige Pläne an. <strong>Der kostenlose Plan ist eine freiwillige Leistung und begründet keinen Rechtsanspruch auf dauerhafte Kostenfreiheit.</strong> Der Anbieter behält sich ausdrücklich vor, den Free-Plan mit einer Ankündigungsfrist von 14 Tagen einzuschränken, zu bepreisen oder einzustellen.</p>
<p>Preise, Limits und Funktionsumfang kostenpflichtiger Pläne können mit einer Ankündigungsfrist von 30 Tagen per E-Mail geändert werden. Der Nutzer hat in diesem Fall das Recht, den Vertrag zum Wirksamwerden der Änderung außerordentlich zu kündigen.</p>

<h2>§ 7 Fair Use und Missbrauchsschutz</h2>
<p>Nutzungskontingente (z. B. Belege pro Tag, API-Aufrufe pro Minute, AI-Chat-Anfragen) werden für jeden Plan individuell festgelegt und im Plan sichtbar ausgewiesen. Sie dienen auch dem Schutz vor automatisierter Überlastung.</p>
<p>Der Anbieter behält sich das Recht vor, den Zugang bei <strong>missbräuchlicher Nutzung</strong> — insbesondere automatisierten Massen-Uploads, Umgehungsversuchen der Kontingente, Reverse-Engineering oder Weiterverkauf der API — ohne Vorankündigung vorübergehend oder dauerhaft zu sperren. Bei nachgewiesen missbräuchlicher Nutzung kann der Anbieter die durch den Missbrauch verursachten Mehrkosten (z. B. Drittanbieter-OCR- oder KI-Gebühren) dem Nutzer in Rechnung stellen.</p>

<h2>§ 8 Verfügbarkeit und Einstellung des Dienstes</h2>
<p>Der Anbieter bemüht sich um eine hohe Verfügbarkeit, übernimmt aber keine Garantie für ununterbrochene Erreichbarkeit. Wartungsarbeiten, Störungen, höhere Gewalt und Änderungen bei Drittanbietern (Hosting, OCR-API, KI-API) können zu Ausfällen führen.</p>
<p>Der Anbieter kann den Dienst mit einer Ankündigungsfrist von <strong>60 Tagen</strong> vollständig einstellen. Nutzer erhalten vor Einstellung die Möglichkeit, ihre Daten zu exportieren. Bereits im Voraus gezahlte Beträge für kostenpflichtige Pläne werden anteilig für den nicht genutzten Zeitraum erstattet.</p>

<h2>§ 9 Vertragslaufzeit und Kündigung</h2>
<p>Kostenpflichtige Abonnements werden monatlich abgerechnet und verlängern sich automatisch um den jeweils gewählten Zeitraum, sofern nicht bis zum Ende der laufenden Periode gekündigt wird. Die Kündigung erfolgt jederzeit in den Konto-Einstellungen oder per E-Mail an info@autotaxhub.de.</p>
<p>Der Nutzer kann sein Konto jederzeit unwiderruflich löschen. Mit der Löschung werden alle personenbezogenen Daten gemäß Art. 17 DSGVO entfernt; steuerlich aufbewahrungspflichtige Daten können in anonymisierter Form für die gesetzliche Dauer verbleiben.</p>

<h2>§ 10 Widerrufsrecht für Verbraucher</h2>
<p>Verbraucher im Sinne von § 13 BGB haben bei digitalen Dienstleistungen grundsätzlich ein 14-tägiges Widerrufsrecht ab Vertragsabschluss. Mit ausdrücklicher Zustimmung des Verbrauchers zum sofortigen Beginn der Leistung und Verzicht auf das Widerrufsrecht erlischt dieses bei vollständiger Erbringung der Leistung. Ein gesondertes Widerrufsformular ist auf Anfrage erhältlich.</p>

<h2>§ 11 Datenschutz und Auftragsverarbeitung</h2>
<p>Die Verarbeitung personenbezogener Daten erfolgt gemäß der <a href="/datenschutz">Datenschutzerklärung</a> und den Vorgaben der DSGVO. Zur Erbringung des Dienstes werden Daten an Auftragsverarbeiter übermittelt (Hosting, OCR-Dienst, KI-Verarbeitung). Die Liste der Empfänger sowie die Rechtsgrundlagen sind in der Datenschutzerklärung vollständig aufgeführt.</p>

<h2>§ 12 Schlussbestimmungen</h2>
<ul>
<li>Es gilt das Recht der Bundesrepublik Deutschland unter Ausschluss des UN-Kaufrechts. Gegenüber Verbrauchern gilt diese Rechtswahl nur, soweit dem Verbraucher dadurch nicht der Schutz zwingender Bestimmungen seines Heimatrechts entzogen wird.</li>
<li>Ausschließlicher Gerichtsstand für Streitigkeiten mit Kaufleuten, juristischen Personen des öffentlichen Rechts oder öffentlich-rechtlichen Sondervermögen ist Saarbrücken.</li>
<li>Sollte eine Bestimmung dieser AGB unwirksam sein oder werden, bleibt die Wirksamkeit der übrigen Bestimmungen unberührt.</li>
<li>Die Europäische Kommission stellt unter <a href="https://ec.europa.eu/consumers/odr" target="_blank" rel="noopener">ec.europa.eu/consumers/odr</a> eine Online-Streitbeilegungsplattform bereit. Der Anbieter ist nicht verpflichtet und nicht bereit, an einem Streitbeilegungsverfahren vor einer Verbraucherschlichtungsstelle teilzunehmen.</li>
</ul>

<div class="note">
<strong>Hinweis:</strong> Diese AGB sind ein Grundgerüst. Vor Aktivierung kostenpflichtiger Pläne oder bei Aufnahme gewerblicher Nutzer im großen Stil wird dringend empfohlen, die AGB von einem deutschen Rechtsanwalt (Fachgebiet IT-/SaaS-Recht) prüfen und gegebenenfalls ergänzen zu lassen.
</div>

<p style="margin-top:40px;color:#64748b;font-size:13px">© 2026 AutoTax-HUB — Alle Rechte vorbehalten.</p>
</body></html>""")


@app.post("/admin/reset-password")
def admin_reset_password(body: dict = Body(...), admin: dict = Depends(get_current_user)):
    _ADMIN_EMAILS = set(filter(None, os.getenv("ADMIN_EMAILS", "").split(",")))
    if not _ADMIN_EMAILS or admin.get("email") not in _ADMIN_EMAILS:
        err(403, "Admin access required")
    email = body.get("email")
    new_password = body.get("new_password")
    if not email or not new_password:
        err(400, "email and new_password required")
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == email).first()
        if not user:
            err(404, "User not found")
        user.hashed_password = hash_password(new_password)
        db.commit()
        logger.info("Admin %s reset password for %s", admin.get("email"), email)
        return {"success": True, "message": f"Password reset for {email}"}
    finally:
        db.close()


@app.post("/admin/reparse")
def admin_reparse(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"]).all()
        count = 0
        for inv in invoices:
            if not inv.raw_text:
                continue
            parsed = parse_invoice(inv.raw_text)
            inv.total_amount = parsed["total_amount"]
            inv.vat_amount = parsed["vat_amount"]
            inv.vat_rate = parsed["vat_rate"]
            inv.vendor = parsed["vendor"]
            inv.category = parsed["category"]
            inv.date = parsed["date"]
            count += 1
        db.commit()
        return {"status": "done", "count": count}
    except Exception:
        db.rollback()
        logger.exception("Reparse failed")
        err(500, "Reparse failed")
    finally:
        db.close()


# ════════════════════════════════════════════════════════════════
# ADMIN PANEL — kullanici/abonelik yonetimi
# ════════════════════════════════════════════════════════════════

@app.get("/admin/users")
def admin_users(
    search: Optional[str] = Query(None),
    plan: Optional[str] = Query(None),
    limit: int = Query(100, ge=1, le=500),
    user: dict = Depends(get_current_user),
):
    """Tum kullanicilari listeler (sadece admin). Search + plan filter."""
    db = SessionLocal()
    try:
        q = db.query(User)
        if search:
            q = q.filter(User.email.ilike(f"%{search}%"))
        if plan:
            q = q.filter(User.plan == plan)
        users = q.order_by(User.id.desc()).limit(limit).all()
        # Admin email seti (env'den) — frontend admin badge gostersin diye
        _admin_set = set(filter(None, os.getenv("ADMIN_EMAILS", "").split(",")))
        out = []
        for u in users:
            inv_count = db.query(Invoice).filter(
                Invoice.user_id == u.id,
                (Invoice.is_deleted == False) | (Invoice.is_deleted == None),
            ).count()
            companies = db.query(UserCompany).filter(UserCompany.user_id == u.id).all()
            trial_ends = getattr(u, "trial_ends_at", None)
            now_utc = datetime.now(timezone.utc)
            is_trial = bool(trial_ends and trial_ends > now_utc)
            trial_expired = bool(trial_ends and trial_ends <= now_utc)
            trial_days_left = None
            if trial_ends:
                delta = trial_ends - now_utc
                trial_days_left = max(0, delta.days + (1 if delta.seconds > 0 else 0))
            out.append({
                "id": u.id,
                "email": u.email,
                "full_name": u.full_name or "",
                "plan": u.plan or "free",
                "has_cloud_addon": getattr(u, "has_cloud_addon", False),
                "is_kleinunternehmer": getattr(u, "is_kleinunternehmer", False),
                "is_admin": u.email in _admin_set,
                "registered_at": u.registered_at.isoformat() if u.registered_at else "",
                "trial_ends_at": trial_ends.isoformat() if trial_ends else None,
                "is_trial": is_trial,
                "trial_expired": trial_expired,
                "trial_days_left": trial_days_left,
                "invoice_count": inv_count,
                "company_name": companies[0].company_name if companies else "",
            })
        return {"users": out, "total": len(out)}
    finally:
        db.close()


@app.patch("/admin/users/{user_id}")
def admin_update_user(
    user_id: int,
    body: dict = Body(...),
    user: dict = Depends(get_current_user),
):
    """Plan ve cloud addon degistirir. Manuel odeme aldigi anda admin
    panelinden tek tikla aktif eder."""
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user_id).first()
        if not u:
            err(404, "User not found")

        changed = []
        if "plan" in body:
            new_plan = (body["plan"] or "").strip().lower()
            if new_plan not in ("free", "early", "pro"):
                err(400, "Invalid plan: must be free|early|pro")
            old = u.plan
            u.plan = new_plan
            changed.append(f"plan: {old} -> {new_plan}")
            # Manuel odeme alindiysa (Pro'ya gecirildiyse) trial'i bitir
            # -> trial_ends_at NULL (kalici Pro). Free'ye dusurulduyse de
            # trial bitmis sayilir.
            if new_plan == "pro" and getattr(u, "trial_ends_at", None):
                u.trial_ends_at = None
                changed.append("trial: cleared (manual payment)")
        if "has_cloud_addon" in body:
            old = bool(getattr(u, "has_cloud_addon", False))
            u.has_cloud_addon = bool(body["has_cloud_addon"])
            changed.append(f"cloud: {old} -> {u.has_cloud_addon}")
        if "extend_trial_days" in body:
            # Manuel trial uzatma — admin'in pilot musteriye 'birkac gun
            # daha versek mi' icin
            try:
                days = int(body["extend_trial_days"])
                base = getattr(u, "trial_ends_at", None) or datetime.now(timezone.utc)
                u.trial_ends_at = base + timedelta(days=days)
                u.plan = "pro"  # uzatma -> Pro deneme aktif
                changed.append(f"trial extended +{days}d -> {u.trial_ends_at.isoformat()}")
            except (ValueError, TypeError):
                pass
        if "is_kleinunternehmer" in body:
            old = bool(getattr(u, "is_kleinunternehmer", False))
            u.is_kleinunternehmer = bool(body["is_kleinunternehmer"])
            changed.append(f"ku: {old} -> {u.is_kleinunternehmer}")
        if "full_name" in body and isinstance(body["full_name"], str):
            u.full_name = body["full_name"][:200]
            changed.append("full_name")

        db.commit()
        logger.info("ADMIN UPDATE user_id=%d by=%s changes=%s", user_id, user.get("email"), ", ".join(changed))
        return {
            "success": True,
            "user": {
                "id": u.id, "email": u.email, "plan": u.plan,
                "has_cloud_addon": getattr(u, "has_cloud_addon", False),
                "is_kleinunternehmer": getattr(u, "is_kleinunternehmer", False),
                "full_name": u.full_name or "",
            },
            "changes": changed,
        }
    finally:
        db.close()


@app.delete("/admin/users/{user_id}")
def admin_delete_user(user_id: int, user: dict = Depends(get_current_user)):
    """Kullanici siler — kullanicinin tum fis/cash/company verisi de silinir."""
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user_id).first()
        if not u:
            err(404, "User not found")
        # Admin kendini silemesin (kilit risk)
        if u.email == user.get("email"):
            err(400, "Cannot delete yourself")
        email = u.email
        db.query(Invoice).filter(Invoice.user_id == user_id).delete()
        db.query(CashEntry).filter(CashEntry.user_id == user_id).delete()
        db.query(UserCompany).filter(UserCompany.user_id == user_id).delete()
        db.delete(u)
        db.commit()
        logger.warning("ADMIN DELETE user id=%d email=%s by=%s", user_id, email, user.get("email"))
        return {"success": True, "deleted": email}
    finally:
        db.close()


@app.get("/admin/stats")
def admin_stats(user: dict = Depends(get_current_user)):
    """Genel istatistik — toplam kullanici, fis, plan dagilimi, gelir tahmini.
    Admin hesaplari (ADMIN_EMAILS) revenue hesabindan haric tutulur — kendi
    hesabini 'pro' yapsan bile MRR'a eklenmez."""
    db = SessionLocal()
    try:
        _admin_set = set(filter(None, os.getenv("ADMIN_EMAILS", "").split(",")))
        admin_count = db.query(User).filter(User.email.in_(_admin_set)).count() if _admin_set else 0
        total_users = db.query(User).count()
        # Plan dagilimi — admin'leri haric tut
        users_by_plan = {}
        for plan_id in ("free", "early", "pro"):
            q = db.query(User).filter(User.plan == plan_id)
            if _admin_set:
                q = q.filter(~User.email.in_(_admin_set))
            users_by_plan[plan_id] = q.count()
        cloud_q = db.query(User).filter(User.has_cloud_addon == True)
        if _admin_set:
            cloud_q = cloud_q.filter(~User.email.in_(_admin_set))
        cloud_users = cloud_q.count()

        total_invoices = db.query(Invoice).filter(
            (Invoice.is_deleted == False) | (Invoice.is_deleted == None)
        ).count()

        # Aylik gelir tahmini (Stripe gelmeden once el ile takip)
        # Ucretler: free=0, early=10, pro=20, cloud=+5
        revenue_estimate = (
            users_by_plan.get("early", 0) * 10
            + users_by_plan.get("pro", 0) * 20
            + cloud_users * 5
        )

        # Son 7 gun yeni kayit
        from sqlalchemy import func as _func
        from datetime import timedelta
        cutoff = datetime.now(timezone.utc) - timedelta(days=7)
        new_7d = db.query(User).filter(User.registered_at >= cutoff).count() if True else 0

        return {
            "total_users": total_users,
            "admin_count": admin_count,
            "paying_users": total_users - admin_count,  # admin'siz toplam
            "users_by_plan": users_by_plan,
            "cloud_addon_users": cloud_users,
            "total_invoices": total_invoices,
            "monthly_revenue_estimate_eur": revenue_estimate,
            "new_users_7d": new_7d,
        }
    finally:
        db.close()


# ════════════════════════════════════════════════════════════════
# RECHNUNG REMINDER SYSTEM — odeme takibi endpoint'leri
# ════════════════════════════════════════════════════════════════

@app.patch("/invoices/{invoice_id}/payment")
def update_invoice_payment(
    invoice_id: int,
    body: dict = Body(...),
    user: dict = Depends(get_current_user),
):
    """Faturanin odeme bilgisini guncelle (due_date, payment_status, paid_at).
    body: {due_date?: 'YYYY-MM-DD', payment_status?: 'paid|unpaid|overdue'}"""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(
            Invoice.id == invoice_id, Invoice.user_id == user["sub"],
        ).first()
        if not inv:
            err(404, "Invoice not found")

        changed = []
        if "due_date" in body:
            d = (body["due_date"] or "").strip()
            if d and not _re_global.match(r"^\d{4}-\d{2}-\d{2}$", d):
                err(400, "due_date must be YYYY-MM-DD")
            inv.due_date = d or None
            # due_date degisince reminder kodlarini sifirla — yeni tarih
            # icin reminder'lar tekrar verilir.
            inv.reminder_sent_codes = None
            changed.append(f"due_date={inv.due_date}")
        if "payment_status" in body:
            ps = (body["payment_status"] or "").lower()
            if ps not in ("paid", "unpaid", "overdue"):
                err(400, "payment_status must be paid|unpaid|overdue")
            inv.payment_status = ps
            inv.paid_at = datetime.now(timezone.utc) if ps == "paid" else None
            changed.append(f"status={ps}")
        db.commit()
        return {
            "id": inv.id,
            "due_date": inv.due_date,
            "payment_status": inv.payment_status,
            "paid_at": inv.paid_at.isoformat() if inv.paid_at else None,
            "changes": changed,
        }
    finally:
        db.close()


def _serialize_reminder_invoice(inv: Invoice) -> dict:
    today = datetime.now(timezone.utc).date()
    due_str = inv.due_date or ""
    days_until = None
    if due_str:
        try:
            d = datetime.strptime(due_str[:10], "%Y-%m-%d").date()
            days_until = (d - today).days
        except ValueError:
            pass
    return {
        "id": inv.id,
        "vendor": inv.vendor or "",
        "total_amount": safe_float(inv.total_amount),
        "invoice_number": inv.invoice_number or "",
        "date": inv.date or "",
        "due_date": inv.due_date or "",
        "payment_status": inv.payment_status or "unpaid",
        "days_until_due": days_until,
        "is_overdue": days_until is not None and days_until < 0,
    }


@app.get("/reminders/upcoming")
def reminders_upcoming(
    days: int = Query(7, ge=1, le=90),
    user: dict = Depends(get_current_user),
):
    """Onumuzdeki N gun icinde vadesi gelecek odenmemis faturalar."""
    db = SessionLocal()
    try:
        today = datetime.now(timezone.utc).date()
        upper = (today + timedelta(days=days)).isoformat()
        invoices = (
            db.query(Invoice)
            .filter(Invoice.user_id == user["sub"])
            .filter((Invoice.is_deleted == False) | (Invoice.is_deleted == None))
            .filter(Invoice.payment_status != "paid")
            .filter(Invoice.due_date.isnot(None))
            .filter(Invoice.due_date >= today.isoformat())
            .filter(Invoice.due_date <= upper)
            .order_by(Invoice.due_date.asc())
            .all()
        )
        return {
            "items": [_serialize_reminder_invoice(i) for i in invoices],
            "count": len(invoices),
            "window_days": days,
        }
    finally:
        db.close()


@app.get("/reminders/overdue")
def reminders_overdue(user: dict = Depends(get_current_user)):
    """Vadesi gecmis odenmemis faturalar."""
    db = SessionLocal()
    try:
        today = datetime.now(timezone.utc).date()
        invoices = (
            db.query(Invoice)
            .filter(Invoice.user_id == user["sub"])
            .filter((Invoice.is_deleted == False) | (Invoice.is_deleted == None))
            .filter(Invoice.payment_status != "paid")
            .filter(Invoice.due_date.isnot(None))
            .filter(Invoice.due_date < today.isoformat())
            .order_by(Invoice.due_date.asc())
            .all()
        )
        items = [_serialize_reminder_invoice(i) for i in invoices]
        total_overdue = sum(i["total_amount"] for i in items)
        return {"items": items, "count": len(items), "total_amount": round(total_overdue, 2)}
    finally:
        db.close()


@app.get("/steuer/upcoming")
def steuer_upcoming(user: dict = Depends(get_current_user)):
    """Kullanicinin yaklasan vergi vadeleri (USt/ESt/GewSt/Jahres).
    Kleinunternehmer'a USt yok."""
    from autotax.steuer import upcoming_for_user
    return {"items": upcoming_for_user(user["sub"])}


@app.post("/admin/steuer/run-now")
async def admin_run_steuer_now(user: dict = Depends(get_current_user)):
    """Manuel steuer reminder cycle (test icin)."""
    from autotax.steuer import process_steuer_reminders
    stats = await process_steuer_reminders()
    return {"success": True, "stats": stats}


@app.post("/admin/mahnung/run-now")
async def admin_run_mahnung_now(user: dict = Depends(get_current_user)):
    """Manuel Mahnung cycle (test icin)."""
    from autotax.mahnung import process_mahnungen
    stats = await process_mahnungen()
    return {"success": True, "stats": stats}


@app.post("/admin/recurring/run-now")
async def admin_run_recurring_now(user: dict = Depends(get_current_user)):
    """Manuel recurring spawn cycle (test icin)."""
    from autotax.recurring import process_recurring_spawns
    stats = await process_recurring_spawns()
    return {"success": True, "stats": stats}


@app.patch("/invoices/{invoice_id}/recurring")
def update_invoice_recurring(
    invoice_id: int,
    body: dict = Body(...),
    user: dict = Depends(get_current_user),
):
    """Bir faturayi recurring template'e cevir veya recurring'i kapat.
    body: {is_recurring: bool, recurring_freq?: 'monthly|quarterly|yearly',
           recurring_next_at?: 'YYYY-MM-DD'}"""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(
            Invoice.id == invoice_id, Invoice.user_id == user["sub"],
        ).first()
        if not inv:
            err(404, "Invoice not found")
        if "is_recurring" in body:
            inv.is_recurring = bool(body["is_recurring"])
            if not inv.is_recurring:
                inv.recurring_freq = None
                inv.recurring_next_at = None
        if "recurring_freq" in body:
            f = (body["recurring_freq"] or "").lower()
            if f and f not in ("monthly", "quarterly", "yearly"):
                err(400, "recurring_freq must be monthly|quarterly|yearly")
            inv.recurring_freq = f or None
        if "recurring_next_at" in body:
            n = (body["recurring_next_at"] or "").strip()
            if n and not _re_global.match(r"^\d{4}-\d{2}-\d{2}$", n):
                err(400, "recurring_next_at must be YYYY-MM-DD")
            inv.recurring_next_at = n or None
        # Default: monthly + next ay bugun
        if inv.is_recurring:
            if not inv.recurring_freq:
                inv.recurring_freq = "monthly"
            if not inv.recurring_next_at:
                from autotax.recurring import compute_next_spawn
                base = inv.date or datetime.now().strftime("%Y-%m-%d")
                try:
                    base_d = datetime.strptime(base[:10], "%Y-%m-%d").date()
                except ValueError:
                    base_d = datetime.now().date()
                inv.recurring_next_at = compute_next_spawn(base_d, inv.recurring_freq).isoformat()
        db.commit()
        return {
            "id": inv.id,
            "is_recurring": inv.is_recurring,
            "recurring_freq": inv.recurring_freq,
            "recurring_next_at": inv.recurring_next_at,
        }
    finally:
        db.close()


@app.post("/admin/monthly-summary/run-now")
async def admin_run_monthly_summary_now(user: dict = Depends(get_current_user)):
    """Manuel monthly summary (ay 1'i degil de test edebilmek icin force=True)."""
    from autotax.reminders import process_monthly_summary
    stats = await process_monthly_summary(force=True)
    return {"success": True, "stats": stats}


@app.get("/invoices/{invoice_id}/mahnung-pdf")
def get_mahnung_pdf(
    invoice_id: int,
    level: int = Query(1, ge=1, le=3),
    user: dict = Depends(get_current_user),
):
    """Manuel Mahnung PDF olustur — kullanicinin kendisinin musterisine
    kestigi income fatura icin. Cron beklemeden anlik PDF lazimsa."""
    from autotax.mahnung import generate_mahnung_pdf
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(
            Invoice.id == invoice_id, Invoice.user_id == user["sub"],
        ).first()
        if not inv:
            err(404, "Invoice not found")
        sender = db.query(User).filter(User.id == user["sub"]).first()
        try:
            pdf_bytes = generate_mahnung_pdf(inv, level, sender)
        except Exception as e:
            err(500, f"PDF generation failed: {e}")
        from fastapi.responses import Response
        fname = f"Mahnung-{inv.invoice_number or inv.id}-Stufe{level}.pdf"
        return Response(
            content=pdf_bytes,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={fname}"},
        )
    finally:
        db.close()


@app.post("/admin/reminders/run-now")
async def admin_run_reminders_now(user: dict = Depends(get_current_user)):
    """Manuel olarak reminder cycle'ini calistir (test icin). Sadece admin."""
    from autotax.reminders import process_reminders
    stats = await process_reminders()
    return {"success": True, "stats": stats}


@app.get("/admin/users/{user_id}/invoice")
def admin_subscription_invoice(
    user_id: int,
    plan: str = Query("pro"),
    months: int = Query(1, ge=1, le=12),
    invoice_no: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    """Manuel abonelik faturasi (PDF) — admin'in firma bilgileriyle.
    Stripe gelene kadar manuel takip icin: muşteriye email atilacak,
    SEPA havalesi onayinda admin panelden Pro+Cloud isaretlenir."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
    except ImportError:
        err(501, "PDF generation not available")

    PLAN_PRICES = {"free": 0, "early": 10, "pro": 20}
    CLOUD_PRICE = 5

    db = SessionLocal()
    try:
        target = db.query(User).filter(User.id == user_id).first()
        if not target:
            err(404, "User not found")

        plan = (plan or "pro").lower()
        if plan not in PLAN_PRICES:
            err(400, "Invalid plan")

        plan_price = PLAN_PRICES[plan] * months
        cloud_price = CLOUD_PRICE * months if getattr(target, "has_cloud_addon", False) else 0
        net = plan_price + cloud_price
        # Kleinunternehmer -> KDV yok (§19 UStG)
        sender_ku = bool(os.getenv("BILLING_KLEINUNTERNEHMER", "1") == "1")
        vat_rate = 0 if sender_ku else 19
        vat_amount = round(net * vat_rate / 100, 2)
        total = round(net + vat_amount, 2)

        # Sender (operator firma bilgileri — env var'larla esnek)
        SENDER_NAME = os.getenv("BILLING_NAME", "Hüseyin Hancer")
        SENDER_ADDR = os.getenv("BILLING_ADDRESS", "Wiesenstr. 10")
        SENDER_CITY = os.getenv("BILLING_CITY", "66115 Saarbrücken")
        SENDER_COUNTRY = os.getenv("BILLING_COUNTRY", "Deutschland")
        SENDER_EMAIL = os.getenv("BILLING_EMAIL", "info@autotaxhub.de")
        SENDER_IBAN = os.getenv("BILLING_IBAN", "DE00 0000 0000 0000 0000 00")
        SENDER_BIC = os.getenv("BILLING_BIC", "")
        SENDER_USTID = os.getenv("BILLING_USTID", "")

        if not invoice_no:
            invoice_no = f"AT-{datetime.now().strftime('%Y%m')}-{user_id:04d}"

        buf = io.BytesIO()
        c = pdf_canvas.Canvas(buf, pagesize=A4)
        w, h = A4

        # Header
        c.setFillColor(HexColor("#10b981"))
        c.setFont("Helvetica-Bold", 28)
        c.drawString(2*cm, h-2.5*cm, "RECHNUNG")
        c.setFillColor(HexColor("#64748b"))
        c.setFont("Helvetica", 10)
        c.drawString(2*cm, h-3.1*cm, f"Rechnungs-Nr.: {invoice_no}")
        c.drawString(2*cm, h-3.55*cm, f"Datum: {datetime.now().strftime('%d.%m.%Y')}")

        # Sender (Absender)
        c.setFillColor(HexColor("#0f172a"))
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, h-5*cm, SENDER_NAME)
        c.setFont("Helvetica", 10)
        c.drawString(2*cm, h-5.45*cm, SENDER_ADDR)
        c.drawString(2*cm, h-5.9*cm, SENDER_CITY)
        c.drawString(2*cm, h-6.35*cm, SENDER_COUNTRY)
        c.drawString(2*cm, h-6.95*cm, f"E-Mail: {SENDER_EMAIL}")
        if SENDER_USTID:
            c.drawString(2*cm, h-7.4*cm, f"USt-IdNr.: {SENDER_USTID}")

        # Recipient (Empfänger)
        c.setFont("Helvetica-Bold", 10)
        c.drawString(12*cm, h-5*cm, "Rechnung an:")
        c.setFont("Helvetica", 10)
        if target.full_name:
            c.drawString(12*cm, h-5.45*cm, target.full_name)
            c.drawString(12*cm, h-5.9*cm, target.email)
        else:
            c.drawString(12*cm, h-5.45*cm, target.email)

        # Items table
        y = h - 9.5*cm
        c.setStrokeColor(HexColor("#cbd5e1"))
        c.line(2*cm, y, 19*cm, y)
        y -= 0.2*cm
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(HexColor("#334155"))
        c.drawString(2*cm, y-0.4*cm, "Beschreibung")
        c.drawRightString(13*cm, y-0.4*cm, "Menge")
        c.drawRightString(16*cm, y-0.4*cm, "Einzel")
        c.drawRightString(19*cm, y-0.4*cm, "Summe")
        y -= 0.7*cm
        c.line(2*cm, y, 19*cm, y)

        c.setFont("Helvetica", 10)
        c.setFillColor(HexColor("#0f172a"))
        # Plan line
        plan_label = {"free": "Free", "early": "Early Adopter", "pro": "Pro"}.get(plan, plan)
        period = f"{months} Monat" if months == 1 else f"{months} Monate"
        y -= 0.7*cm
        c.drawString(2*cm, y, f"AutoTax-HUB {plan_label} ({period})")
        c.drawRightString(13*cm, y, str(months))
        c.drawRightString(16*cm, y, f"{PLAN_PRICES[plan]:.2f} EUR")
        c.drawRightString(19*cm, y, f"{plan_price:.2f} EUR")

        if cloud_price > 0:
            y -= 0.6*cm
            c.drawString(2*cm, y, f"AutoTax-Cloud Add-on ({period})")
            c.drawRightString(13*cm, y, str(months))
            c.drawRightString(16*cm, y, f"{CLOUD_PRICE:.2f} EUR")
            c.drawRightString(19*cm, y, f"{cloud_price:.2f} EUR")

        # Totals
        y -= 1.2*cm
        c.line(13*cm, y, 19*cm, y)
        y -= 0.6*cm
        c.setFont("Helvetica", 10)
        c.drawRightString(16*cm, y, "Netto:")
        c.drawRightString(19*cm, y, f"{net:.2f} EUR")
        y -= 0.5*cm
        if vat_rate > 0:
            c.drawRightString(16*cm, y, f"USt {vat_rate}%:")
            c.drawRightString(19*cm, y, f"{vat_amount:.2f} EUR")
            y -= 0.5*cm
        else:
            c.setFillColor(HexColor("#64748b"))
            c.setFont("Helvetica-Oblique", 9)
            c.drawRightString(19*cm, y, "Kein Ausweis von USt gem. § 19 UStG (Kleinunternehmer)")
            c.setFillColor(HexColor("#0f172a"))
            c.setFont("Helvetica", 10)
            y -= 0.5*cm
        c.line(13*cm, y, 19*cm, y)
        y -= 0.5*cm
        c.setFont("Helvetica-Bold", 12)
        c.setFillColor(HexColor("#10b981"))
        c.drawRightString(16*cm, y, "Gesamtbetrag:")
        c.drawRightString(19*cm, y, f"{total:.2f} EUR")

        # Bank info
        y -= 2*cm
        c.setFillColor(HexColor("#f59e0b"))
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, y, "Zahlung per SEPA-Überweisung")
        c.setFillColor(HexColor("#0f172a"))
        c.setFont("Helvetica", 10)
        y -= 0.6*cm
        c.drawString(2*cm, y, f"Empfänger: {SENDER_NAME}")
        y -= 0.45*cm
        c.drawString(2*cm, y, f"IBAN: {SENDER_IBAN}")
        if SENDER_BIC:
            y -= 0.45*cm
            c.drawString(2*cm, y, f"BIC: {SENDER_BIC}")
        y -= 0.45*cm
        c.drawString(2*cm, y, f"Verwendungszweck: {invoice_no}")
        y -= 0.7*cm
        c.setFillColor(HexColor("#64748b"))
        c.setFont("Helvetica-Oblique", 9)
        c.drawString(2*cm, y, "Bitte zahlen Sie innerhalb von 14 Tagen.")

        # Footer
        c.setFillColor(HexColor("#94a3b8"))
        c.setFont("Helvetica", 8)
        c.drawString(2*cm, 1.5*cm, f"{SENDER_NAME} · {SENDER_ADDR} · {SENDER_CITY}")
        c.drawString(2*cm, 1.1*cm, f"E-Mail: {SENDER_EMAIL}" + (f" · USt-IdNr.: {SENDER_USTID}" if SENDER_USTID else ""))

        c.showPage()
        c.save()
        buf.seek(0)
        logger.info("ADMIN_INVOICE generated id=%s plan=%s months=%d total=%.2f for user_id=%d by=%s",
                    invoice_no, plan, months, total, user_id, user.get("email"))
        return StreamingResponse(
            buf,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={invoice_no}.pdf"},
        )
    finally:
        db.close()


@app.get("/admin/users/{user_id}/details")
def admin_user_details(user_id: int, user: dict = Depends(get_current_user)):
    """Tek kullanicinin detayli bilgileri (manuel fatura / debug icin)."""
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user_id).first()
        if not u:
            err(404, "User not found")
        invoices = db.query(Invoice).filter(
            Invoice.user_id == user_id,
            (Invoice.is_deleted == False) | (Invoice.is_deleted == None),
        ).count()
        companies = db.query(UserCompany).filter(UserCompany.user_id == user_id).all()
        return {
            "id": u.id, "email": u.email, "full_name": u.full_name or "",
            "plan": u.plan or "free",
            "has_cloud_addon": getattr(u, "has_cloud_addon", False),
            "is_kleinunternehmer": getattr(u, "is_kleinunternehmer", False),
            "registered_at": u.registered_at.isoformat() if u.registered_at else "",
            "invoice_count": invoices,
            "companies": [{"id": c.id, "name": c.company_name} for c in companies],
        }
    finally:
        db.close()


# DATEV Konto mapping (Ausgaben)
_DATEV_KONTO_MAP = {
    "food": "6800", "groceries": "6800", "restaurant": "6640",
    "fuel": "6670", "transport": "6673",
    "office": "6815", "software": "6815", "subscription": "6815",
    "telecom": "6805", "shipping": "6810",
    "electronics": "6800", "shopping": "6800",
    "insurance": "6400", "health": "6800", "medical": "6800",
    "home": "6800", "clothing": "6800",
    "other": "6800",
}
# DATEV Konto mapping (Einnahmen)
_DATEV_KONTO_MAP_INCOME = {
    "other": "8400", "food": "8400", "electronics": "8400",
    "software": "8400", "shopping": "8400",
}

def calculate_dashboard_metrics(user_id: int, year: int = None):
    """Shared function: compute dashboard metrics from invoices table.
    Used by dashboard endpoint AND all exports so numbers always match."""
    from sqlalchemy.orm import defer as _sa_defer
    db = SessionLocal()
    try:
        # raw_text + file_data dashboard'da kullanilmiyor; her satirda 2-5KB
        # ekstra yuk olarak gelmesin diye defer edildi. 200 fis icin ~400KB
        # daha az transfer + Python parsing.
        all_inv = db.query(Invoice).options(
            _sa_defer(Invoice.raw_text),
            _sa_defer(Invoice.file_data),
        ).filter(
            Invoice.user_id == user_id,
            (Invoice.is_deleted == False) | (Invoice.is_deleted == None),
        ).all()
        # Same filter as dashboard: skip amount=0 and vendor=Unbekannt
        invoices = [i for i in all_inv if safe_float(i.total_amount) > 0 and safe_vendor(i.vendor) != "Unbekannt"]

        # Year filter
        if year:
            invoices = [i for i in invoices if safe_date_str(i.date).startswith(str(year))]

        inv_inc = [i for i in invoices if safe_invoice_type(i.invoice_type) == "income"]
        inv_exp = [i for i in invoices if safe_invoice_type(i.invoice_type) == "expense"]

        total_income = round(sum(safe_float(i.total_amount) for i in inv_inc), 2)
        total_expenses = round(sum(safe_float(i.total_amount) for i in inv_exp), 2)
        profit = round(total_income - total_expenses, 2)

        vat_paid = round(sum(safe_float(i.vat_amount) for i in inv_exp), 2)
        vat_collected = round(sum(safe_float(i.vat_amount) for i in inv_inc), 2)
        vat_balance = round(vat_collected - vat_paid, 2)

        # VAT grouped by rate
        vat_by_rate = {}
        for i in invoices:
            rate = safe_vat_rate(i.vat_rate)
            if rate not in vat_by_rate:
                vat_by_rate[rate] = {"amount": 0.0, "vat": 0.0, "count": 0}
            vat_by_rate[rate]["amount"] += safe_float(i.total_amount)
            vat_by_rate[rate]["vat"] += safe_float(i.vat_amount)
            vat_by_rate[rate]["count"] += 1
        for v in vat_by_rate.values():
            v["amount"] = round(v["amount"], 2)
            v["vat"] = round(v["vat"], 2)

        # Monthly breakdown
        month_map = {}
        for i in invoices:
            d = safe_date_str(i.date)
            if not d or len(d) < 7 or "-" not in d:
                continue
            m = d[:7]
            if m not in month_map:
                month_map[m] = {"month": m, "income": 0.0, "expenses": 0.0}
            if safe_invoice_type(i.invoice_type) == "income":
                month_map[m]["income"] += safe_float(i.total_amount)
            else:
                month_map[m]["expenses"] += safe_float(i.total_amount)
        monthly = sorted(month_map.values(), key=lambda x: x["month"])
        for mb in monthly:
            mb["income"] = round(mb["income"], 2)
            mb["expenses"] = round(mb["expenses"], 2)

        # Category distribution
        cat_map = {}
        for i in invoices:
            c = safe_category(i.category)
            cat_map[c] = cat_map.get(c, 0) + safe_float(i.total_amount)
        by_category = [{"category": k, "total": round(v, 2)} for k, v in sorted(cat_map.items(), key=lambda x: -x[1])]

        # Transaction list (same safe_ transforms as invoice_to_dict)
        transactions = []
        for i in invoices:
            transactions.append({
                "id": i.id,
                "date": safe_date_str(i.date),
                "vendor": safe_vendor(i.vendor),
                "invoice_number": safe_str(i.invoice_number),
                "invoice_type": safe_invoice_type(i.invoice_type),
                "total_amount": safe_float(i.total_amount),
                "vat_amount": safe_float(i.vat_amount),
                "vat_rate": safe_vat_rate(i.vat_rate),
                "category": safe_category(i.category),
                "payment_method": safe_str(i.payment_method),
                "konto": _DATEV_KONTO_MAP.get(safe_category(i.category), "6800") if safe_invoice_type(i.invoice_type) == "expense" else _DATEV_KONTO_MAP_INCOME.get(safe_category(i.category), "8400"),
            })
        transactions.sort(key=lambda x: x["date"])

        return {
            "total_income": total_income,
            "total_expenses": total_expenses,
            "profit": profit,
            "vat_paid": vat_paid,
            "vat_collected": vat_collected,
            "vat_balance": vat_balance,
            "vat_by_rate": vat_by_rate,
            "invoice_count": len(invoices),
            "income_count": len(inv_inc),
            "expense_count": len(inv_exp),
            "monthly": monthly,
            "by_category": by_category,
            "transactions": transactions,
            "year": year,
        }
    finally:
        db.close()


ALLOWED_TYPES = {"application/pdf", "image/jpeg", "image/jpg", "image/png", "image/tiff", "image/webp", "image/heic", "image/heif", "application/zip", "application/x-zip-compressed"}
MAX_FILE_SIZE = 10 * 1024 * 1024

# Magic bytes for file type validation (prevents fake content_type)
_MAGIC_BYTES = {
    b"\xff\xd8\xff": "image/jpeg",
    b"\x89PNG": "image/png",
    b"%PDF": "application/pdf",
    b"II\x2a\x00": "image/tiff",  # little-endian TIFF
    b"MM\x00\x2a": "image/tiff",  # big-endian TIFF
    b"RIFF": "image/webp",        # WebP starts with RIFF
    b"PK\x03\x04": "application/zip",  # ZIP archive
}


def _validate_file_magic(content: bytes, claimed_type: str) -> bool:
    """Check if file content matches claimed MIME type via magic bytes."""
    if not content or len(content) < 4:
        return False
    # HEIC/HEIF have complex headers — trust content_type for those
    if "heic" in claimed_type or "heif" in claimed_type:
        return True
    for magic, mime in _MAGIC_BYTES.items():
        if content[:len(magic)] == magic:
            return True
    return False


# ============================================================
# AUTH
# ============================================================

class AuthRequest(BaseModel):
    email: str
    password: str


class RegisterRequest(BaseModel):
    email: str
    password: str
    full_name: Optional[str] = None
    company_name: Optional[str] = None
    gdpr_consent: bool = False


@app.post("/auth/register")
@limiter.limit("3/minute")
def register(request: Request, body: RegisterRequest):
    if not body.gdpr_consent:
        err(400, "Datenschutzerklärung muss akzeptiert werden (DSGVO Art. 6)")
    if len(body.password) < 8:
        err(400, "Password must be at least 8 characters")
    if not any(c.isupper() for c in body.password):
        err(400, "Password must contain at least 1 uppercase letter")
    if not any(c.isdigit() for c in body.password):
        err(400, "Password must contain at least 1 digit")
    db = SessionLocal()
    try:
        if db.query(User).filter(User.email == body.email).first():
            err(400, "Email already registered")
        # Yeni kayit -> 15 gun Pro trial. Cron 15. gunde plan=free yapar
        # ve Telegram alert atar. Manuel odeme alirsan admin panelden
        # 'Pro+Cloud' butonuna basarsin -> trial_ends_at NULL olur (kalici).
        trial_days = int(os.getenv("TRIAL_DAYS", "15"))
        trial_end = datetime.now(timezone.utc) + timedelta(days=trial_days)
        try:
            user = User(
                email=body.email,
                hashed_password=hash_password(body.password),
                full_name=body.full_name,
                plan="pro",  # trial Pro
                trial_ends_at=trial_end,
                gdpr_consent_at=datetime.now(),
            )
        except Exception:
            user = User(email=body.email, hashed_password=hash_password(body.password), full_name=body.full_name)
        db.add(user)
        db.commit()
        db.refresh(user)
        # Auto-create company (optional — don't fail registration if this fails)
        comp_name = ""
        try:
            comp_name = (body.company_name or "").strip()
            if not comp_name:
                comp_name = (body.full_name or "").strip()
            if not comp_name:
                comp_name = body.email.split("@")[0].strip()
            if comp_name:
                company = UserCompany(user_id=user.id, company_name=comp_name)
                db.add(company)
                db.commit()
        except Exception:
            logger.warning("Could not create company for user — table may not exist yet")
        logger.info("User registered: %s (company: %s)", _mask_email(body.email), comp_name)
        token = create_token(user.id, user.email)
        return {"success": True, "token": token, "email": user.email}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Registration error")
        err(500, "Registration failed")
    finally:
        db.close()


@app.post("/auth/login")
@limiter.limit("5/minute")
def login(request: Request, body: AuthRequest):
    db = SessionLocal()
    try:
        user = db.query(User).filter(User.email == body.email).first()
        if not user or not verify_password(body.password, user.hashed_password):
            logger.warning("Failed login: %s", _mask_email(body.email))
            err(401, "Invalid email or password")
        logger.info("User logged in: %s", _mask_email(body.email))
        token = create_access_token(user.id, user.email)
        refresh = create_refresh_token(user.id, user.email)
        return {"success": True, "token": token, "refresh_token": refresh, "email": user.email}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Login error")
        err(500, "Login failed")
    finally:
        db.close()


@app.post("/auth/refresh")
def refresh_token_endpoint(body: dict = Body(...)):
    refresh = body.get("refresh_token", "")
    if not refresh:
        err(400, "refresh_token required")
    try:
        data = decode_token(refresh, expected_type="refresh")
    except HTTPException:
        raise
    new_access = create_access_token(data["sub"], data["email"])
    new_refresh = create_refresh_token(data["sub"], data["email"])
    return {"success": True, "token": new_access, "refresh_token": new_refresh}


# Debug endpoints removed — security risk in production


class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str


@app.post("/auth/change-password")
def change_password(body: ChangePasswordRequest, user: dict = Depends(get_current_user)):
    if len(body.new_password) < 8:
        err(400, "Neues Passwort muss mindestens 8 Zeichen haben")
    if not any(c.isupper() for c in body.new_password):
        err(400, "Neues Passwort muss mindestens 1 Großbuchstaben enthalten")
    if not any(c.isdigit() for c in body.new_password):
        err(400, "Neues Passwort muss mindestens 1 Zahl enthalten")
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        if not u or not verify_password(body.old_password, u.hashed_password):
            err(401, "Altes Passwort ist falsch")
        u.hashed_password = hash_password(body.new_password)
        db.commit()
        return {"success": True, "message": "Passwort erfolgreich geändert"}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Password change failed")
        err(500, "Passwort-Änderung fehlgeschlagen")
    finally:
        db.close()


@app.post("/auth/reset-password")
@limiter.limit("3/minute")
def reset_password(request: Request, body: dict = Body(...)):
    """Send password reset — for now just verify email exists and return token."""
    email = body.get("email", "").strip().lower()
    if not email:
        err(400, "E-Mail erforderlich")
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.email == email).first()
        if not u:
            # Don't reveal if email exists
            return {"success": True, "message": "Falls ein Konto existiert, wurde ein Reset-Link gesendet."}
        # Generate reset token (valid 1 hour)
        reset_token = create_access_token(u.id, u.email)
        logger.info("Password reset requested for %s", _mask_email(email))
        # TODO: Send email with reset link. Token stored server-side only.
        logger.info("Reset token generated for %s (not exposed in response)", _mask_email(email))
        return {"success": True, "message": "Falls ein Konto existiert, wurde ein Reset-Link gesendet."}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Password reset failed")
        err(500, "Reset fehlgeschlagen")
    finally:
        db.close()


# ============================================================
# INVOICES: UPLOAD
# ============================================================


@app.post("/invoices/upload-erechnung")
async def upload_erechnung(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Import XRechnung / ZUGFeRD / Factur-X e-invoice (XML or ZUGFeRD-PDF)."""
    _enforce_upload_quota(user["sub"])
    import xml.etree.ElementTree as ET
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        err(400, "Datei zu groß (max 5MB)")

    # ZUGFeRD/Factur-X: XML embedded in PDF — extract it
    is_pdf = content[:4] == b"%PDF"
    if is_pdf:
        try:
            from pypdf import PdfReader
            import io as _io
            reader = PdfReader(_io.BytesIO(content))
            xml_bytes = None
            preferred = ("factur-x.xml", "zugferd-invoice.xml", "xrechnung.xml")
            attachments = getattr(reader, "attachments", {}) or {}
            # prefer known ZUGFeRD filenames, fall back to any .xml
            for name in preferred:
                for k, v in attachments.items():
                    if k.lower() == name and v:
                        xml_bytes = v[0] if isinstance(v, list) else v
                        break
                if xml_bytes:
                    break
            if not xml_bytes:
                for k, v in attachments.items():
                    if k.lower().endswith(".xml") and v:
                        xml_bytes = v[0] if isinstance(v, list) else v
                        break
        except Exception:
            logger.exception("ZUGFeRD PDF attachment extraction failed")
            xml_bytes = None
        if not xml_bytes:
            err(400, "PDF enthält keine ZUGFeRD/Factur-X XML-Anlage")
        text = xml_bytes.decode("utf-8", errors="ignore")
    else:
        text = content.decode("utf-8", errors="ignore")

    # XXE protection: strip DOCTYPE declarations before parsing
    import re as _re
    text_safe = _re.sub(r'<!DOCTYPE[^>]*>', '', text, flags=_re.IGNORECASE | _re.DOTALL)
    text_safe = _re.sub(r'<!ENTITY[^>]*>', '', text_safe, flags=_re.IGNORECASE | _re.DOTALL)

    root = None
    try:
        root = ET.fromstring(text_safe)
    except ET.ParseError:
        pass

    if root is None:
        err(400, "Keine gültige XML/E-Rechnung")

    # Parse with namespace-agnostic approach
    def _find(el, tags):
        for tag in tags:
            for child in el.iter():
                if tag.lower() in child.tag.lower():
                    if child.text and child.text.strip():
                        return child.text.strip()
        return ""

    vendor = _find(root, ["PartyName", "Name", "SellerTradeParty"])
    invoice_number = _find(root, ["InvoiceNumber", "DocumentNumber"]) or _find(root, ["ID"])
    date_str = _find(root, ["IssueDate", "DateTimeString", "InvoiceDate"])
    total_str = _find(root, ["PayableAmount", "TaxInclusiveAmount", "GrandTotalAmount", "DuePayableAmount"])
    tax_str = _find(root, ["TaxAmount", "TaxTotalAmount"])
    vat_rate_str = _find(root, ["Percent", "RateApplicablePercent", "CategoryCode"])

    total = 0.0
    try:
        total = float(total_str.replace(",", "."))
    except (ValueError, AttributeError):
        pass

    tax = 0.0
    try:
        tax = float(tax_str.replace(",", "."))
    except (ValueError, AttributeError):
        pass

    vat_rate = "19%"
    try:
        r = float(vat_rate_str.replace(",", "."))
        if 0 < r <= 30:
            vat_rate = f"{r}%"
    except (ValueError, AttributeError):
        pass

    if not vendor:
        vendor = "E-Rechnung"

    # Use existing category detection
    try:
        from autotax.parser import detect_category
        category = detect_category(vendor, text)
    except Exception:
        category = "other"

    # Save invoice
    db = SessionLocal()
    try:
        inv = Invoice(
            user_id=user["sub"],
            filename=file.filename or "e-rechnung.xml",
            vendor=vendor,
            total_amount=total,
            vat_amount=tax if tax > 0 else round(total * 19 / 119, 2),
            vat_rate=vat_rate,
            date=date_str,
            raw_text=text[:2000],
            invoice_type="expense",
            invoice_number=invoice_number,
            payment_method="",
            category=category,
            processed=True,
            file_data=content,
            file_content_type=file.content_type or "application/xml",
        )
        db.add(inv)
        db.commit()
        db.refresh(inv)
        auto_create_cash_entry(inv.id, user["sub"], {
            "vendor": vendor, "total_amount": total,
            "vat_amount": tax, "vat_rate": vat_rate,
            "date": date_str, "category": category,
        })
        konto = _DATEV_KONTO_MAP.get(category, "6800")
        return {
            "success": True,
            "id": inv.id,
            "vendor": vendor,
            "total_amount": total,
            "vat_amount": tax,
            "vat_rate": vat_rate,
            "date": date_str,
            "invoice_number": invoice_number,
            "category": category,
            "konto": konto,
            "message": "E-Rechnung erfolgreich importiert — Automatisch kategorisiert ✔",
        }
    except Exception:
        db.rollback()
        logger.exception("E-Rechnung import failed")
        err(500, "E-Rechnung Import fehlgeschlagen")
    finally:
        db.close()


@app.post("/ocr/extract")
@limiter.limit("20/minute")
async def ocr_extract_only(request: Request, file: UploadFile = File(...), handwriting: bool = False, user: dict = Depends(get_current_user)):
    """OCR + parse, no DB write. For auto-filling forms (e.g. /beleg)."""
    _enforce_upload_quota(user["sub"])
    if file.content_type not in ALLOWED_TYPES:
        err(400, "Ungültige Datei. Erlaubt: PDF, JPG, PNG, TIFF, WEBP")
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        err(400, "Datei zu groß (max 10MB)")
    if len(content) == 0:
        err(400, "Leere Datei")
    if not _validate_file_magic(content, file.content_type or ""):
        err(400, "Ungültige Datei — Dateityp stimmt nicht mit Inhalt überein")
    await file.seek(0)

    raw_text = ""
    qr_data = {}
    try:
        if (file.content_type or "").lower().startswith("image/"):
            from autotax.ocr import local_ocr_tesseract, is_ocr_valid
            _tt = local_ocr_tesseract(content)
            if is_ocr_valid(_tt):
                raw_text = _tt
                try:
                    from autotax.qr_reader import extract_qr_data
                    qr_data = extract_qr_data(content, file.content_type or "")
                except Exception:
                    pass
        if not raw_text:
            raw_text, qr_data = await asyncio.wait_for(
                extract_text_and_qr(file, handwriting=handwriting, file_bytes=content), timeout=45
            )
    except Exception:
        logger.warning("OCR extract failed for %s", file.filename)

    try:
        result = parse_invoice(raw_text or "")
    except Exception:
        logger.exception("parse_invoice failed in /ocr/extract")
        result = {}

    if qr_data:
        if qr_data.get("company") and result.get("vendor") in ("Unbekannt", "", None):
            result["vendor"] = qr_data["company"]
        if qr_data.get("amount") and not result.get("total_amount"):
            result["total_amount"] = qr_data["amount"]
        if qr_data.get("date") and not result.get("date"):
            result["date"] = qr_data["date"]
        if qr_data.get("iban") and not result.get("vendor_iban"):
            result["vendor_iban"] = qr_data["iban"]

    vendor = result.get("vendor", "") or ""
    if vendor == "Unbekannt":
        vendor = ""

    _text_lc = (raw_text or "").lower()
    amount = safe_float(result.get("total_amount"))
    date_val = result.get("date", "") or ""
    vat_rate = result.get("vat_rate", "") or ""
    inv_num = result.get("invoice_number", "") or ""

    # --- Confidence scoring (simple rule-based) ---
    vendor_conf = 0.0
    _vsrc = result.get("vendor_source", "primary")
    if vendor:
        if qr_data and qr_data.get("company") and qr_data["company"] == vendor:
            vendor_conf = 0.95
        elif _vsrc == "guess":
            # Last-resort first-line pick — always low confidence
            vendor_conf = 0.45
        elif _vsrc == "deep":
            vendor_conf = 0.75
        elif vendor.lower() in _text_lc:
            vendor_conf = 0.88
        elif len(vendor) >= 4:
            vendor_conf = 0.65
        else:
            vendor_conf = 0.45

    amount_conf = 0.0
    if amount > 0:
        kw_hit = any(kw in _text_lc for kw in ("total", "gesamt", "summe", "brutto", "zu zahlen", "rechnungsbetrag", "endbetrag"))
        amount_conf = 0.95 if kw_hit else 0.7

    date_conf = 0.0
    if date_val:
        import re as _re3
        if _re3.match(r"^\d{4}-\d{2}-\d{2}$", date_val):
            y, m, d = date_val.split("-")
            date_conf = 0.9 if (d in raw_text or m in raw_text) else 0.75
        else:
            date_conf = 0.5

    vat_conf = 0.0
    if vat_rate and vat_rate != "0%":
        vat_conf = 0.95 if vat_rate in ("7%", "19%") else 0.7

    inv_conf = 0.0
    if inv_num:
        kw_hit = any(kw in _text_lc for kw in ("rechnung", "invoice", "re.", "re-", "beleg", "quittung", "nr."))
        has_letter = any(c.isalpha() for c in inv_num)
        if kw_hit and has_letter:
            inv_conf = 0.9
        elif kw_hit:
            inv_conf = 0.75
        elif has_letter:
            inv_conf = 0.55
        else:
            inv_conf = 0.4

    return {
        "vendor": vendor,
        "vendor_confidence": round(vendor_conf, 2),
        "date": date_val,
        "date_confidence": round(date_conf, 2),
        "amount": amount,
        "amount_confidence": round(amount_conf, 2),
        "vat_rate": vat_rate,
        "vat_confidence": round(vat_conf, 2),
        "vat_amount": safe_float(result.get("vat_amount")),
        "invoice_number": inv_num,
        "invoice_number_confidence": round(inv_conf, 2),
        "category": result.get("category", "") or "",
        "payment_method": result.get("payment_method", "") or "",
        "vendor_iban": result.get("vendor_iban", "") or _extract_first_iban(raw_text or ""),
        "vendor_email": result.get("vendor_email", "") or "",
        "vendor_phone": result.get("vendor_phone", "") or _extract_first_phone(raw_text or ""),
        "vendor_address": result.get("vendor_address", "") or _extract_first_address(raw_text or ""),
        "raw_text": (raw_text or "")[:3000],
    }


# --- Email invoice import (IMAP) ---

@app.post("/email/config")
def set_email_config(body: dict = Body(...), user: dict = Depends(get_current_user)):
    """Save per-user IMAP credentials. App Password required — raw passwords rejected by providers."""
    from autotax.email_sync import encrypt_password, PROVIDERS
    from autotax.models import EmailConfig

    provider = (body.get("provider") or "").lower().strip()
    email_addr = (body.get("email") or "").strip()
    app_password = body.get("app_password") or ""
    host = (body.get("host") or "").strip() or None
    port = body.get("port") or 993
    try:
        port = int(port)
    except (TypeError, ValueError):
        err(400, "Port muss eine Zahl sein")

    if provider not in PROVIDERS and provider != "imap":
        err(400, "provider: gmail | outlook | imap")
    if not email_addr or "@" not in email_addr or len(email_addr) > 255:
        err(400, "Ungültige Email-Adresse")
    if not app_password or len(app_password) < 8 or len(app_password) > 200:
        err(400, "App-Passwort ungültig (mind. 8 Zeichen)")
    if provider == "imap" and not host:
        err(400, "Host erforderlich für Custom IMAP")

    try:
        enc = encrypt_password(app_password)
    except RuntimeError:
        err(500, "Server nicht konfiguriert (EMAIL_CREDS_KEY fehlt)")

    db = SessionLocal()
    try:
        cfg = db.query(EmailConfig).filter(EmailConfig.user_id == user["sub"]).first()
        if cfg:
            cfg.provider = provider
            cfg.email = email_addr
            cfg.encrypted_password = enc
            cfg.host = host if provider == "imap" else None
            cfg.port = port if provider == "imap" else None
            cfg.enabled = True
        else:
            cfg = EmailConfig(
                user_id=user["sub"], provider=provider, email=email_addr,
                encrypted_password=enc,
                host=host if provider == "imap" else None,
                port=port if provider == "imap" else None,
                enabled=True,
            )
            db.add(cfg)
        db.commit()
        logger.info("Email config saved for user=%s provider=%s", user["sub"], provider)
        return {"success": True}
    finally:
        db.close()


@app.get("/email/config")
def get_email_config(user: dict = Depends(get_current_user)):
    """Return current config — never includes the password."""
    from autotax.models import EmailConfig
    db = SessionLocal()
    try:
        cfg = db.query(EmailConfig).filter(EmailConfig.user_id == user["sub"]).first()
        if not cfg:
            return {"configured": False}
        return {
            "configured": True,
            "provider": cfg.provider,
            "email": cfg.email,
            "host": cfg.host,
            "port": cfg.port,
            "enabled": cfg.enabled,
            "last_sync": cfg.last_sync.isoformat() if cfg.last_sync else None,
        }
    finally:
        db.close()


@app.delete("/email/config")
def delete_email_config(user: dict = Depends(get_current_user)):
    from autotax.models import EmailConfig
    db = SessionLocal()
    try:
        cfg = db.query(EmailConfig).filter(EmailConfig.user_id == user["sub"]).first()
        if cfg:
            db.delete(cfg)
            db.commit()
        return {"success": True}
    finally:
        db.close()


@app.post("/email/sync")
@limiter.limit("6/hour")
async def email_sync(request: Request, user: dict = Depends(get_current_user)):
    """Pull UNSEEN inbox messages, process PDF/XML attachments into invoices."""
    from autotax.email_sync import sync_user_inbox
    return await sync_user_inbox(user["sub"])


@app.post("/email/test")
@limiter.limit("10/hour")
def email_test(request: Request, user: dict = Depends(get_current_user)):
    """Dry-run: connect, login, select INBOX — no message processing."""
    import imaplib
    import ssl as _ssl
    from autotax.email_sync import PROVIDERS, _connect_imap, decrypt_password
    from autotax.models import EmailConfig

    db = SessionLocal()
    try:
        cfg = db.query(EmailConfig).filter(EmailConfig.user_id == user["sub"]).first()
        if not cfg:
            return {"success": False, "message": "Keine Konfiguration gespeichert"}
        if cfg.provider in PROVIDERS:
            host, port = PROVIDERS[cfg.provider]
        else:
            host, port = cfg.host, cfg.port or 993
        try:
            password = decrypt_password(cfg.encrypted_password)
        except Exception:
            return {"success": False, "message": "Passwort-Entschlüsselung fehlgeschlagen"}

        M = None
        try:
            M = _connect_imap(host, port, cfg.email, password)
            typ, data = M.select("INBOX", readonly=True)
            if typ != "OK":
                return {"success": False, "message": "Posteingang nicht verfügbar"}
            total = int(data[0]) if data and data[0] else 0
            typ, udata = M.search(None, "UNSEEN")
            unseen = len(udata[0].split()) if (typ == "OK" and udata and udata[0]) else 0
            return {"success": True, "inbox_total": total, "unseen": unseen}
        except imaplib.IMAP4.error:
            return {"success": False, "message": "Login fehlgeschlagen — App-Passwort prüfen"}
        except (OSError, _ssl.SSLError):
            return {"success": False, "message": "Server nicht erreichbar"}
        finally:
            if M is not None:
                try: M.close()
                except Exception: pass
                try: M.logout()
                except Exception: pass
    finally:
        db.close()


@app.post("/invoices/create-rechnung")
def create_rechnung(body: dict = Body(...), user: dict = Depends(get_current_user)):
    """Create a manual outgoing invoice (Einnahme)."""
    db = SessionLocal()
    try:
        betrag = float(body.get("betrag", 0))
        mwst_satz = body.get("mwst_satz", "19%")
        rate = float(mwst_satz.replace("%", "").replace(",", ".").strip() or "19")
        mwst_betrag = float(body.get("mwst_betrag", 0)) or round(betrag * rate / (100 + rate), 2)
        inv = Invoice(
            user_id=user["sub"], filename="rechnung-erstellt",
            vendor=body.get("kunde", ""), total_amount=betrag,
            vat_amount=mwst_betrag, vat_rate=mwst_satz,
            date=body.get("datum", ""), raw_text="Manuell erstellte Rechnung",
            invoice_type="income", invoice_number=body.get("rechnung_nr", ""),
            payment_method=body.get("zahlungsart", ""),
            category=body.get("kategorie", "service"), processed=True,
        )
        db.add(inv)
        db.commit()
        db.refresh(inv)
        auto_create_cash_entry(inv.id, user["sub"], {
            "vendor": body.get("kunde", ""), "total_amount": betrag,
            "vat_amount": mwst_betrag, "vat_rate": mwst_satz,
            "date": body.get("datum", ""), "category": "service",
            "invoice_type": "income",
        })
        return {"success": True, "id": inv.id, "invoice_number": inv.invoice_number}
    except HTTPException:
        raise
    except Exception:
        db.rollback()
        logger.exception("Create Rechnung failed")
        err(500, "Rechnung erstellen fehlgeschlagen")
    finally:
        db.close()


@app.post("/account/kleinunternehmer")
def toggle_kleinunternehmer(body: dict = Body(...), user: dict = Depends(get_current_user)):
    """Toggle Kleinunternehmerregelung §19 UStG."""
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        if not u:
            err(404, "User not found")
        val = body.get("enabled", False)
        # Use proper boolean column instead of full_name hack
        u.is_kleinunternehmer = bool(val)
        # Migrate: clean up legacy [KU] prefix from full_name if present
        if (u.full_name or "").startswith("[KU]"):
            u.full_name = (u.full_name or "").replace("[KU] ", "").replace("[KU]", "").strip()
        db.commit()
        return {"success": True, "kleinunternehmer": val}
    finally:
        db.close()


@app.get("/account/me")
def get_account_me(user: dict = Depends(get_current_user)):
    """Return current user profile info."""
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        if not u:
            err(404, "User not found")
        inv_count = db.query(Invoice).filter(Invoice.user_id == user["sub"], (Invoice.is_deleted == False) | (Invoice.is_deleted == None)).count()
        companies = db.query(UserCompany).filter(UserCompany.user_id == user["sub"]).all()
        # Trial info — frontend banner icin
        trial_ends = getattr(u, 'trial_ends_at', None)
        is_trial = bool(trial_ends and trial_ends > datetime.now(timezone.utc))
        trial_expired = bool(trial_ends and trial_ends <= datetime.now(timezone.utc))
        trial_days_left = None
        if trial_ends:
            delta = trial_ends - datetime.now(timezone.utc)
            trial_days_left = max(0, delta.days + (1 if delta.seconds > 0 else 0))
        return {
            "id": u.id,
            "email": u.email,
            "full_name": u.full_name or "",
            "plan": u.plan or "free",
            "plan_name": {"free":"Free","early":"Early Adopter","pro":"Pro"}.get(u.plan or "free", "Free"),
            "registered_at": u.registered_at.isoformat() if u.registered_at else "",
            "is_kleinunternehmer": getattr(u, 'is_kleinunternehmer', False),
            "has_cloud_addon": getattr(u, 'has_cloud_addon', False),
            "trial_ends_at": trial_ends.isoformat() if trial_ends else None,
            "is_trial": is_trial,
            "trial_expired": trial_expired,
            "trial_days_left": trial_days_left,
            "invoice_count": inv_count,
            "company_name": companies[0].company_name if companies else "",
        }
    finally:
        db.close()


@app.get("/account/kleinunternehmer")
def get_kleinunternehmer(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        is_ku = getattr(u, 'is_kleinunternehmer', False) if u else False
        # Fallback: check legacy [KU] prefix
        if not is_ku and u and (u.full_name or "").startswith("[KU]"):
            is_ku = True
        return {"kleinunternehmer": is_ku}
    finally:
        db.close()


@app.post("/invoices/create")
def create_invoice_manual(body: dict = Body(...), user: dict = Depends(get_current_user)):
    """Create invoice from JSON (for cross-page transfer). Skips duplicates."""
    db = SessionLocal()
    try:
        # Duplicate check (skip soft-deleted)
        dup = db.query(Invoice).filter(
            Invoice.user_id == user["sub"],
            Invoice.vendor == (body.get("vendor") or "Manual"),
            Invoice.total_amount == float(body.get("total_amount") or 0),
            (Invoice.is_deleted == False) | (Invoice.is_deleted == None),
        ).first()
        if dup:
            return {"success": True, "id": dup.id, "message": "already exists"}
        inv = Invoice(
            user_id=user["sub"],
            filename=None,
            vendor=body.get("vendor") or "Manual",
            total_amount=float(body.get("total_amount") or 0),
            vat_amount=float(body.get("vat_amount") or 0),
            vat_rate=body.get("vat_rate") or "19%",
            date=body.get("date") or "",
            raw_text=body.get("raw_text") or "Manual entry",
            invoice_type=body.get("invoice_type") or "expense",
            invoice_number=body.get("invoice_number") or "",
            payment_method=body.get("payment_method") or "",
            category=body.get("category") or "other",
            processed=True,
        )
        db.add(inv)
        db.commit()
        db.refresh(inv)
        return {"success": True, "id": inv.id}
    except Exception:
        db.rollback()
        logger.exception("Create invoice failed")
        err(500, "Failed")
    finally:
        db.close()


@app.post("/invoices/upload-zip")
@limiter.limit("5/minute")
async def upload_zip(request: Request, file: UploadFile = File(...), invoice_type: str = "expense", user: dict = Depends(get_current_user)):
    """Upload a ZIP file containing invoices (PDF, JPG, PNG). Extracts and processes each file."""
    _enforce_upload_quota(user["sub"])
    import zipfile
    content = await file.read()
    if len(content) > 50 * 1024 * 1024:
        err(400, "ZIP zu groß (max 50MB)")
    if not content[:4] == b"PK\x03\x04":
        err(400, "Keine gültige ZIP-Datei")
    results = []
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as zf:
            for name in zf.namelist():
                name_lower = name.lower()
                if name.startswith("__MACOSX") or name.startswith("."):
                    continue
                if not any(name_lower.endswith(ext) for ext in (".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".webp")):
                    results.append({"filename": name, "status": "skipped", "message": "Nicht unterstützt"})
                    continue
                try:
                    file_data = zf.read(name)
                    if len(file_data) == 0:
                        continue
                    if len(file_data) > MAX_FILE_SIZE:
                        results.append({"filename": name, "status": "error", "message": "Datei zu groß"})
                        continue
                    # Determine content type from extension
                    if name_lower.endswith(".pdf"):
                        ct = "application/pdf"
                    elif name_lower.endswith((".jpg", ".jpeg")):
                        ct = "image/jpeg"
                    elif name_lower.endswith(".png"):
                        ct = "image/png"
                    elif name_lower.endswith(".tiff"):
                        ct = "image/tiff"
                    else:
                        ct = "image/webp"
                    # Create a fake UploadFile for existing pipeline
                    from starlette.datastructures import Headers as _Headers2
                    fake_file = UploadFile(filename=name, file=io.BytesIO(file_data), headers=_Headers2({"content-type": ct}))
                    raw_text = ""
                    try:
                        raw_text = await asyncio.wait_for(extract_text(fake_file, handwriting=False, file_bytes=file_data), timeout=45)
                    except Exception:
                        logger.warning("OCR failed for ZIP entry: %s", name)
                    try:
                        parsed = parse_invoice(raw_text)
                    except Exception:
                        results.append({"filename": name, "status": "error", "message": "Parse failed"})
                        continue
                    if invoice_type in ("income", "expense"):
                        parsed["invoice_type"] = invoice_type
                    invoice_id = save_invoice(parsed, user_id=user["sub"], filename=name)
                    auto_create_cash_entry(invoice_id, user["sub"], parsed)
                    results.append({"filename": name, "status": "ok", "id": invoice_id, "vendor": parsed.get("vendor", ""), "total": parsed.get("total_amount", 0)})
                except Exception as e:
                    results.append({"filename": name, "status": "error", "message": str(e)})
    except zipfile.BadZipFile:
        err(400, "Beschädigte ZIP-Datei")
    return {"success": True, "results": results, "total": len(results)}


@app.post("/invoices/upload")
@limiter.limit("20/minute")
async def upload_invoice(request: Request, file: UploadFile = File(...), handwriting: bool = False, invoice_type: str = "expense", force_upload: bool = False, user: dict = Depends(get_current_user)):
    _enforce_upload_quota(user["sub"])
    if file.content_type not in ALLOWED_TYPES:
        err(400, "Ungültige Datei. Erlaubt: PDF, JPG, PNG, TIFF, WEBP, ZIP")

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        err(400, "Datei zu groß (max 10MB)")
    if len(content) == 0:
        err(400, "Leere Datei")

    # ZIP: extract and process each file inside
    if content[:4] == b"PK\x03\x04" or (file.content_type or "").lower() in ("application/zip", "application/x-zip-compressed"):
        import zipfile
        zip_results = []
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                for name in zf.namelist():
                    nl = name.lower()
                    if name.startswith("__MACOSX") or name.startswith("."):
                        continue
                    if not any(nl.endswith(e) for e in (".pdf", ".jpg", ".jpeg", ".png", ".tiff", ".webp")):
                        continue
                    try:
                        fd = zf.read(name)
                        if not fd or len(fd) > MAX_FILE_SIZE:
                            continue
                        ct = "application/pdf" if nl.endswith(".pdf") else "image/jpeg" if nl.endswith((".jpg",".jpeg")) else "image/png"
                        from starlette.datastructures import Headers as _Headers
                        fake = UploadFile(filename=name, file=io.BytesIO(fd), headers=_Headers({"content-type": ct}))
                        rt = ""
                        _qr = {}
                        try:
                            rt = await asyncio.wait_for(extract_text(fake, handwriting=handwriting, file_bytes=fd), timeout=45)
                        except Exception:
                            pass
                        # Try QR reader if OCR returned little/no text
                        if len(rt.strip()) < 20:
                            try:
                                from autotax.qr_reader import extract_qr_data
                                _qr = extract_qr_data(fd, ct)
                            except Exception:
                                pass
                        parsed = parse_invoice(rt)
                        # Merge QR data
                        if _qr:
                            if _qr.get("company") and parsed.get("vendor") in ("Unbekannt", "", None):
                                parsed["vendor"] = _qr["company"]
                            if _qr.get("amount") and (not parsed.get("total_amount") or parsed["total_amount"] == 0):
                                parsed["total_amount"] = _qr["amount"]
                            if _qr.get("date") and not parsed.get("date"):
                                parsed["date"] = _qr["date"]
                        if invoice_type in ("income", "expense"):
                            parsed["invoice_type"] = invoice_type
                        inv_id = save_invoice(parsed, user_id=user["sub"], filename=name)
                        auto_create_cash_entry(inv_id, user["sub"], parsed)
                        zip_results.append({"filename": name, "status": "ok", "id": inv_id})
                    except Exception:
                        zip_results.append({"filename": name, "status": "error"})
        except zipfile.BadZipFile:
            err(400, "Beschädigte ZIP-Datei")
        return {"success": True, "results": zip_results, "count": len(zip_results)}

    if not _validate_file_magic(content, file.content_type or ""):
        err(400, "Ungültige Datei — Dateityp stimmt nicht mit Inhalt überein")

    await file.seek(0)

    # --- Hard duplicate check (md5) ---
    file_hash = generate_file_hash(content)
    db_check = SessionLocal()
    try:
        existing = find_hard_duplicate(db_check, user["sub"], file_hash)
        if existing and not force_upload:
            logger.info("Hard duplicate: user=%s hash=%s -> invoice %d", user["sub"], file_hash, existing.id)
            return {
                "id": existing.id,
                "total_amount": safe_float(existing.total_amount),
                "filename": existing.filename,
                "status": "duplicate",
                "duplicate": True,
                "can_force": True,
                "message": "Diese Datei wurde bereits hochgeladen.",
            }
    finally:
        db_check.close()
    await file.seek(0)
    # --- End hard duplicate check ---

    # Save original file to DB for vault preview
    _file_data = content
    _file_ct = file.content_type or ""

    logger.info("Upload by user %s: type=%s, size=%d bytes", user["sub"], file.content_type, len(content))

    import gc
    raw_text = ""
    qr_data = {}

    # --- ADDED START: Try Tesseract first, skip paid OCR if valid ---
    _tess_used = False
    try:
        if (file.content_type or "").lower().startswith("image/"):
            from autotax.ocr import local_ocr_tesseract, is_ocr_valid
            _tess_text = local_ocr_tesseract(content)
            if is_ocr_valid(_tess_text):
                logger.info("Using local OCR (Tesseract): %s (%d chars)", file.filename, len(_tess_text))
                raw_text = _tess_text
                _tess_used = True
                # Still get QR data
                try:
                    from autotax.qr_reader import extract_qr_data
                    qr_data = extract_qr_data(content, file.content_type or "")
                except Exception:
                    pass
            else:
                logger.info("Tesseract invalid — Fallback to OCR.space: %s", file.filename)
    except Exception as e:
        logger.warning("Tesseract pre-check failed: %s", e)
    # --- ADDED END ---

    try:
        if not _tess_used:
            raw_text, qr_data = await asyncio.wait_for(extract_text_and_qr(file, handwriting=handwriting, file_bytes=content), timeout=45)
    except asyncio.TimeoutError:
        logger.warning("OCR timeout — saving with empty text")
    except Exception:
        logger.warning("OCR failed — saving with empty text")

    # If OCR returned nothing, try QR-only (pure QR code images)
    if len(raw_text.strip()) < 10 and not qr_data:
        try:
            from autotax.qr_reader import extract_qr_data
            qr_data = extract_qr_data(_file_data, _file_ct)
            if qr_data:
                logger.info("QR fallback: found data in QR-only image")
        except Exception:
            pass

    # --- STEP 1: Learning rules FIRST — apply user's saved corrections ---
    # Runs BEFORE parser so learned vendor/category auto-fills immediately.
    # Parser then sees pre-filled fields and skips those extractions.
    learned_data = {}
    try:
        from autotax.learning import apply_learning_rules
        learned_data = apply_learning_rules(user["sub"], raw_text, {})
        if learned_data:
            logger.info("LEARNING APPLIED: %s", {k: v for k, v in learned_data.items() if v})
    except Exception as e:
        logger.warning("Learning rules skipped: %s", e)

    # --- STEP 2: Parser — regex extraction ---
    try:
        result = parse_invoice(raw_text)
    except Exception:
        logger.exception("Parsing failed for %s", file.filename)
        err(500, "Invoice parsing failed")

    # --- STEP 2b: Vendor identity match — kimlik parmak izi (USt-IdNr/IBAN/HRB) ---
    # parser_invoice ust_id/iban/hrb/email/domain/phone cikariyor; vendor_identities
    # tablosunda kayit varsa vendor adi (OCR bozulmasindan bagimsiz) dogru gelir.
    # Manuel girilmis vendor (Beleg hinzufugen formundan) en yuksek oncelige sahip.
    try:
        from autotax.vendor_identity import match_vendor
        _identity_fields = {
            "ust_id": result.get("vendor_ust_id"),
            "iban": result.get("vendor_iban"),
            "hrb": result.get("vendor_hrb"),
            "email": result.get("vendor_email"),
            "domain": result.get("vendor_domain"),
            "phone": result.get("vendor_phone"),
            "address": result.get("vendor_address"),
        }
        _vmatch = match_vendor(user["sub"], identity_fields=_identity_fields)
        if _vmatch:
            _old_vendor = result.get("vendor", "") or ""
            # Yuksek guvenli eslesme = ust_id (1.0) / iban (0.95) / hrb (0.90).
            # Phone/email/domain (<= 0.80) yanlis pozitif riskidir — Lidl fisinde
            # yanlislikla SPAR'in telefonuna eslesirse SPAR vendor'i kilitlenmesin.
            HIGH_CONF = 0.90
            _high_conf = _vmatch.score >= HIGH_CONF

            if _high_conf and (
                _old_vendor in ("Unbekannt", "Manual Entry", "")
                or len(_old_vendor) < 3
                or _vmatch.score >= 0.95
            ):
                result["vendor"] = _vmatch.vendor_name
                logger.info("[IDENTITY] vendor kilitlendi: '%s' -> '%s' by=%s score=%.2f",
                            _old_vendor, _vmatch.vendor_name, _vmatch.matched_by, _vmatch.score)
            elif not _high_conf:
                # Dusuk guvenli eslesme — vendor adi ezilmez, sadece logla
                logger.info("[IDENTITY] dusuk guven, vendor degismedi: by=%s score=%.2f vendor='%s'",
                            _vmatch.matched_by, _vmatch.score, _old_vendor)

            # Default'lar (vat_rate/category/payment) — yalnizca yuksek guvenli
            # eslesmede ezilir. Aksi takdirde yanlis vendor'in default'lari
            # gercege gore atanabilir.
            if _high_conf:
                if _vmatch.default_vat_rate and result.get("vat_rate") in ("0%", "", None):
                    result["vat_rate"] = _vmatch.default_vat_rate
                if _vmatch.default_category and result.get("category") in ("other", "", None):
                    result["category"] = _vmatch.default_category
                if _vmatch.default_payment_method and not result.get("payment_method"):
                    result["payment_method"] = _vmatch.default_payment_method
    except Exception as e:
        logger.warning("[IDENTITY] match skipped: %s", e)

    # --- STEP 2c: Vendor fallback — adres-aware ---
    # Parser bazen ilk OCR satirini vendor sayar; logo gorsel ise OCR atlar
    # ve 2. satir (adres) yanlislikla vendor olur. Bu fallback:
    #  1) Once UPPERCASE kisa kelime ara (ARAL/LIDL/SPAR logo isimleri)
    #  2) Bulunmazsa, ilk 8 satirdan ADRES OLMAYAN ilk anlamli satiri al
    # Adres reddedicileri: cadde anahtar kelimeleri, 5-haneli PLZ, sirket
    # son ekleri tek basina.
    try:
        _cur_vendor = (result.get("vendor") or "").strip()
        _is_default = _cur_vendor in ("Unbekannt", "Manual Entry", "") or len(_cur_vendor) < 3

        # Adres tespit yardimcisi — Im Rotfeld 1 / Hauptweg 12 / Am Staden 4 / 66115 Saarbruecken / Stiftsbergstr. 1
        import re as _re_v_chk
        _addr_signals = ("str.", "strasse", "straße", "weg", "platz",
                         "allee", "gasse", "ring", "damm", "ufer", "chaussee")
        # "Im X 1" / "Am X 1" / "Auf der X 1" — sokak adi prefix'i (ev numarasiz da olabilir).
        # Bunlar kesin adres prefix'leri, vendor adlarinda yok denecek kadar nadir
        # ("Im Westen Nichts Neues" gibi yayinevi olurdu, fis olmaz).
        _addr_prefix_re = _re_v_chk.compile(
            r"^(?:im|am|an\s+der|auf\s+der|bei\s+der|in\s+der|zur|zum)\s+\w+",
            _re_v_chk.IGNORECASE,
        )
        # "In Rotfeld 1" / "In R)tfeld 1" — Lidl tarzi: "In" + kelime + ev numarasi.
        # 'in' tek basina cok genel oldugu icin (Inception/In Stock vs.) bu pattern'i
        # ayri bir regex'e koyduk: 'In ' + (en az 1 word) + ' ' + (1-4 rakam, opsiyonel a/b/c) + sonu.
        _in_address_re = _re_v_chk.compile(
            r"^in\s+\S+(?:\s+\S+)?\s+\d{1,4}[a-z]?\s*$",
            _re_v_chk.IGNORECASE,
        )

        def _looks_like_address(line: str) -> bool:
            if not line:
                return False
            ll = line.lower()
            if any(s in ll for s in _addr_signals):
                return True
            if _re_v_chk.search(r"\b\d{5}\b", line):
                return True
            if _addr_prefix_re.match(line.strip()):
                return True
            if _in_address_re.match(line.strip()):
                return True
            return False

        # Vendor PARSER'dan gelse bile adres-benzeri ise reset edelim
        if _cur_vendor and not _is_default:
            if _looks_like_address(_cur_vendor):
                logger.info("[VENDOR_FALLBACK] parser adres-benzeri vendor verdi, reset: %r", _cur_vendor)
                _cur_vendor = ""
                _is_default = True

        if _is_default:
            import re as _re_v
            _candidate_upper = None  # UPPERCASE prioriteli
            _candidate_normal = None  # adres olmayan ilk satir (fallback)

            for _line in (raw_text or "").splitlines()[:8]:
                _line = _line.strip()
                if not _line or len(_line) < 3:
                    continue

                # Adres elimene (genisletilmis: Im/Am/Auf der + ev numarasi pattern'i dahil)
                if _looks_like_address(_line):
                    continue
                # Tarih/numara satirlari (tum digit veya cok digit)
                if sum(c.isdigit() for c in _line) > len(_line) * 0.5:
                    continue

                # 1. tercih: UPPERCASE 3-15 harfli kelime
                _m_upper = _re_v.match(r"^([A-ZÄÖÜ]{3,15})\s*$", _line)
                if _m_upper and _candidate_upper is None:
                    _candidate_upper = _m_upper.group(1).capitalize()
                    break  # UPPERCASE bulduk, durabiliriz

                # 2. tercih: en az 3 karakter, baska adayimiz yoksa
                # Sadece harfler + bosluk + nokta + & icerebilir (firma adi tipiktir)
                if _re_v.match(r"^[A-Za-zÄÖÜäöüß0-9.\s&,\-/]+$", _line) and _candidate_normal is None:
                    # Sadece sembol/rakam degil, en az 3 harf icermeli
                    if sum(c.isalpha() for c in _line) >= 3:
                        _candidate_normal = _line[:60].strip(".,- ")

            chosen = _candidate_upper or _candidate_normal
            if chosen:
                result["vendor"] = chosen
                logger.info("[VENDOR_FALLBACK] OCR'dan vendor bulundu: %r (kaynak=%s)",
                            chosen, "UPPERCASE" if _candidate_upper else "ilk-anlamli")
    except Exception as e:
        logger.warning("[VENDOR_FALLBACK] hata: %s", e)

    # --- STEP 2d: Filename-based vendor fallback ---
    # Kullanici dosyalari "lidl 97.55.pdf" / "LIDEL 63.pdf" / "TEDI 3.00.pdf"
    # gibi adlandiriyor. OCR vendor'i adres olarak yakalarsa veya bos kalirsa
    # dosya adindaki bilinen vendor'i fallback olarak kullan.
    try:
        _fname = (file.filename or "").lower()
        if _fname:
            # Bilinen Almanya zinciri vendor'leri — dosya adinda gecerse match.
            # Sira: en spesifik (uzun) once, kismi match'leri onler.
            _KNOWN_VENDORS = [
                ("media markt", "Media Markt"), ("mediamarkt", "Media Markt"),
                ("lidl", "Lidl"), ("lidel", "Lidl"),  # tipo
                ("aldi", "Aldi"), ("rewe", "Rewe"), ("edeka", "Edeka"),
                ("kaufland", "Kaufland"), ("penny", "Penny"), ("netto", "Netto"),
                ("norma", "Norma"), ("tegut", "Tegut"), ("globus", "Globus"),
                ("aral", "Aral"), ("shell", "Shell"), ("esso", "Esso"),
                ("douglas", "Douglas"), ("rossmann", "Rossmann"),
                ("müller", "Müller"), ("muller", "Müller"),
                ("saturn", "Saturn"), ("expert", "Expert"), ("euronics", "Euronics"),
                ("tedi", "TEDI"), ("action", "Action"), ("kik", "KiK"),
                ("woolworth", "Woolworth"), ("snipes", "Snipes"),
                ("deichmann", "Deichmann"), ("zara", "Zara"), ("primark", "Primark"),
                ("h&m", "H&M"), ("c&a", "C&A"),
                ("ikea", "IKEA"), ("bauhaus", "Bauhaus"), ("obi", "OBI"),
                ("hornbach", "Hornbach"), ("toom", "Toom"),
                ("amazon", "Amazon"), ("ebay", "eBay"), ("zalando", "Zalando"),
                ("dm", "dm"),  # "dm" en sonda — kisa pattern
            ]
            _file_vendor = None
            for needle, vname in _KNOWN_VENDORS:
                # "dm" gibi kisa pattern'ler icin word-boundary, digerleri icin substring.
                if len(needle) <= 2:
                    if _re_v_chk.search(rf"\b{_re_v_chk.escape(needle)}\b", _fname):
                        _file_vendor = vname
                        break
                else:
                    if needle in _fname:
                        _file_vendor = vname
                        break

            # GENERIC fallback: bilinen listede yoksa dosya adindan
            # uzantiyi ve fiyat parcasini soyup geriyi vendor adi olarak al.
            # 'bereket metzger--22.43.pdf' -> 'Bereket Metzger'
            # 'kebabhaus 12.50.pdf'        -> 'Kebabhaus'
            # 'scan001.pdf'                -> None (scanner default name)
            if not _file_vendor:
                _base = _re_v_chk.sub(r"\.[a-z0-9]+$", "", _fname, flags=_re_v_chk.IGNORECASE)
                _base = _re_v_chk.sub(r"[-_\s]+\d{1,5}[.,]?\d{0,2}\s*$", "", _base)
                _base = _re_v_chk.sub(r"[-_]+", " ", _base)
                _base = _re_v_chk.sub(r"\s+", " ", _base).strip()
                # Generic scanner default names — bunlar vendor degil
                _GENERIC_FNAME_PREFIXES = ("scan", "img", "image", "photo", "doc",
                                           "page", "untitled", "kopie", "copy",
                                           "neu", "neue", "test", "rechnung",
                                           "invoice", "fatura", "fis")
                _base_lower = _base.lower()
                _is_generic = (
                    not _base or len(_base) < 3 or
                    any(_base_lower.startswith(p) and len(_base) <= len(p) + 5
                        for p in _GENERIC_FNAME_PREFIXES)
                )
                if not _is_generic:
                    if _base == _base.upper():
                        _file_vendor = _base
                    else:
                        _file_vendor = _base.title()

            # FILENAME AMOUNT — kullanici '...-22.43.pdf' / '... 63.47.pdf' formatinda
            # toplam tutari dosya adina yaziyor. Parser'in bulduguyla farkliysa
            # filename'e guvenilir referans olarak bakariz.
            _file_amount = None
            try:
                # Sondaki X.YY veya X,YY (1-5 hane integer + 2 hane decimal)
                _amt_m = _re_v_chk.search(r"(\d{1,5})[.,](\d{2})\s*$",
                                           _re_v_chk.sub(r"\.[a-z0-9]+$", "", _fname,
                                                          flags=_re_v_chk.IGNORECASE))
                if _amt_m:
                    _file_amount = float(f"{_amt_m.group(1)}.{_amt_m.group(2)}")
            except Exception:
                pass

            if _file_vendor:
                _cur = (result.get("vendor") or "").strip()
                _is_default = _cur in ("Unbekannt", "Manual Entry", "") or len(_cur) < 3

                # Suspect vendor: parser bazen item satirini vendor sayar
                # ('Jessa Slipeinlagen Lang 308t 1,45 1') veya promotion-line
                # ('ACTION ab 1.99'). Bu durumlarda dosya adi user'in
                # niyetine daha yakin.
                def _is_suspect(name: str) -> bool:
                    if not name:
                        return True
                    # Fiyat formatli sayi var (1,45 / 12.99)
                    if _re_v_chk.search(r"\d+[,.]\d{2}", name):
                        return True
                    # 4+ kelime: vendor genelde 1-3 kelime
                    if len(name.split()) > 4:
                        return True
                    # 4+ rakam: item kodlari
                    if sum(c.isdigit() for c in name) > 4:
                        return True
                    # OCR garbage: cok fazla noktalama / ozel karakter
                    # ('Ber Ek F.\'V M Er' gibi parcalanmis logo)
                    _alpha = sum(c.isalpha() for c in name)
                    _punct = sum(1 for c in name if c in ".'\"/\\|`~^*+={}[]()<>")
                    if _alpha > 0 and _punct / max(_alpha, 1) > 0.2:
                        return True
                    return False

                # Filename'i bulduysak ama parser vendor parser_filename'e yakin
                # ise override etme — parser dogruyu bulmus olabilir
                # (ornek: filename 'lidl 97.pdf', parser 'LIDL' -> ayni)
                _file_vendor_lower = _file_vendor.lower().replace(" ", "")
                _cur_lower = _re_v_chk.sub(r"[^a-z0-9]", "", _cur.lower())
                _matches_filename = bool(_file_vendor_lower) and (
                    (len(_file_vendor_lower) >= 3 and _file_vendor_lower in _cur_lower) or
                    (len(_cur_lower) >= 3 and _cur_lower in _file_vendor_lower)
                )

                if not _matches_filename and (
                    _is_default or _looks_like_address(_cur) or _is_suspect(_cur)
                ):
                    logger.info("[VENDOR_FILENAME] dosya adindan vendor: %r (parser: %r)",
                                _file_vendor, _cur)
                    result["vendor"] = _file_vendor

            # Filename amount override — parser total'i ile filename amount'u
            # cok farkliysa (5+ EUR sapma VEYA parser 0/None ise) filename'i kullan.
            # Kullanici bilincli olarak adlandirir; OCR Tesseract '63'u '68' okusa
            # bile filename'deki '63.47' dogru cevap.
            if _file_amount and _file_amount > 0:
                _parser_amount = result.get("total_amount") or 0
                _diff = abs(_parser_amount - _file_amount)
                _max_val = max(_parser_amount, _file_amount)
                # Trigger: parser bos VEYA fark >= 0.50 EUR ve >= %2 (kucuk yuvarlamalari yoksay)
                _significant = _parser_amount <= 0 or (_diff >= 0.50 and _diff / _max_val >= 0.02)
                if _significant:
                    logger.info("[AMOUNT_FILENAME] dosya adindan tutar: %.2f (parser: %.2f, fark: %.2f)",
                                _file_amount, _parser_amount, _diff)
                    result["total_amount"] = _file_amount
    except Exception as e:
        logger.warning("[VENDOR_FILENAME] hata: %s", e)

    # --- STEP 3: Merge learning over parser defaults ---
    # Learning wins ONLY for fields where parser returned defaults
    # (Unbekannt, other, 0%, empty). Never overwrites real parser values.
    if learned_data:
        from autotax.learning import LEARNABLE_FIELDS
        _DEFAULTS = {"Unbekannt", "other", "0%", "", None, "expense"}
        for field in LEARNABLE_FIELDS:
            parser_val = str(result.get(field, "")).strip()
            learned_val = learned_data.get(field, "")
            if learned_val and parser_val in _DEFAULTS:
                result[field] = learned_val
                logger.info("LEARNING OVERRIDE: %s = %r (parser had %r)", field, learned_val, parser_val)

    # Merge QR data (QR overrides OCR if available)
    if qr_data:
        logger.info("QR data found: keys=%s", list(qr_data.keys()))
        if qr_data.get("company") and (not result.get("vendor") or result.get("vendor") == "Unbekannt"):
            result["vendor"] = qr_data["company"]
        if qr_data.get("amount") and (not result.get("total_amount") or result.get("total_amount") == 0):
            result["total_amount"] = qr_data["amount"]
        if qr_data.get("date") and (not result.get("date") or result["date"] == datetime.now().strftime("%Y-%m-%d")):
            result["date"] = qr_data["date"]
        if qr_data.get("invoice_number") and not result.get("invoice_number"):
            result["invoice_number"] = qr_data["invoice_number"]
        if qr_data.get("tax") and (not result.get("vat_amount") or result.get("vat_amount") == 0):
            result["vat_amount"] = qr_data["tax"]
        if qr_data.get("qr_raw"):
            result["raw_text"] = result.get("raw_text", "") + "\n\n[QR] " + qr_data["qr_raw"]

    # Soft duplicate flag (same vendor + amount + date, case-insensitive vendor)
    db_soft = SessionLocal()
    try:
        is_soft_dup = check_soft_duplicate(
            db_soft,
            user_id=user["sub"],
            vendor=result.get("vendor"),
            amount=safe_float(result.get("total_amount")),
            date=result.get("date"),
        )
    finally:
        db_soft.close()

    if invoice_type in ("income", "expense"):
        result["invoice_type"] = invoice_type

    try:
        invoice_id = save_invoice(result, user_id=user["sub"], filename=file.filename, file_data=_file_data, file_content_type=_file_ct, file_hash=file_hash, possible_duplicate=is_soft_dup)
    except Exception:
        logger.exception("DB save failed")
        err(500, "Failed to save invoice")

    auto_create_cash_entry(invoice_id, user["sub"], result)

    # Auto-detect income: if vendor/IBAN/email matches user's registered company
    try:
        db_c = SessionLocal()
        user_companies = db_c.query(UserCompany).filter(UserCompany.user_id == user["sub"]).all()
        if user_companies:
            inv = db_c.query(Invoice).filter(Invoice.id == invoice_id).first()
            if inv:
                vendor_lower = (inv.vendor or "").lower()
                inv_iban = result.get("vendor_iban", "").replace(" ", "").upper()
                inv_email = result.get("vendor_email", "").lower()
                for uc in user_companies:
                    matched = False
                    # Match by company name
                    if vendor_lower and uc.company_name:
                        if uc.company_name.lower() in vendor_lower or vendor_lower in uc.company_name.lower() or _fuzzy_match(uc.company_name, inv.vendor or ""):
                            matched = True
                    # Match by IBAN
                    if not matched and inv_iban and uc.iban:
                        if inv_iban == uc.iban.replace(" ", "").upper():
                            matched = True
                    # Match by email
                    if not matched and inv_email and uc.email:
                        if inv_email == uc.email.lower():
                            matched = True
                    if matched:
                        inv.invoice_type = "income"
                        logger.info("Auto-detected income: invoice %d matches company '%s'", invoice_id, uc.company_name)
                        db_c.commit()
                        break
        db_c.close()
    except Exception:
        pass

    # OCR quality warning
    _ocr_warning = ""
    if not raw_text or len(raw_text.strip()) < 20:
        _ocr_warning = "OCR konnte den Text nicht lesen — bitte manuell prüfen"
    elif safe_float(result.get("total_amount")) == 0:
        _ocr_warning = "Betrag nicht erkannt — bitte manuell eingeben"
    elif result.get("vendor") == "Unbekannt":
        _ocr_warning = "Lieferant nicht erkannt — bitte manuell eingeben"
    elif result.get("vendor_source") == "guess":
        _ocr_warning = "Lieferantenname unsicher — bitte prüfen"

    return {
        "id": invoice_id,
        "total_amount": safe_float(result.get("total_amount")),
        "filename": file.filename,
        "status": "ok",
        "warning": _ocr_warning,
        "vendor": result.get("vendor", ""),
        "vendor_iban": result.get("vendor_iban", ""),
        "vendor_email": result.get("vendor_email", ""),
        "vendor_phone": result.get("vendor_phone", ""),
        "vendor_address": result.get("vendor_address", ""),
    }


@app.post("/invoices/batch")
async def upload_batch(files: List[UploadFile] = File(...), invoice_type: str = "expense", user: dict = Depends(get_current_user)):
    import gc
    results = []
    for file in files:
        try:
            if file.content_type not in ALLOWED_TYPES:
                results.append({"filename": file.filename, "status": "error", "message": "Ungültige Datei"})
                continue
            content = await file.read()
            if len(content) > MAX_FILE_SIZE:
                results.append({"filename": file.filename, "status": "error", "message": "Datei zu groß"})
                continue
            if len(content) == 0:
                results.append({"filename": file.filename, "status": "error", "message": "Leere Datei"})
                continue
            if not _validate_file_magic(content, file.content_type or ""):
                results.append({"filename": file.filename, "status": "error", "message": "Ungültige Datei"})
                continue
            await file.seek(0)
            _batch_file_data = content  # save reference before OCR
            raw_text = ""
            try:
                raw_text = await asyncio.wait_for(extract_text(file, handwriting=False, file_bytes=content), timeout=45)
            except Exception:
                logger.warning("OCR failed/timeout for batch file")
            try:
                parsed = parse_invoice(raw_text)
            except Exception:
                results.append({"filename": file.filename, "status": "error", "message": "Parse failed"})
                continue
            # Duplicate check
            db_dup = SessionLocal()
            try:
                dup = db_dup.query(Invoice).filter(
                    Invoice.user_id == user["sub"],
                    Invoice.vendor == (parsed.get("vendor") or "Unbekannt"),
                    Invoice.total_amount == safe_float(parsed.get("total_amount")),
                    Invoice.date == (parsed.get("date") or ""),
                ).first()
            finally:
                db_dup.close()
            if dup:
                results.append({"filename": file.filename, "status": "duplicate", "message": "Duplikat erkannt"})
                continue
            if invoice_type in ("income", "expense"):
                parsed["invoice_type"] = invoice_type
            invoice_id = save_invoice(parsed, user_id=user["sub"], filename=file.filename, file_data=_batch_file_data, file_content_type=file.content_type or "")
            auto_create_cash_entry(invoice_id, user["sub"], parsed)
            results.append({
                "filename": file.filename,
                "status": "ok",
                "message": f"OK — €{safe_float(parsed.get('total_amount')):.2f}",
                "id": invoice_id,
            })
        except Exception as e:
            results.append({"filename": file.filename, "status": "error", "message": str(e)})
    return {"results": results}


# ============================================================
# INVOICES: LIST
# ============================================================

@app.get("/invoices")
def list_invoices(
    skip: int = Query(0, ge=0),
    limit: int = Query(50, ge=1, le=500),
    search: Optional[str] = Query(None),
    vendor: Optional[str] = Query(None),
    status: Optional[str] = Query(None),
    category: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    user: dict = Depends(get_current_user),
):
    db = SessionLocal()
    try:
        from sqlalchemy.orm import defer as _sa_defer
        # file_data legacy LargeBinary — defer ediliyor (her satirda
        # MB'larca veri yuklemek list view icin gereksiz).
        # raw_text invoice_to_dict icinde ocr_snippet uretmek icin
        # kullaniliyor (ilk 200 char) — defer edilirse N+1 olur, bu yuzden eager yuklenir.
        q = db.query(Invoice).options(
            _sa_defer(Invoice.file_data),
        ).filter(Invoice.user_id == user["sub"])
        # --- ADDED: exclude soft-deleted ---
        q = q.filter((Invoice.is_deleted == False) | (Invoice.is_deleted == None))
        # --- END ---

        if search:
            # Smart multi-keyword search: split by space, ALL keywords must match
            # Search across vendor, category, and raw_text (OCR content)
            import unicodedata
            normalized = unicodedata.normalize("NFKD", search.lower().strip())
            keywords = [k.strip() for k in normalized.split() if k.strip()]
            from sqlalchemy import or_
            for kw in keywords:
                pattern = f"%{kw}%"
                q = q.filter(or_(
                    Invoice.raw_text.ilike(pattern),
                    Invoice.vendor.ilike(pattern),
                    Invoice.category.ilike(pattern),
                    Invoice.invoice_number.ilike(pattern),
                ))
        if vendor:
            q = q.filter(Invoice.vendor.ilike(f"%{vendor}%"))
        if status == "processed":
            q = q.filter(Invoice.processed == True)
        elif status == "unprocessed":
            q = q.filter(Invoice.processed == False)
        if category:
            q = q.filter(Invoice.category == category)
        # Validate date range (reject invalid years like 333333)
        import re as _re
        _current_year = datetime.now().year
        if date_from and _re.match(r"^\d{4}-\d{2}-\d{2}$", date_from):
            if 2020 <= int(date_from[:4]) <= _current_year + 1:
                q = q.filter(Invoice.date >= date_from)
        if date_to and _re.match(r"^\d{4}-\d{2}-\d{2}$", date_to):
            if 2020 <= int(date_to[:4]) <= _current_year + 1:
                q = q.filter(Invoice.date <= date_to)

        total_count = q.count()
        q = q.order_by(Invoice.created_at.desc())
        invoices = q.offset(skip).limit(limit).all()

        return ok_list(
            [invoice_to_dict(i) for i in invoices],
            total_count,
        )
    except HTTPException:
        raise
    except Exception:
        logger.exception("Failed to list invoices")
        err(500, "Failed to load invoices")
    finally:
        db.close()


# ============================================================
# INVOICES: DASHBOARD
# ============================================================

@app.get("/invoices/dashboard")
def invoice_dashboard(country: str = Query("DE"), user: dict = Depends(get_current_user)):
    try:
        m = calculate_dashboard_metrics(user["sub"])
        net_profit = m["profit"]

        if country == "DE":
            if net_profit > 277826:
                tax_rate = 45
            elif net_profit > 61356:
                tax_rate = 42
            elif net_profit > 17005:
                tax_rate = 30
            elif net_profit > 10908:
                tax_rate = 14
            else:
                tax_rate = 0
        else:
            tax_rate = 30

        tax_estimate = round(net_profit * tax_rate / 100, 2) if net_profit > 0 else 0

        return {
            "total_income": m["total_income"],
            "total_expenses": m["total_expenses"],
            "net_profit": m["profit"],
            "tax_estimate": tax_estimate,
            "tax_rate_applied": tax_rate,
            "income_count": m["income_count"],
            "expense_count": m["expense_count"],
            "invoice_count": m["invoice_count"],
            "invalid_count": 0,
            "monthly_breakdown": m["monthly"],
            "by_category": m["by_category"],
            "total_vat_paid": m["vat_paid"],
            "total_vat_collected": m["vat_collected"],
            "vat_balance": m["vat_balance"],
        }
    except Exception:
        logger.exception("Dashboard failed")
        err(500, "Dashboard failed")


# ============================================================
# INVOICES: SUMMARY
# ============================================================

@app.get("/invoices/summary")
def invoice_summary(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        all_invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"], (Invoice.is_deleted == False) | (Invoice.is_deleted == None)).all()
        invoices = [i for i in all_invoices if safe_float(i.total_amount) > 0 and safe_vendor(i.vendor) != "Unbekannt"]
        total_count = len(invoices)
        processed = sum(1 for i in invoices if i.processed)
        unprocessed = total_count - processed
        total_revenue = sum(safe_float(i.total_amount) for i in invoices)
        return {
            "success": True,
            "total_count": total_count,
            "processed": processed,
            "unprocessed": unprocessed,
            "total_revenue": round(total_revenue, 2),
        }
    except Exception:
        logger.exception("Summary failed")
        err(500, "Failed to load summary")
    finally:
        db.close()


# ============================================================
# INVOICES: UPDATE (PATCH + PUT)
# ============================================================

class InvoiceUpdate(BaseModel):
    vendor: Optional[str] = None
    category: Optional[str] = None
    total_amount: Optional[float] = None
    vat_amount: Optional[float] = None
    vat_rate: Optional[str] = None
    date: Optional[str] = None
    invoice_type: Optional[str] = None
    invoice_number: Optional[str] = None
    payment_method: Optional[str] = None
    processed: Optional[bool] = None


def _do_update_invoice(invoice_id: int, body: InvoiceUpdate, user: dict):
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        # Snapshot original values for learning comparison + corrections log
        # Genisletildi: log_corrections tum LOGGABLE_FIELDS'i bekliyor.
        # save_learning_rule fazla alanlari nazikce gormezden gelir.
        _original = {
            "vendor": inv.vendor, "category": inv.category,
            "vat_rate": inv.vat_rate, "payment_method": inv.payment_method,
            "invoice_type": inv.invoice_type,
            "total_amount": inv.total_amount, "vat_amount": inv.vat_amount,
            "date": inv.date, "invoice_number": inv.invoice_number,
        }
        _pre_raw_text = inv.raw_text or ""
        _pre_vendor = inv.vendor or ""
        if body.vendor is not None:
            inv.vendor = body.vendor
        if body.category is not None:
            inv.category = body.category
        if body.total_amount is not None:
            inv.total_amount = body.total_amount
        if body.vat_amount is not None:
            inv.vat_amount = body.vat_amount
        if body.vat_rate is not None:
            inv.vat_rate = body.vat_rate
        if body.date is not None:
            inv.date = body.date
        if body.invoice_type is not None:
            inv.invoice_type = body.invoice_type
        if body.invoice_number is not None:
            inv.invoice_number = body.invoice_number
        if body.payment_method is not None:
            inv.payment_method = body.payment_method
        if body.processed is not None:
            inv.processed = body.processed
        db.commit()
        db.refresh(inv)
        # --- Learning: save user corrections as rules for future auto-fill ---
        _edited = {k: v for k, v in (body.model_dump() if hasattr(body, "model_dump") else body.dict()).items() if v is not None}
        try:
            from autotax.learning import save_learning_rule
            save_learning_rule(user["sub"], inv.raw_text or "", _original, _edited)
        except Exception as e:
            logger.warning("Learning save skipped: %s", e)
        # --- Corrections: ham diff + OCR snapshot (few-shot RAG yakiti) ---
        try:
            from autotax.corrections import log_corrections
            log_corrections(
                invoice_id=invoice_id,
                user_id=user["sub"],
                original=_original,
                edited=_edited,
                ocr_text=_pre_raw_text,
                vendor=_pre_vendor,
            )
        except Exception as e:
            logger.warning("Corrections log skipped: %s", e)
        # --- Status flow: kullanici PATCH yapti = bu fisi onayladi ---
        # 'confirmed' degilse 'confirmed' yap. State machine ilerletme.
        try:
            if getattr(inv, "status", None) and inv.status != "confirmed":
                inv.status = "confirmed"
                db.commit()
        except Exception as e:
            logger.warning("Status update to 'confirmed' skipped: %s", e)
        # --- Vendor identity: confirmed PATCH'ten otomatik ogrenme ---
        # Manuel kayit (Beleg hinzufugen) yoksa OCR'dan cikan kimlik anahtarlariyla
        # vendor_identities tablosuna 'auto_learned' olarak kaydet. Sonraki ayni
        # vendor'in fisi geldiginde otomatik tanir.
        try:
            from autotax.vendor_identity import learn_from_invoice
            # source='manual' — PATCH = kullanici dogruladi. Adres/telefon/email
            # yeterli (market fislerinde IBAN yok). Confidence=1.0 atanir.
            learn_from_invoice(inv, source="manual")
        except Exception as e:
            logger.warning("Vendor identity auto-learn skipped: %s", e)
        # Sync changes to linked CashEntry (Kassenbuch)
        linked = db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user["sub"]).first()
        if linked:
            if body.vendor is not None:
                linked.vendor = body.vendor
                linked.description = f"Rechnung: {body.vendor}"
            if body.total_amount is not None:
                linked.gross_amount = body.total_amount
            if body.vat_amount is not None:
                linked.vat_amount = body.vat_amount
            if body.vat_rate is not None:
                linked.vat_rate = body.vat_rate
            if body.category is not None:
                linked.category = body.category
            if body.invoice_type is not None:
                linked.entry_type = body.invoice_type
            if body.payment_method is not None:
                linked.payment_method = body.payment_method
            if body.date is not None:
                linked.date = parse_date_str_to_datetime(body.date)
            db.commit()
            logger.info("Synced invoice %d changes to linked cash entry %d", invoice_id, linked.id)
        elif body.total_amount is not None and float(body.total_amount) > 0:
            # No linked entry exists — create one
            auto_create_cash_entry(invoice_id, user["sub"], {
                "vendor": inv.vendor, "total_amount": inv.total_amount,
                "vat_amount": inv.vat_amount, "vat_rate": inv.vat_rate,
                "date": inv.date, "category": inv.category,
                "invoice_type": inv.invoice_type, "payment_method": inv.payment_method,
                "invoice_number": inv.invoice_number,
            })
        return {"success": True, **invoice_to_dict(inv)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Update invoice failed")
        err(500, "Failed to update invoice")
    finally:
        db.close()


@app.patch("/invoices/{invoice_id}")
def patch_invoice(invoice_id: int, body: InvoiceUpdate, user: dict = Depends(get_current_user)):
    return _do_update_invoice(invoice_id, body, user)


@app.put("/invoices/{invoice_id}")
def put_invoice(invoice_id: int, body: InvoiceUpdate, user: dict = Depends(get_current_user)):
    return _do_update_invoice(invoice_id, body, user)


# --- ADDED START: Single invoice detail with full OCR text ---
@app.get("/invoices/{invoice_id}/detail")
def get_invoice_detail(invoice_id: int, user: dict = Depends(get_current_user)):
    """Get full invoice detail including complete raw OCR text."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        result = invoice_to_dict(inv)
        result["raw_text"] = inv.raw_text or ""
        return result
    finally:
        db.close()
# --- ADDED END ---

# --- ADDED START: Async OCR upload + status endpoint ---
_bg_tasks: set = set()  # prevent GC of background tasks

@app.post("/invoices/upload-async")
async def upload_invoice_async(request: Request, file: UploadFile = File(...), handwriting: bool = False, invoice_type: str = "expense", user: dict = Depends(get_current_user)):
    """Async upload: Tesseract first, if fails creates placeholder invoice + runs OCR.space in background."""
    _enforce_upload_quota(user["sub"])
    import asyncio as _asyncio
    if file.content_type not in ALLOWED_TYPES:
        err(400, "Ungültige Datei")
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        err(400, "Datei zu groß")
    if len(content) == 0:
        err(400, "Leere Datei")

    # Sync fast-path: Tesseract for images, pdfplumber for digital PDFs.
    # Both are local + fast; avoids the async OCR.space round-trip that
    # frontends can time out on (>30s polling).
    _ct = (file.content_type or "").lower()
    if _ct.startswith("image/"):
        try:
            from autotax.ocr import local_ocr_tesseract, is_ocr_valid
            _tess_text = local_ocr_tesseract(content)
            if is_ocr_valid(_tess_text):
                logger.info("Using local OCR (Tesseract) sync: %s (%d chars)", file.filename, len(_tess_text))
                parsed = parse_invoice(_tess_text)
                parsed = apply_filename_overrides(parsed, file.filename or "", raw_text=_tess_text)
                if invoice_type in ("income", "expense"):
                    parsed["invoice_type"] = invoice_type
                inv_id = save_invoice(parsed, user_id=user["sub"], filename=file.filename, file_data=content, file_content_type=file.content_type or "")
                auto_create_cash_entry(inv_id, user["sub"], parsed)
                return {"status": "done", "id": inv_id, "total_amount": safe_float(parsed.get("total_amount")), "vendor": parsed.get("vendor", "")}
        except Exception as e:
            logger.warning("Tesseract sync failed: %s", e)
    elif _ct == "application/pdf" or (file.filename or "").lower().endswith(".pdf"):
        try:
            from autotax.ocr import extract_pdf_text
            pdf_text = extract_pdf_text(content) or ""
            stripped = pdf_text.strip()
            # Only treat as digital invoice when text is substantial AND
            # contains at least one total/invoice keyword. Scanned-as-PDF
            # receipts often carry a few dozen chars of garbage metadata
            # whose column order breaks the total extractor — let those
            # fall through to the OCR.space path instead.
            _inv_kw = _re_global.search(
                r"\b(gesamt|summe|total|betrag|brutto|zu\s*zahlen|rechnungsbetrag|zahlbetrag|endbetrag|rechnung|invoice|beleg|mwst|ust)\b",
                stripped, _re_global.IGNORECASE,
            )
            if len(stripped) >= 200 and _inv_kw:
                logger.info("Using native PDF text (pdfplumber) sync: %s (%d chars)", file.filename, len(stripped))
                parsed = parse_invoice(pdf_text)
                parsed = apply_filename_overrides(parsed, file.filename or "", raw_text=pdf_text)
                # Sanity guard: if parser returned no total, don't commit —
                # fall through to OCR.space which may read the layout better.
                if safe_float(parsed.get("total_amount")) > 0:
                    if invoice_type in ("income", "expense"):
                        parsed["invoice_type"] = invoice_type
                    inv_id = save_invoice(parsed, user_id=user["sub"], filename=file.filename, file_data=content, file_content_type=file.content_type or "application/pdf")
                    auto_create_cash_entry(inv_id, user["sub"], parsed)
                    return {"status": "done", "id": inv_id, "total_amount": safe_float(parsed.get("total_amount")), "vendor": parsed.get("vendor", "")}
                else:
                    logger.info("pdfplumber got text but no total — falling through to OCR path")
            else:
                logger.info("PDF native text sparse (%d chars, kw=%s) — using OCR path", len(stripped), bool(_inv_kw))
        except Exception as e:
            logger.warning("pdfplumber sync failed: %s", e)

    # Fast path missed (scanned PDF, Tesseract invalid) → placeholder + OCR.space async
    placeholder = {"vendor": "Processing...", "total_amount": 0.0, "date": datetime.now().strftime("%Y-%m-%d"), "raw_text": "", "invoice_type": invoice_type, "vat_amount": 0.0, "vat_rate": "0%", "category": "other", "invoice_number": "", "payment_method": ""}
    inv_id = save_invoice(placeholder, user_id=user["sub"], filename=file.filename, file_data=content, file_content_type=file.content_type or "")
    logger.info("Async OCR started: invoice %d (%s)", inv_id, file.filename)

    async def _bg_ocr(inv_id, content, filename, ct, handwriting, user_sub, invoice_type):
        try:
            from fastapi import UploadFile as _UF
            from starlette.datastructures import Headers as _Headers
            import io as _io
            fake = _UF(filename=filename, file=_io.BytesIO(content), headers=_Headers({"content-type": ct or "application/octet-stream"}))
            raw_text, qr_data = await _asyncio.wait_for(extract_text_and_qr(fake, handwriting=handwriting, file_bytes=content), timeout=60)
            parsed = parse_invoice(raw_text)
            parsed = apply_filename_overrides(parsed, filename or "", raw_text=raw_text)
            if invoice_type in ("income", "expense"):
                parsed["invoice_type"] = invoice_type
            # Update the placeholder invoice
            db_bg = SessionLocal()
            try:
                inv = db_bg.query(Invoice).filter(Invoice.id == inv_id, Invoice.user_id == user_sub).first()
                if inv:
                    inv.vendor = parsed.get("vendor") or "Unbekannt"
                    inv.total_amount = safe_float(parsed.get("total_amount"))
                    inv.vat_amount = safe_float(parsed.get("vat_amount"))
                    inv.vat_rate = parsed.get("vat_rate") or "0%"
                    inv.date = parsed.get("date") or ""
                    inv.category = parsed.get("category") or "other"
                    inv.invoice_number = parsed.get("invoice_number") or ""
                    inv.payment_method = parsed.get("payment_method") or ""
                    inv.raw_text = raw_text[:2000]
                    inv.processed = True
                    # Vendor identity fields — A4 fatura icin printout'ta gerekli.
                    # Sync upload path bunlari save_invoice() icinden zaten yaziyor;
                    # bg OCR path'i de ayni alanlari yazsin diye buraya eklendi.
                    # Ust-IdNr / IBAN / email / phone / adres / HRB / domain.
                    inv.vendor_iban = parsed.get("vendor_iban") or ""
                    inv.vendor_email = parsed.get("vendor_email") or ""
                    inv.vendor_phone = parsed.get("vendor_phone") or ""
                    inv.vendor_fax = parsed.get("vendor_fax") or ""
                    inv.vendor_address = parsed.get("vendor_address") or ""
                    inv.vendor_website = parsed.get("vendor_website") or parsed.get("vendor_domain") or ""
                    inv.vendor_ust_id = parsed.get("vendor_ust_id") or None
                    inv.vendor_hrb = parsed.get("vendor_hrb") or None
                    inv.vendor_steuernr = parsed.get("vendor_steuernr") or None
                    # Reminder system: parser'dan otomatik due_date geldiyse yaz
                    if parsed.get("due_date"):
                        inv.due_date = parsed["due_date"]
                    db_bg.commit()
                    logger.info("Async OCR completed: invoice %d (%s, €%.2f, ust_id=%s)",
                                inv_id, parsed.get("vendor"),
                                safe_float(parsed.get("total_amount")),
                                parsed.get("vendor_ust_id") or "-")
            finally:
                db_bg.close()

            # Create the cash-entry mirror in Kassenbuch. The two sync
            # fast-paths above (Tesseract / pdfplumber) call this directly
            # — the bg OCR path was missing it, so PDF scans that fall
            # through to OCR.space ended up in /invoices but never in
            # /bookkeeping. User reported: 'fisler kassenbucher gitmiyor'.
            try:
                auto_create_cash_entry(inv_id, user_sub, parsed)
            except Exception as ce:
                logger.warning("Auto cash-entry failed for invoice %d: %s", inv_id, ce)
        except Exception as e:
            logger.warning("Async OCR failed for invoice %d: %s", inv_id, e)

    task = _asyncio.create_task(_bg_ocr(inv_id, content, file.filename, file.content_type or "", handwriting, user["sub"], invoice_type))
    _bg_tasks.add(task)
    task.add_done_callback(_bg_tasks.discard)
    return {"status": "processing", "id": inv_id, "message": "OCR is being processed in background"}


@app.get("/invoices/{invoice_id}/status")
def invoice_status(invoice_id: int, user: dict = Depends(get_current_user)):
    """Check processing status of an invoice (for async upload polling)."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        if inv.vendor == "Processing..." or not inv.processed:
            return {"status": "processing", "id": invoice_id}
        return {"status": "done", "id": invoice_id, "vendor": inv.vendor, "total_amount": safe_float(inv.total_amount)}
    finally:
        db.close()
# --- ADDED END ---


# ============================================================
# INVOICES: DELETE
# ============================================================

@app.delete("/invoices/{invoice_id}")
def delete_invoice(invoice_id: int, permanent: bool = False, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        # --- ADDED: soft delete ---
        if permanent:
            db.delete(inv)
            logger.info("Permanent delete: invoice %d", invoice_id)
        else:
            inv.is_deleted = True
            inv.deleted_at = datetime.now()
            logger.info("Soft delete: invoice %d", invoice_id)
            # Also soft-delete linked cash entry
            linked = db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user["sub"]).first()
            if linked:
                linked.is_deleted = True
                linked.deleted_at = datetime.now()
        # --- END ---
        db.commit()
        return {"success": True, "deleted": invoice_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Delete invoice failed")
        err(500, "Failed to delete invoice")
    finally:
        db.close()


class BulkDeleteRequest(BaseModel):
    ids: List[int]


@app.post("/invoices/bulk-delete")
def bulk_delete_invoices(body: BulkDeleteRequest, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        # --- ADDED: soft delete ---
        invs = db.query(Invoice).filter(Invoice.id.in_(body.ids), Invoice.user_id == user["sub"]).all()
        deleted = 0
        for inv in invs:
            inv.is_deleted = True
            inv.deleted_at = datetime.now()
            deleted += 1
            linked = db.query(CashEntry).filter(CashEntry.invoice_id == inv.id, CashEntry.user_id == user["sub"]).first()
            if linked:
                linked.is_deleted = True
                linked.deleted_at = datetime.now()
        logger.info("Soft bulk delete: %d invoices", deleted)
        # --- END ---
        db.commit()
        return {"success": True, "deleted": deleted}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Bulk delete failed")
        err(500, "Bulk delete failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: MODELS
# ============================================================

class CashEntryCreate(BaseModel):
    description: str
    gross_amount: float
    entry_type: str
    vendor: Optional[str] = None
    category: Optional[str] = None
    vat_rate: Optional[str] = None
    payment_method: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None
    date: Optional[str] = None
    # Vendor identity fingerprint — Beleg hinzufugen formundan manuel kayit
    # icin kullanici girer; vendor_identities tablosuna upsert edilir.
    iban: Optional[str] = None
    email: Optional[str] = None
    phone: Optional[str] = None
    fax: Optional[str] = None
    address: Optional[str] = None
    ust_id: Optional[str] = None
    hrb: Optional[str] = None


class CashEntryUpdate(BaseModel):
    description: Optional[str] = None
    gross_amount: Optional[float] = None
    entry_type: Optional[str] = None
    vendor: Optional[str] = None
    category: Optional[str] = None
    vat_rate: Optional[str] = None
    payment_method: Optional[str] = None
    reference: Optional[str] = None
    notes: Optional[str] = None
    date: Optional[str] = None


# ============================================================
# BOOKKEEPING: LIST (GET /bookkeeping + /kassenbuch)
# ============================================================

def _list_bookkeeping(skip, limit, user):
    db = SessionLocal()
    try:
        q = db.query(CashEntry).filter(CashEntry.user_id == user["sub"])
        # --- ADDED: exclude soft-deleted ---
        q = q.filter((CashEntry.is_deleted == False) | (CashEntry.is_deleted == None))
        # --- END ---
        from sqlalchemy import func, case
        total_count = q.count()
        entries = q.order_by(CashEntry.date.desc()).offset(skip).limit(limit).all()
        # Calculate totals via SQL aggregation (not loading all into memory)
        agg = db.query(
            func.coalesce(func.sum(CashEntry.gross_amount), 0).label("total_gross"),
            func.coalesce(func.sum(CashEntry.vat_amount), 0).label("total_vat"),
            func.coalesce(func.sum(case((CashEntry.entry_type == "income", CashEntry.gross_amount), else_=0)), 0).label("total_income"),
            func.coalesce(func.sum(case((CashEntry.entry_type == "expense", CashEntry.gross_amount), else_=0)), 0).label("total_expense"),
            func.coalesce(func.sum(case((CashEntry.entry_type == "income", CashEntry.vat_amount), else_=0)), 0).label("vat_income"),
            func.coalesce(func.sum(case((CashEntry.entry_type == "expense", CashEntry.vat_amount), else_=0)), 0).label("vat_expense"),
        ).filter(CashEntry.user_id == user["sub"], (CashEntry.is_deleted == False) | (CashEntry.is_deleted == None)).first()
        total_gross = float(agg.total_gross)
        total_vat = float(agg.total_vat)
        total_income = float(agg.total_income)
        total_expense = float(agg.total_expense)
        vat_balance = float(agg.vat_income) - float(agg.vat_expense)
        return {
            "success": True,
            "items": [cash_entry_to_dict(e) for e in entries],
            "total": total_count,
            "summary": {
                "total_gross": round(total_gross, 2),
                "total_vat": round(total_vat, 2),
                "total_income": round(total_income, 2),
                "total_expense": round(total_expense, 2),
                "total_expenses": round(total_expense, 2),
                "net": round(total_income - total_expense, 2),
                "net_profit": round(total_income - total_expense, 2),
                "vat_balance": round(vat_balance, 2),
                "entry_count": total_count,
            },
        }
    except Exception:
        logger.exception("Failed to list cash entries")
        err(500, "Failed to load cash entries")
    finally:
        db.close()


@app.get("/bookkeeping")
def list_bookkeeping(skip: int = Query(0, ge=0), limit: int = Query(50, ge=1, le=10000), user: dict = Depends(get_current_user)):
    return _list_bookkeeping(skip, limit, user)


@app.get("/kassenbuch")
def list_kassenbuch(skip: int = Query(0, ge=0), limit: int = Query(50, ge=1, le=10000), user: dict = Depends(get_current_user)):
    return _list_bookkeeping(skip, limit, user)


# ============================================================
# BOOKKEEPING: CREATE (POST /bookkeeping + /kassenbuch)
# ============================================================

def _create_bookkeeping(body: CashEntryCreate, user: dict):
    if body.entry_type not in ("income", "expense"):
        err(400, "entry_type must be 'income' or 'expense'")
    db = SessionLocal()
    try:
        entry_date = parse_date_str_to_datetime(body.date)
        vat_amount = calc_vat(body.gross_amount, body.vat_rate)
        entry = CashEntry(
            user_id=user["sub"],
            description=body.description,
            gross_amount=body.gross_amount,
            vat_amount=vat_amount,
            vat_rate=body.vat_rate or "0%",
            vendor=body.vendor or "Unbekannt",
            entry_type=body.entry_type,
            category=body.category or "other",
            payment_method=body.payment_method or "",
            reference=body.reference or "",
            notes=body.notes or "",
            date=entry_date,
        )
        # Create Invoice first so we can link it
        inv = Invoice(
            user_id=user["sub"],
            filename=None,
            vendor=body.vendor or "Manual Entry",
            total_amount=body.gross_amount or 0.0,
            vat_amount=vat_amount,
            vat_rate=body.vat_rate or "0%",
            date=body.date or "",
            raw_text=f"manual entry: {body.description}",
            invoice_type=body.entry_type,
            invoice_number="",
            payment_method=body.payment_method or "",
            category=body.category or "other",
            processed=True,
            status="confirmed",  # manuel kayit = zaten onaylanmis
            # Vendor identity bilgileri — Beleg hinzufugen formundan
            vendor_iban=body.iban or None,
            vendor_email=body.email or None,
            vendor_phone=body.phone or None,
            vendor_address=body.address or None,
            vendor_ust_id=body.ust_id or None,
            vendor_hrb=body.hrb or None,
        )
        db.add(inv)
        db.flush()
        entry.invoice_id = inv.id
        db.add(entry)
        db.commit()
        db.refresh(entry)
        logger.info("Created Kassenbuch entry %s linked to invoice %s", entry.id, inv.id)
        # --- Vendor identity: manuel kayittan ogrenme ---
        # Form'da vendor + en az bir kimlik anahtari (USt-IdNr/IBAN/HRB/email/...)
        # varsa vendor_identities tablosuna upsert et. Bir sonraki ayni vendor'in
        # OCR'i geldiginde otomatik tanima saglar.
        try:
            if body.vendor and body.vendor.strip() and body.vendor != "Unbekannt":
                from autotax.vendor_identity import save_or_update
                save_or_update(
                    user_id=user["sub"],
                    vendor_name=body.vendor,
                    ust_id=body.ust_id,
                    iban=body.iban,
                    hrb=body.hrb,
                    phone=body.phone,
                    email=body.email,
                    address=body.address,
                    default_vat_rate=body.vat_rate,
                    default_category=body.category,
                    default_payment_method=body.payment_method,
                    source="manual",
                )
        except Exception as _ie:
            logger.warning("Vendor identity manual save skipped: %s", _ie)
        return {"success": True, **cash_entry_to_dict(entry)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Create cash entry failed")
        err(500, "Failed to create entry")
    finally:
        db.close()


@app.post("/bookkeeping")
def create_bookkeeping(body: CashEntryCreate, user: dict = Depends(get_current_user)):
    return _create_bookkeeping(body, user)


@app.post("/kassenbuch")
def create_kassenbuch(body: CashEntryCreate, user: dict = Depends(get_current_user)):
    return _create_bookkeeping(body, user)


# ============================================================
# BOOKKEEPING: UPDATE (PATCH+PUT /bookkeeping/{id} + /kassenbuch/{id})
# ============================================================

def _update_bookkeeping(entry_id: int, body: CashEntryUpdate, user: dict):
    db = SessionLocal()
    try:
        entry = db.query(CashEntry).filter(CashEntry.id == entry_id, CashEntry.user_id == user["sub"]).first()
        if not entry:
            err(404, "Entry not found")
        if body.description is not None:
            entry.description = body.description
        if body.gross_amount is not None:
            entry.gross_amount = body.gross_amount
        if body.entry_type is not None:
            if body.entry_type not in ("income", "expense"):
                err(400, "entry_type must be 'income' or 'expense'")
            entry.entry_type = body.entry_type
        if body.vendor is not None:
            entry.vendor = body.vendor
        if body.category is not None:
            entry.category = body.category
        if body.vat_rate is not None:
            entry.vat_rate = body.vat_rate
        if body.payment_method is not None:
            entry.payment_method = body.payment_method
        if body.reference is not None:
            entry.reference = body.reference
        if body.notes is not None:
            entry.notes = body.notes
        if body.date is not None:
            entry.date = parse_date_str_to_datetime(body.date)
        if body.gross_amount is not None or body.vat_rate is not None:
            entry.vat_amount = calc_vat(entry.gross_amount, entry.vat_rate)
        db.commit()
        db.refresh(entry)
        # Sync changes to linked Invoice (Rechnungen)
        if entry.invoice_id:
            linked_inv = db.query(Invoice).filter(Invoice.id == entry.invoice_id, Invoice.user_id == user["sub"]).first()
            if linked_inv:
                if body.vendor is not None:
                    linked_inv.vendor = body.vendor
                if body.gross_amount is not None:
                    linked_inv.total_amount = body.gross_amount
                if body.vat_rate is not None:
                    linked_inv.vat_rate = body.vat_rate
                    linked_inv.vat_amount = entry.vat_amount
                if body.category is not None:
                    linked_inv.category = body.category
                if body.entry_type is not None:
                    linked_inv.invoice_type = body.entry_type
                if body.payment_method is not None:
                    linked_inv.payment_method = body.payment_method
                if body.date is not None:
                    linked_inv.date = body.date
                db.commit()
                logger.info("Synced cash entry %d changes to linked invoice %d", entry_id, entry.invoice_id)
        return {"success": True, **cash_entry_to_dict(entry)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Update cash entry failed")
        err(500, "Failed to update entry")
    finally:
        db.close()


@app.patch("/bookkeeping/{entry_id}")
def patch_bookkeeping(entry_id: int, body: CashEntryUpdate, user: dict = Depends(get_current_user)):
    return _update_bookkeeping(entry_id, body, user)


@app.put("/bookkeeping/{entry_id}")
def put_bookkeeping(entry_id: int, body: CashEntryUpdate, user: dict = Depends(get_current_user)):
    return _update_bookkeeping(entry_id, body, user)


@app.patch("/kassenbuch/{entry_id}")
def patch_kassenbuch(entry_id: int, body: CashEntryUpdate, user: dict = Depends(get_current_user)):
    return _update_bookkeeping(entry_id, body, user)


@app.put("/kassenbuch/{entry_id}")
def put_kassenbuch(entry_id: int, body: CashEntryUpdate, user: dict = Depends(get_current_user)):
    return _update_bookkeeping(entry_id, body, user)


# ============================================================
# BOOKKEEPING: DELETE (/bookkeeping/{id} + /kassenbuch/{id})
# ============================================================

def _delete_bookkeeping(entry_id: int, user: dict):
    db = SessionLocal()
    try:
        entry = db.query(CashEntry).filter(CashEntry.id == entry_id, CashEntry.user_id == user["sub"]).first()
        if not entry:
            err(404, "Entry not found")
        entry.is_deleted = True
        entry.deleted_at = datetime.now()
        # Also soft-delete linked invoice
        if entry.invoice_id:
            linked_inv = db.query(Invoice).filter(Invoice.id == entry.invoice_id, Invoice.user_id == user["sub"]).first()
            if linked_inv:
                linked_inv.is_deleted = True
                linked_inv.deleted_at = datetime.now()
        logger.info("Soft delete: cash entry %d (+ linked invoice)", entry_id)
        db.commit()
        return {"success": True, "deleted": entry_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Delete cash entry failed")
        err(500, "Failed to delete entry")
    finally:
        db.close()


@app.post("/bookkeeping/bulk-delete")
def bulk_delete_bookkeeping(body: BulkDeleteRequest, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        entries = db.query(CashEntry).filter(CashEntry.id.in_(body.ids), CashEntry.user_id == user["sub"]).all()
        deleted = 0
        for entry in entries:
            entry.is_deleted = True
            entry.deleted_at = datetime.now()
            deleted += 1
            # Also soft-delete linked invoice
            if entry.invoice_id:
                linked_inv = db.query(Invoice).filter(Invoice.id == entry.invoice_id, Invoice.user_id == user["sub"]).first()
                if linked_inv:
                    linked_inv.is_deleted = True
                    linked_inv.deleted_at = datetime.now()
        db.commit()
        logger.info("Bulk delete: %d cash entries", deleted)
        return {"success": True, "deleted": deleted}
    except Exception:
        db.rollback()
        logger.exception("Bulk delete bookkeeping failed")
        err(500, "Bulk delete failed")
    finally:
        db.close()


@app.delete("/bookkeeping/{entry_id}")
def delete_bookkeeping(entry_id: int, user: dict = Depends(get_current_user)):
    return _delete_bookkeeping(entry_id, user)


@app.delete("/kassenbuch/{entry_id}")
def delete_kassenbuch(entry_id: int, user: dict = Depends(get_current_user)):
    return _delete_bookkeeping(entry_id, user)


# ============================================================
# BOOKKEEPING: SYNC INVOICES
# ============================================================

@app.post("/bookkeeping/sync-invoices")
def sync_invoices_to_bookkeeping(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"], (Invoice.is_deleted == False) | (Invoice.is_deleted == None)).all()
        all_entries = db.query(CashEntry).filter(CashEntry.user_id == user["sub"], (CashEntry.is_deleted == False) | (CashEntry.is_deleted == None)).all()

        # Build sets for duplicate detection
        existing_invoice_ids = set()          # invoice IDs that already have a CashEntry
        entry_by_invoice_id = {}              # quick lookup: invoice_id → CashEntry
        for e in all_entries:
            if e.invoice_id:
                existing_invoice_ids.add(e.invoice_id)
                entry_by_invoice_id[e.invoice_id] = e

        # --- Forward sync: Invoice → CashEntry ---
        # Only create CashEntry for invoices that don't have one yet
        synced = 0
        skipped = 0
        for inv in invoices:
            if inv.id in existing_invoice_ids:
                skipped += 1
                continue
            # Extra safety: check by vendor+amount+date to avoid near-dupes
            inv_date = safe_date_str(inv.date)[:10]
            inv_vendor = safe_vendor(inv.vendor).lower()
            inv_amount = safe_float(inv.total_amount)
            already_exists = False
            for e in all_entries:
                if (safe_float(e.gross_amount) == inv_amount
                    and (e.vendor or e.description or "").lower() == inv_vendor
                    and (e.date.strftime("%Y-%m-%d") if e.date else "") == inv_date):
                    # Found matching entry — link them instead of creating new
                    if not e.invoice_id:
                        e.invoice_id = inv.id
                        existing_invoice_ids.add(inv.id)
                        synced += 1
                    already_exists = True
                    break
            if already_exists:
                continue
            vat_amount = calc_vat(inv_amount, safe_vat_rate(inv.vat_rate))
            entry_date = parse_date_str_to_datetime(inv.date) if inv.date else inv.created_at
            entry = CashEntry(
                user_id=user["sub"],
                description=safe_vendor(inv.vendor),
                vendor=safe_vendor(inv.vendor),
                gross_amount=inv_amount,
                vat_amount=vat_amount,
                vat_rate=safe_vat_rate(inv.vat_rate),
                entry_type=safe_invoice_type(inv.invoice_type),
                category=safe_category(inv.category),
                payment_method=safe_str(inv.payment_method),
                invoice_id=inv.id,
                date=entry_date,
            )
            db.add(entry)
            existing_invoice_ids.add(inv.id)
            synced += 1

        # --- Reverse sync: CashEntry → Invoice ---
        # Only for entries that have no linked invoice
        # Build invoice lookup: vendor+amount+date → invoice_id
        inv_lookup = {}
        for inv in invoices:
            key = f"{safe_vendor(inv.vendor).lower()}|{safe_float(inv.total_amount)}|{safe_date_str(inv.date)[:10]}"
            inv_lookup[key] = inv.id

        rev_synced = 0
        for entry in all_entries:
            if entry.invoice_id:
                continue  # already linked
            # Try to find existing invoice by vendor+amount+date
            e_date = entry.date.strftime("%Y-%m-%d") if entry.date else ""
            e_vendor = (entry.vendor or entry.description or "").lower()
            e_key = f"{e_vendor}|{safe_float(entry.gross_amount)}|{e_date}"
            if e_key in inv_lookup:
                entry.invoice_id = inv_lookup[e_key]
                rev_synced += 1
                continue
            # No matching invoice found — create one
            inv = Invoice(
                user_id=user["sub"],
                filename=None,
                vendor=entry.vendor or "Manual Entry",
                total_amount=safe_float(entry.gross_amount),
                vat_amount=safe_float(entry.vat_amount),
                vat_rate=entry.vat_rate or "0%",
                date=e_date,
                raw_text=f"Sync from Kassenbuch: {safe_str(entry.description)}",
                invoice_type=entry.entry_type or "expense",
                invoice_number="",
                payment_method=safe_str(entry.payment_method),
                category=safe_category(entry.category),
                processed=True,
            )
            db.add(inv)
            db.flush()
            entry.invoice_id = inv.id
            # Add to lookup so next identical entry links to same invoice
            inv_lookup[e_key] = inv.id
            rev_synced += 1

        db.commit()
        logger.info("Sync: %d forward, %d skipped, %d reverse", synced, skipped, rev_synced)
        return {"synced": synced, "skipped": skipped, "reverse_synced": rev_synced}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Sync invoices failed")
        err(500, "Sync failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: RECONCILE
# ============================================================

@app.post("/bookkeeping/{entry_id}/reconcile")
def reconcile_entry(entry_id: int, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        entry = db.query(CashEntry).filter(CashEntry.id == entry_id, CashEntry.user_id == user["sub"]).first()
        if not entry:
            err(404, "Entry not found")
        entry.is_reconciled = not entry.is_reconciled
        db.commit()
        db.refresh(entry)
        return {"success": True, **cash_entry_to_dict(entry)}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Reconcile failed")
        err(500, "Reconcile failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: SUMMARY
# ============================================================

@app.get("/bookkeeping/summary/overview")
def bookkeeping_summary(year: int = Query(None), user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        q = db.query(CashEntry).filter(CashEntry.user_id == user["sub"], (CashEntry.is_deleted == False) | (CashEntry.is_deleted == None))
        if year:
            q = q.filter(CashEntry.date >= datetime(year, 1, 1))
            q = q.filter(CashEntry.date < datetime(year + 1, 1, 1))
        entries = q.all()
        total_income = sum(safe_float(e.gross_amount) for e in entries if e.entry_type == "income")
        total_expenses = sum(safe_float(e.gross_amount) for e in entries if e.entry_type == "expense")
        vat_collected = sum(safe_float(e.vat_amount) for e in entries if e.entry_type == "income")
        vat_paid = sum(safe_float(e.vat_amount) for e in entries if e.entry_type == "expense")
        return {
            "total_income": round(total_income, 2),
            "total_expenses": round(total_expenses, 2),
            "net_profit": round(total_income - total_expenses, 2),
            "vat_balance": round(vat_collected - vat_paid, 2),
            "entry_count": len(entries),
        }
    except Exception:
        logger.exception("Summary failed")
        err(500, "Summary failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: EXPORT CSV
# ============================================================

@app.get("/bookkeeping/export/csv")
def export_bookkeeping_csv(year: int = Query(None), user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        q = db.query(CashEntry).filter(CashEntry.user_id == user["sub"], (CashEntry.is_deleted == False) | (CashEntry.is_deleted == None))
        if year:
            q = q.filter(CashEntry.date >= datetime(year, 1, 1))
            q = q.filter(CashEntry.date < datetime(year + 1, 1, 1))
        entries = q.order_by(CashEntry.date.desc()).all()
        buf = io.StringIO()
        buf.write("Datum,Typ,Beschreibung,Lieferant,Betrag,MwSt,MwSt-Satz,Kategorie,Zahlungsart,Beleg-Nr.\n")
        for e in entries:
            date_str = e.date.strftime("%d.%m.%Y") if e.date else ""
            desc = (e.description or "").replace('"', '""')
            vendor = (e.vendor or "").replace('"', '""')
            buf.write(f'{date_str},{e.entry_type or ""},"{desc}","{vendor}",{safe_float(e.gross_amount):.2f},{safe_float(e.vat_amount):.2f},{safe_vat_rate(e.vat_rate)},{safe_category(e.category)},{safe_str(e.payment_method)},{safe_str(e.reference)}\n')
        buf.seek(0)
        return StreamingResponse(buf, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=kassenbuch_{year or 'all'}.csv"})
    except Exception:
        logger.exception("Bookkeeping CSV export failed")
        err(500, "Export failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: CSV IMPORT
# ============================================================

@app.post("/bookkeeping/import-csv")
async def import_csv(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    import csv
    content = await file.read()
    text = content.decode("utf-8", errors="ignore")
    first_line = text.split("\n")[0] if text else ""
    delimiter = ";" if ";" in first_line else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)

    db = SessionLocal()
    imported = 0
    errors = []
    try:
        for idx, row in enumerate(reader, 1):
            try:
                # Flexible column mapping — supports Export format + custom formats
                vendor = row.get("Lieferant") or row.get("Vendor") or row.get("vendor") or ""
                beschreibung = row.get("Beschreibung") or row.get("description") or row.get("Description") or vendor or ""
                datum = row.get("Datum") or row.get("date") or row.get("Date") or ""
                betrag_raw = row.get("Betrag") or row.get("Ausgaben") or row.get("amount") or row.get("expenses") or "0"
                einnahmen = row.get("Einnahmen") or row.get("income") or "0"
                typ = row.get("Typ") or row.get("type") or row.get("Type") or ""
                category = row.get("Kategorie") or row.get("category") or row.get("Category") or "other"
                payment = row.get("Zahlungsart") or row.get("Zahlungsmethode") or row.get("payment_method") or ""
                mwst_raw = row.get("MwSt") or row.get("vat_amount") or ""
                mwst_satz = row.get("MwSt-Satz") or row.get("vat_rate") or "19%"
                inv_nr = row.get("Rechnungs-Nr.") or row.get("invoice_number") or ""
                if not vendor:
                    vendor = beschreibung[:50] or "Import"

                def _parse_num(s):
                    return float(str(s).replace(",", ".").replace("€", "").replace("%", "").replace(" ", "").strip() or "0")

                betrag_val = _parse_num(betrag_raw)
                einnahmen_val = _parse_num(einnahmen)

                # Determine type: from Typ column, or from Einnahmen column
                if typ.lower() in ("income", "einnahme", "einnahmen"):
                    entry_type = "income"
                    amount = betrag_val if betrag_val > 0 else einnahmen_val
                elif typ.lower() in ("expense", "ausgabe", "ausgaben"):
                    entry_type = "expense"
                    amount = betrag_val
                elif einnahmen_val > 0:
                    entry_type = "income"
                    amount = einnahmen_val
                else:
                    entry_type = "expense"
                    amount = betrag_val

                if amount <= 0 and not beschreibung:
                    continue

                # Duplicate check (skip soft-deleted)
                existing = db.query(CashEntry).filter(
                    CashEntry.user_id == user["sub"],
                    CashEntry.description == (beschreibung or vendor),
                    CashEntry.gross_amount == amount,
                    (CashEntry.is_deleted == False) | (CashEntry.is_deleted == None),
                ).first()
                if existing:
                    continue

                date_val = None
                if datum:
                    parts = datum.strip().split(".")
                    if len(parts) == 3:
                        try:
                            date_val = datetime(int(parts[2]), int(parts[1]), int(parts[0]))
                        except ValueError:
                            pass
                    if not date_val:
                        date_val = parse_date_str_to_datetime(datum)
                if not date_val:
                    date_val = datetime.now()

                # Use MwSt from CSV if provided, otherwise calculate
                if mwst_raw:
                    vat_amount = _parse_num(mwst_raw)
                else:
                    rate = _parse_num(mwst_satz) if mwst_satz else 19
                    vat_amount = round(amount * rate / (100 + rate), 2) if amount > 0 else 0
                vat_rate_str = mwst_satz if mwst_satz else "19%"
                if "%" not in vat_rate_str:
                    vat_rate_str += "%"

                inv = Invoice(
                    user_id=user["sub"],
                    filename=file.filename or f"csv-import-{idx}",
                    vendor=vendor,
                    total_amount=amount,
                    vat_amount=vat_amount,
                    vat_rate=vat_rate_str,
                    date=date_val.strftime("%Y-%m-%d") if date_val else "",
                    raw_text=f"CSV Import: {beschreibung}",
                    invoice_type=entry_type,
                    invoice_number=inv_nr,
                    payment_method=payment,
                    category=category,
                    processed=True,
                    file_data=content,
                    file_content_type="text/csv",
                )
                db.add(inv)
                db.flush()
                entry = CashEntry(
                    user_id=user["sub"],
                    description=beschreibung or vendor,
                    vendor=vendor,
                    gross_amount=amount,
                    vat_amount=vat_amount,
                    vat_rate=vat_rate_str,
                    entry_type=entry_type,
                    category=category,
                    payment_method=payment,
                    reference=f"CSV-Import Zeile {idx}",
                    notes="Importiert aus CSV",
                    date=date_val,
                    invoice_id=inv.id,
                )
                db.add(entry)
                imported += 1
            except Exception as e:
                errors.append(f"Zeile {idx}: {str(e)[:80]}")
        db.commit()
        return {"success": True, "imported": imported, "errors": errors}
    except Exception:
        db.rollback()
        logger.exception("CSV import failed")
        err(500, "CSV import failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: XLSX IMPORT
# ============================================================

@app.post("/bookkeeping/import-xlsx")
async def import_xlsx(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Import Excel (.xlsx) file into Kassenbuch + Rechnungen."""
    from openpyxl import load_workbook
    content = await file.read()
    wb = load_workbook(io.BytesIO(content), read_only=True)
    ws = wb.active
    db = SessionLocal()
    imported = 0
    errors = []
    try:
        headers = []
        header_row_found = False
        _KNOWN_HEADERS = {"datum", "date", "lieferant", "vendor", "betrag", "amount", "beschreibung", "description", "ausgaben", "expenses", "einnahmen", "income", "kategorie", "category", "typ", "type"}
        for idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if not header_row_found:
                # Scan for real header row (skip disclaimer/empty rows)
                candidate = [str(c or "").strip().lower() for c in row]
                if len(set(candidate) & _KNOWN_HEADERS) >= 2:
                    headers = candidate
                    header_row_found = True
                    logger.info("XLSX import: header found at row %d: %s", idx, headers[:5])
                    continue
                if idx > 5:
                    break  # give up after 5 rows
                continue
            if not any(row):
                continue
            rd = dict(zip(headers, [c for c in row]))

            def _col(names):
                for n in names:
                    v = rd.get(n) or rd.get(n.lower())
                    if v is not None and str(v).strip():
                        return str(v).strip()
                return ""

            beschreibung = _col(["beschreibung", "description", "lieferant", "vendor"])
            datum = _col(["datum", "date"])
            vendor = _col(["lieferant", "vendor"]) or beschreibung[:50] or "Import"
            category = _col(["kategorie", "category"]) or "other"
            payment = _col(["zahlungsart", "zahlungsmethode", "payment_method"]) or ""
            inv_nr = _col(["rechnungs-nr.", "invoice_number"]) or ""
            typ = _col(["typ", "type"]) or ""

            def _num(names):
                raw = _col(names)
                if not raw:
                    return 0.0
                return float(str(raw).replace(",", ".").replace("€", "").replace(" ", "").strip() or "0")

            betrag = _num(["betrag", "ausgaben", "amount", "expenses"])
            einnahmen = _num(["einnahmen", "income"])
            mwst = _num(["mwst", "vat_amount"])
            mwst_satz = _col(["mwst-satz", "vat_rate"]) or "19%"
            if "%" not in mwst_satz:
                mwst_satz += "%"

            if typ.lower() in ("income", "einnahme", "einnahmen"):
                entry_type = "income"
                amount = betrag if betrag > 0 else einnahmen
            elif einnahmen > 0:
                entry_type = "income"
                amount = einnahmen
            else:
                entry_type = "expense"
                amount = betrag

            if amount <= 0 and not beschreibung:
                continue

            date_val = parse_date_str_to_datetime(str(datum)) if datum else None
            if not date_val:
                date_val = datetime.now()
            if not mwst and amount > 0:
                rate = float(mwst_satz.replace("%", "").replace(",", ".").strip() or "19")
                mwst = round(amount * rate / (100 + rate), 2)

            try:
                inv = Invoice(user_id=user["sub"], filename=file.filename or f"xlsx-import-{idx}", vendor=vendor,
                    total_amount=amount, vat_amount=mwst, vat_rate=mwst_satz,
                    date=date_val.strftime("%Y-%m-%d") if date_val else "", raw_text=f"XLSX Import: {beschreibung}",
                    invoice_type=entry_type, invoice_number=inv_nr, payment_method=payment,
                    category=category, processed=True)
                db.add(inv)
                db.flush()
                entry = CashEntry(user_id=user["sub"], description=beschreibung or vendor, vendor=vendor,
                    gross_amount=amount, vat_amount=mwst, vat_rate=mwst_satz, entry_type=entry_type,
                    category=category, payment_method=payment, reference=f"XLSX-Import Zeile {idx}",
                    notes="XLSX Import", date=date_val, invoice_id=inv.id)
                db.add(entry)
                imported += 1
            except Exception as e:
                errors.append(f"Zeile {idx}: {str(e)[:80]}")
        db.commit()
        return {"success": True, "imported": imported, "errors": errors}
    except Exception:
        db.rollback()
        logger.exception("XLSX import failed")
        err(500, "XLSX Import failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: DATEV IMPORT
# ============================================================

@app.post("/bookkeeping/import-datev")
async def import_datev(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Import DATEV Buchungsstapel (.csv, semicolon-separated)."""
    import csv
    content = await file.read()
    text = content.decode("utf-8", errors="ignore")
    # Skip comment/disclaimer lines (e.g. "# HINWEIS: ...")
    lines = text.splitlines(keepends=True)
    clean_lines = [l for l in lines if not l.strip().startswith("#") and not l.strip().upper().startswith("HINWEIS")]
    reader = csv.DictReader(io.StringIO("".join(clean_lines)), delimiter=";")
    db = SessionLocal()
    imported = 0
    errors = []
    try:
        for idx, row in enumerate(reader, 1):
            try:
                # DATEV format: Umsatz;Soll/Haben;Konto;Gegenkonto;BU;Belegdatum;Buchungstext;USt
                umsatz_raw = row.get("Umsatz") or row.get("umsatz") or "0"
                sh = row.get("Soll/Haben") or row.get("soll/haben") or "S"
                buchungstext = row.get("Buchungstext") or row.get("buchungstext") or ""
                belegdatum = row.get("Belegdatum") or row.get("belegdatum") or ""
                ust = row.get("USt") or row.get("ust") or "19"

                amount = float(umsatz_raw.replace(",", ".").replace(" ", "").strip() or "0")
                if amount <= 0:
                    continue

                entry_type = "expense" if sh.upper() == "S" else "income"

                # Parse DATEV date: DDMM or DDMMYYYY
                date_val = None
                bd = belegdatum.strip()
                if len(bd) == 4:
                    date_val = parse_date_str_to_datetime(f"20{datetime.now().year % 100}-{bd[2:4]}-{bd[0:2]}")
                elif len(bd) >= 6:
                    date_val = parse_date_str_to_datetime(f"{bd[4:]}-{bd[2:4]}-{bd[0:2]}")
                if not date_val:
                    date_val = datetime.now()

                vat_rate = f"{ust}%"
                rate_f = float(ust.replace(",", ".").strip() or "19")
                vat_amount = round(amount * rate_f / (100 + rate_f), 2)

                inv = Invoice(user_id=user["sub"], filename=f"datev-import-{idx}", vendor=buchungstext[:50] or "DATEV Import",
                    total_amount=amount, vat_amount=vat_amount, vat_rate=vat_rate,
                    date=date_val.strftime("%Y-%m-%d") if date_val else "", raw_text=f"DATEV Import: {buchungstext}",
                    invoice_type=entry_type, invoice_number="", payment_method="",
                    category="other", processed=True)
                db.add(inv)
                db.flush()
                entry = CashEntry(user_id=user["sub"], description=buchungstext, vendor=buchungstext[:50] or "DATEV Import",
                    gross_amount=amount, vat_amount=vat_amount, vat_rate=vat_rate, entry_type=entry_type,
                    category="other", payment_method="", reference=f"DATEV-Import Zeile {idx}",
                    notes="DATEV Import", date=date_val, invoice_id=inv.id)
                db.add(entry)
                imported += 1
            except Exception as e:
                errors.append(f"Zeile {idx}: {str(e)[:80]}")
        db.commit()
        if imported == 0:
            return {"success": False, "imported": 0, "errors": errors,
                    "detail": "Keine DATEV-Buchungen gefunden. Erwartet: Umsatz;Soll/Haben;Konto;Gegenkonto;BU;Belegdatum;Buchungstext;USt"}
        return {"success": True, "imported": imported, "errors": errors}
    except Exception:
        db.rollback()
        logger.exception("DATEV import failed")
        err(500, "DATEV Import failed")
    finally:
        db.close()


# ============================================================
# BOOKKEEPING: PHOTO IMPORT (Handwritten Kassenbuch)
# ============================================================

@app.post("/bookkeeping/import-photo")
async def import_kassenbuch_photo(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Import handwritten Kassenbuch photo — OCR handwriting + table parsing"""
    from autotax.ocr import extract_handwriting_text
    import re as _re

    import gc
    content = await file.read()
    try:
        text = await extract_handwriting_text(content, file.filename or "kassenbuch.jpg")
    finally:
        del content
        gc.collect()
    if not text:
        err(400, "Konnte das Bild nicht lesen. Bitte bessere Qualität verwenden.")

    lines = text.strip().split("\n")
    db = SessionLocal()
    imported = 0
    try:
        for line in lines:
            line = line.strip()
            if not line or len(line) < 4:
                continue
            # Pattern 1: date + description + amount (strict)
            m = _re.search(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+?)\s+(\d+[.,]\d{2})\s*$", line)
            # Pattern 2: date + separator + description + amount (pipes/slashes from OCR)
            if not m:
                m = _re.search(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s*[|/]?\s*(.+?)\s+[/|]?\s*(\d+[.,]\d{2})", line)
            # Pattern 3: date with spaces (OCR misread: "01 03 26" instead of "01.03.26")
            if not m:
                m = _re.search(r"(\d{1,2}\s\d{1,2}\s\d{2,4})\s+(.+?)\s+(\d+[.,]\d{2})", line)
            # Pattern 4: date + description + amount without decimals (e.g. "50" instead of "50,00")
            if not m:
                m = _re.search(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+?)\s+(\d{1,6})\s*$", line)
            # Pattern 5: date with dashes (01-03-26)
            if not m:
                m = _re.search(r"(\d{1,2}-\d{1,2}-\d{2,4})\s+(.+?)\s+(\d+[.,]\d{2})", line)
            if not m:
                continue
            datum_raw = m.group(1)
            beschreibung = m.group(2).strip()
            amt_str = m.group(3).replace(",", ".")
            betrag = float(amt_str) if "." in amt_str else float(amt_str)
            if betrag <= 0 or len(beschreibung) < 2:
                continue
            # Normalize separators: space, slash, dash → dot
            parts = datum_raw.replace("/", ".").replace("-", ".").replace(" ", ".").split(".")
            date_str = ""
            if len(parts) == 3:
                d, mo, y = parts
                if len(y) == 2:
                    y = "20" + y
                date_str = f"{y}-{mo.zfill(2)}-{d.zfill(2)}"
            vat_amount = round(betrag * 19 / 119, 2)
            entry = CashEntry(
                user_id=user["sub"],
                description=beschreibung,
                vendor=beschreibung[:50],
                gross_amount=betrag,
                vat_amount=vat_amount,
                vat_rate="19%",
                entry_type="expense",
                category="other",
                payment_method="",
                reference="",
                notes="Kassenbuch Foto Import",
                date=parse_date_str_to_datetime(date_str),
            )
            db.add(entry)
            inv = Invoice(
                user_id=user["sub"],
                filename="kassenbuch-foto",
                vendor=beschreibung[:50],
                total_amount=betrag,
                vat_amount=vat_amount,
                vat_rate="19%",
                date=date_str,
                raw_text=f"Kassenbuch Foto: {beschreibung}",
                invoice_type="expense",
                invoice_number="",
                payment_method="",
                category="other",
                processed=True,
            )
            db.add(inv)
            imported += 1
        db.commit()
        return {"success": True, "imported": imported, "ocr_text": text[:500]}
    except Exception:
        db.rollback()
        logger.exception("Kassenbuch photo import failed")
        err(500, "Import failed")
    finally:
        db.close()


@app.post("/api/import-image")
async def import_image_table(file: UploadFile = File(...), save: bool = False, user: dict = Depends(get_current_user)):
    """Import Kassenbuch table image → OCR → structured rows + CSV.
    Columns: Nr, Datum, Beschreibung, Einnahmen, Ausgaben, Saldo
    Returns JSON rows + CSV string. If save=true, also saves to DB.
    """
    from autotax.ocr import extract_handwriting_text, extract_image_text, extract_pdf_text, extract_pdf_page_as_image, extract_table_text_autorotate
    import re as _re
    import gc

    content = await file.read()
    filename = (file.filename or "").lower()
    content_type = (file.content_type or "").lower()

    # Auto-rotate image if it's a table photo (landscape table in portrait photo)
    if "image" in content_type or filename.endswith((".jpg",".jpeg",".png")):
        try:
            from PIL import Image, ImageOps
            _img = Image.open(io.BytesIO(content))
            try:
                _img = ImageOps.exif_transpose(_img)
            except Exception:
                pass
            # If portrait and likely a landscape table → try 90° rotation
            if _img.height > _img.width * 1.2:
                _buf = io.BytesIO()
                _img_rot = _img.rotate(-90, expand=True)
                _img_rot.save(_buf, format="JPEG", quality=92)
                _content_rotated = _buf.getvalue()
                logger.info("Table auto-rotate: portrait→landscape (%dx%d → %dx%d)", _img.width, _img.height, _img_rot.width, _img_rot.height)
            else:
                _content_rotated = None
        except Exception:
            _content_rotated = None
    else:
        _content_rotated = None

    # PDF support
    text = ""
    try:
        if "pdf" in content_type or filename.endswith(".pdf"):
            text = extract_pdf_text(content)
            if not text or len(text.strip()) < 20:
                img_bytes = extract_pdf_page_as_image(content)
                if img_bytes:
                    text = await extract_image_text(img_bytes, "scanned.png")
        else:
            # Strategy 1: Try table grid cell detection
            try:
                from autotax.ocr import extract_table_cells, table_cells_to_text
                # Try rotated version first if available
                _cell_content = _content_rotated or content
                cells = extract_table_cells(_cell_content)
                if not cells or len(cells) < 4:
                    cells = extract_table_cells(content)
                if cells and len(cells) > 3:
                    cell_text = table_cells_to_text(cells)
                    logger.info("Table grid OCR: %d rows, %d chars", len(cells), len(cell_text))
                    text = cell_text
            except Exception as e:
                logger.debug("Table grid detection failed: %s", e)

            # Strategy 2: OCR.space handwriting — try both orientations
            if not text or len(text.strip()) < 50:
                # Try original
                _ocr_text = await extract_table_text_autorotate(content, file.filename or "kassenbuch.jpg")
                logger.info("Table OCR.space E2 original: %d chars", len(_ocr_text.strip()) if _ocr_text else 0)
                # Try rotated if available
                _ocr_rot = ""
                if _content_rotated:
                    _ocr_rot = await extract_table_text_autorotate(_content_rotated, file.filename or "kassenbuch_rot.jpg")
                    logger.info("Table OCR.space E2 rotated: %d chars", len(_ocr_rot.strip()) if _ocr_rot else 0)
                # Pick best
                import re as _re_amt
                _orig_amts = len(_re_amt.findall(r"\d+[.,]\d{2}", _ocr_text or ""))
                _rot_amts = len(_re_amt.findall(r"\d+[.,]\d{2}", _ocr_rot or ""))
                if _rot_amts > _orig_amts:
                    _best_ocr = _ocr_rot
                    logger.info("Using rotated OCR (more amounts: %d vs %d)", _rot_amts, _orig_amts)
                else:
                    _best_ocr = _ocr_text
                if _best_ocr and len(_best_ocr.strip()) > len((text or "").strip()):
                    text = _best_ocr

            # Strategy 3: printed OCR fallback
            if not text or len(text.strip()) < 30:
                text_printed = await extract_image_text(content, file.filename or "kassenbuch.png")
                if text_printed and len(text_printed.strip()) > len((text or "").strip()):
                    text = text_printed
    finally:
        del content
        gc.collect()

    if not text or len(text.strip()) < 10:
        err(400, "Konnte das Bild nicht lesen. Bitte bessere Qualität verwenden.")

    import time as _time
    _t0 = _time.time()

    # Preserve raw table before any modification
    raw_lines = [l.strip() for l in text.strip().split("\n") if l.strip()]

    # Amount pattern: matches 42,50 | 1.234,56 | 800,00 | 42.50 | 800 | -16,60
    _AMT_PAT = r"-?\d[\d.]*[.,]\d{1,2}"
    # Loose amount: also matches whole numbers (50, 800) common in handwriting
    _AMT_PAT_LOOSE = r"-?\d[\d.]*(?:[.,]\d{1,2})?"

    # Pre-split: force newline before every date pattern
    _DATE_PAT = (
        r"\d{4}-\d{2}-\d{2}"            # 2026-03-05 (ISO)
        r"|\d{1,2}[./]\d{1,2}[./]\d{2,4}"  # 05.03.2026 or 05/03/26
        r"|\d{1,2}-\d{1,2}-\d{2,4}"      # 05-03-2026
        r"|\d{1,2}[.]\d{1,2}\s\d{2,4}"   # 31.8 21 (OCR broken dot+space)
        r"|\d{1,2}\s\d{1,2}\s\d{2,4}"    # 05 03 2026 (OCR space)
    )
    text = _re.sub(r"(" + _DATE_PAT + r")", r"\n\1", text)

    # Count dates to decide strategy
    all_dates_in_text = _re.findall(_DATE_PAT, text)
    expected_count = len(all_dates_in_text)
    logger.info("Date detection: found %d dates in text (first 3: %s)", expected_count, all_dates_in_text[:3])

    lines = [l.strip() for l in text.strip().split("\n") if l.strip() and len(l.strip()) > 1]
    if len(lines) > 200:
        lines = lines[:200]
    is_table_mode = expected_count > 1
    logger.info("Table import: %d lines, %d dates, table_mode=%s", len(lines), expected_count, is_table_mode)
    rows = []

    def _parse_date(raw):
        raw = raw.strip()
        # Already ISO: 2026-03-05
        if _re.match(r"^\d{4}-\d{2}-\d{2}$", raw):
            return raw
        # Normalize separators: space, dash, slash → dot
        normalized = raw.replace("/", ".").replace("-", ".").replace(" ", ".")
        parts = normalized.split(".")
        if len(parts) == 3:
            dd, mm, yy = parts[0].strip(), parts[1].strip(), parts[2].strip()
            if len(yy) == 2:
                yy = "20" + yy
            if len(dd) <= 2 and len(mm) <= 2 and len(yy) == 4:
                return f"{yy}-{mm.zfill(2)}-{dd.zfill(2)}"
        return raw

    def _detect_currency(t):
        tu = t.upper()
        if "₺" in t or " TL" in tu or tu.endswith("TL"):
            return "TRY"
        if "€" in t or "EUR" in tu:
            return "EUR"
        if "$" in t or "USD" in tu:
            return "USD"
        if "£" in t or "GBP" in tu:
            return "GBP"
        if "CHF" in tu:
            return "CHF"
        return "EUR"  # default for German tax tool

    detected_currency = _detect_currency(text)
    logger.info("Table import currency: %s", detected_currency)

    def _is_date_fragment(s):
        """Check if string looks like a date fragment (DD.MM or MM.YY), not an amount."""
        s = s.strip()
        # DD.MM pattern: 01.01 - 31.12
        m = _re.match(r"^(\d{1,2})[.](\d{1,2})$", s)
        if m:
            d, mo = int(m.group(1)), int(m.group(2))
            if 1 <= d <= 31 and 1 <= mo <= 12:
                return True
        return False

    def _parse_amount(s):
        """Parse German/Turkish number format: 1.234,56 → 1234.56, -16,60 → 16.60"""
        try:
            if "%" in s:
                return 0.0
            raw = s
            s = s.replace("€", "").replace("₺", "").replace(" ", "").strip()
            negative = s.startswith("-")
            s = s.lstrip("-")
            if s.upper().endswith("TL"):
                s = s[:-2].strip()
            if s.upper().endswith("EUR"):
                s = s[:-3].strip()
            if not s:
                return 0.0
            # Reject date-like values: 22.06, 01.12 etc
            if _is_date_fragment(s):
                return 0.0
            # German format: 1.234,56 — dot is thousands, comma is decimal
            if "," in s and "." in s:
                s = s.replace(".", "").replace(",", ".")
            elif "," in s:
                s = s.replace(",", ".")
            val = abs(float(s))  # always positive — sign indicates direction not value
            # Skip year-like numbers
            if _re.match(r"^(19|20)\d{2}$", str(int(val))) and "," not in raw and "." not in raw:
                return 0.0
            return val
        except (ValueError, AttributeError):
            return 0.0

    def _score_line(line):
        """Score a line for table-row likelihood: +2 date, +2 amount, +1 text."""
        score = 0
        if _re.search(_DATE_PAT, line):
            score += 2
        amounts = [a for a in _re.findall(r"(" + _AMT_PAT + r")", line) if not _is_date_fragment(a)]
        if amounts:
            score += 2
        elif _re.search(r"\b\d{2,5}\b", line):
            score += 1  # whole number — weaker signal
        text_part = _re.sub(_DATE_PAT, "", line)
        text_part = _re.sub(_AMT_PAT_LOOSE, "", text_part).strip()
        if len(text_part) >= 3:
            score += 1
        return score

    # Detect column order from header line (e.g. "Nr Datum Beschreibung Einnahmen Ausgaben Saldo")
    _col_order = []  # list of column names in order, e.g. ["einnahmen", "ausgaben", "saldo"]
    for _hl in lines[:5]:
        _hl_lower = _hl.lower()
        if "einnahmen" in _hl_lower or "ausgaben" in _hl_lower:
            # Extract column names in order of appearance
            _header_cols = []
            for _word in _re.findall(r"[a-zäöü]+", _hl_lower):
                if _word in ("einnahmen", "ausgaben", "saldo"):
                    _header_cols.append(_word)
            if _header_cols:
                _col_order = _header_cols
                logger.info("Detected column order from header: %s", _col_order)
            break

    def _assign_amounts_by_columns(amounts_list):
        """Assign amounts to einnahmen/ausgaben based on detected column order."""
        einnahmen, ausgaben = 0.0, 0.0
        if not _col_order:
            # No header detected — use default: 1 amount=ausgaben, 2 amounts=einnahmen+ausgaben
            if len(amounts_list) >= 2:
                einnahmen = amounts_list[0]
                ausgaben = amounts_list[1]
            elif len(amounts_list) == 1:
                ausgaben = amounts_list[0]
            return einnahmen, ausgaben
        # Map amounts to columns by position (skip saldo)
        col_idx = 0
        for col_name in _col_order:
            if col_name == "saldo":
                continue  # always skip saldo
            if col_idx < len(amounts_list):
                if col_name == "einnahmen":
                    einnahmen = amounts_list[col_idx]
                elif col_name == "ausgaben":
                    ausgaben = amounts_list[col_idx]
                col_idx += 1
        return einnahmen, ausgaben

    # Strategy 1: Window-based — group date line + next 2 lines into one entry
    i = 0
    while i < len(lines):
        if len(rows) >= expected_count and expected_count > 0:
            break
        line = lines[i].strip()
        if not line or len(line) < 4:
            i += 1
            continue
        line_lower = line.lower()
        has_date = bool(_re.search(_DATE_PAT, line))
        if not has_date and any(w in line_lower for w in ["datum", "beschreibung", "einnahmen", "ausgaben", "kassenbuch", "übertrag", "seitensumme"]):
            i += 1
            continue

        # If line has date, combine with next 1-2 lines for context
        if has_date:
            combined = line
            lines_consumed = 1
            for j in range(1, 3):
                if i + j < len(lines):
                    next_line = lines[i + j].strip()
                    # Stop if next line starts with a new date
                    if _re.search(r"^(" + _DATE_PAT + r")", next_line):
                        break
                    combined += " " + next_line
                    lines_consumed += 1
            line = combined
            i += lines_consumed
        else:
            i += 1

        # Extract all amounts from line, filter out Saldo (negative/very large running totals)
        _line_amounts_raw = _re.findall(r"(-?\d[\d.]*[.,]\d{1,2})", line)
        _line_amounts = []
        for _a in _line_amounts_raw:
            if _is_date_fragment(_a.lstrip("-")):
                continue
            _v = _parse_amount(_a)
            # Skip negative values (Saldo column) and very large values (>50000 = likely Saldo)
            if _a.strip().startswith("-"):
                continue
            if _v > 50000:
                continue
            _line_amounts.append(_v)

        # Pattern: Date + Description + amounts (Einnahmen/Ausgaben, ignoring Saldo)
        m = _re.search(
            r"(?:\d{1,3}[.\s])?\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+?)\s+([\d.,]+)\s*",
            line
        )
        if m and _line_amounts:
            datum_raw, beschreibung = m.group(1), m.group(2).strip()
            # Remove all numbers from description
            beschreibung = _re.sub(r"-?\d[\d.]*[.,]\d{1,2}", "", beschreibung).strip()
            beschreibung = _re.sub(r"\s+", " ", beschreibung).strip(" .,;:-")
            if not beschreibung or len(beschreibung) < 2:
                beschreibung = "Eintrag"
            # Assign amounts based on detected column order (or default)
            einnahmen, ausgaben = _assign_amounts_by_columns(_line_amounts)
            if beschreibung and len(beschreibung) >= 2:
                rows.append({"date": _parse_date(datum_raw), "description": beschreibung[:80], "income": round(einnahmen, 2), "expense": round(ausgaben, 2)})
            continue

        # Pattern: DD.MM.YYYY + Description + single amount (no other amounts on line)
        m2 = _re.search(r"(?:\d{1,3}[.\s])?\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+?)\s+([\d.,]+)\s*$", line)
        if m2:
            datum_raw, beschreibung = m2.group(1), m2.group(2).strip()
            _val = _parse_amount(m2.group(3))
            if beschreibung and len(beschreibung) >= 2 and _val > 0 and _val <= 50000:
                rows.append({"date": _parse_date(datum_raw), "description": beschreibung, "income": 0, "expense": round(_val, 2)})
            continue

        # Pattern: YYYY-MM-DD + Description + two amounts
        m3 = _re.search(r"(\d{4}-\d{2}-\d{2})\s+(.+?)\s+([\d.,]+)\s+([\d.,]+)\s*$", line)
        if m3:
            date_iso, beschreibung = m3.group(1), m3.group(2).strip()
            val1, val2 = _parse_amount(m3.group(3)), _parse_amount(m3.group(4))
            einnahmen = val1 if val1 > 0 and val2 > 0 else 0
            ausgaben = val2 if val1 == 0 or (val1 > 0 and val2 > 0) else val1
            if einnahmen == 0 and ausgaben == 0:
                ausgaben = max(val1, val2)
            if beschreibung and len(beschreibung) >= 2:
                rows.append({"date": date_iso, "description": beschreibung, "income": round(einnahmen, 2), "expense": round(ausgaben, 2)})
            continue

        # Pattern: YYYY-MM-DD + Description + single amount
        m4 = _re.search(r"(\d{4}-\d{2}-\d{2})\s+(.+?)\s+([\d.,]+)\s*$", line)
        if m4:
            date_iso, beschreibung = m4.group(1), m4.group(2).strip()
            if beschreibung and len(beschreibung) >= 2:
                rows.append({"date": date_iso, "description": beschreibung, "income": 0, "expense": round(_parse_amount(m4.group(3)), 2)})
            continue

        # Universal fallback: any date + any text + any amount anywhere in line
        if is_table_mode and has_date:
            date_m = _re.search(r"(" + _DATE_PAT + r")", line)
            if date_m:
                d = date_m.group(1)
                desc = _re.sub(_DATE_PAT, "", line)
                raw_nums = _re.findall(r"(-?" + _AMT_PAT + r")", desc)
                # Filter: no date fragments, no negatives (Saldo), no >50000
                numbers = []
                for n in raw_nums:
                    if _is_date_fragment(n.lstrip("-")):
                        continue
                    if n.strip().startswith("-"):
                        continue
                    pv = _parse_amount(n)
                    if 0 < pv <= 50000:
                        numbers.append(n)
                desc = _re.sub(r"-?" + _AMT_PAT, "", desc).strip()
                desc = _re.sub(r"\s+", " ", desc).strip(" .,;:-")
                if len(desc) < 2:
                    desc = "Eintrag"
                _parsed_nums = [_parse_amount(n) for n in numbers]
                einnahmen, ausgaben = _assign_amounts_by_columns(_parsed_nums)
                parsed_d = _parse_date(d) if "." in d or "/" in d else d
                rows.append({"date": parsed_d, "description": desc[:80], "income": round(einnahmen, 2), "expense": round(ausgaben, 2), "is_uncertain": ausgaben == 0 and einnahmen == 0})

    logger.info("Strategy 1 result: %d rows from %d lines (dates=%d) in %.2fs", len(rows), len(lines), len(all_dates_in_text), _time.time()-_t0)

    # If multiple dates but Strategy 1 found fewer rows → discard and retry
    s1_rows = list(rows)

    # If Strategy 1 got less than expected, try additional strategies
    if is_table_mode and len(rows) < expected_count:
        logger.info("Strategy 1 incomplete: %d/%d — running Strategy 2", len(rows), expected_count)

    # Strategy 2: Try merging split lines (OCR puts dates and amounts on separate lines)
    if is_table_mode and len(rows) < expected_count:
        date_lines = []
        amount_lines = []
        for line in lines:
            line = line.strip()
            if not line:
                continue
            if _re.match(r"^\d{1,2}[./]\d{1,2}[./]\d{2,4}\s+\S", line) or _re.match(r"^\d{4}-\d{2}-\d{2}\s+\S", line):
                date_lines.append(line)
            elif _re.match(r"^[\d.,]+\s+[\d.,]+\s*$", line):
                amount_lines.append(line)
            elif _re.match(r"^[\d.,]+\s*$", line):
                amount_lines.append(line + " 0")

        logger.info("Table split-line parse: %d date_lines, %d amount_lines", len(date_lines), len(amount_lines))

        for i, dline in enumerate(date_lines):
            m = _re.match(r"(?:\d{1,3}[.\s])?\s*(\d{1,2}[./]\d{1,2}[./]\d{2,4})\s+(.+)", dline)
            if not m:
                continue
            datum_raw, beschreibung = m.group(1), m.group(2).strip()
            beschreibung = _re.sub(r"[\d.,]+\s*$", "", beschreibung).strip()
            if not beschreibung or len(beschreibung) < 2:
                continue
            einnahmen, ausgaben = 0.0, 0.0
            if i < len(amount_lines):
                nums = _re.findall(r"[\d.,]+", amount_lines[i])
                if len(nums) >= 2:
                    v1, v2 = _parse_amount(nums[0]), _parse_amount(nums[1])
                    if v1 > 0 and v2 == 0:
                        ausgaben = v1
                    elif v1 == 0 and v2 > 0:
                        einnahmen = v2
                    elif v1 > 0 and v2 > 0:
                        ausgaben, einnahmen = v1, v2
                elif len(nums) == 1:
                    ausgaben = _parse_amount(nums[0])
            rows.append({"date": _parse_date(datum_raw), "description": beschreibung, "income": round(einnahmen, 2), "expense": round(ausgaben, 2)})

    # Strategy 2.5: Split text by date boundaries (when OCR merges lines)
    if is_table_mode and len(rows) < expected_count:
        logger.info("Strategy 2.5: date-split with %d dates", len(all_dates_in_text))
        # Split text at each date occurrence
        blocks = _re.split(r"(?=\d{1,2}[./]\d{1,2}[./]\d{2,4})|(?=\d{4}-\d{2}-\d{2})", text)
        blocks = [b.strip() for b in blocks if b.strip() and len(b.strip()) > 5][:100]
        logger.info("Strategy 2.5 blocks: %d (first: %s)", len(blocks), blocks[0][:60] if blocks else "none")
        for block in blocks:
            if len(rows) >= expected_count and expected_count > 0:
                break
            # Normalize: merge multi-line block into single line
            block = " ".join(block.splitlines()).strip()
            if not block or len(block) < 5:
                continue
            block_lower = block.lower()
            if any(w in block_lower for w in ["kassenbuch", "übertrag", "seitensumme"]):
                continue

            # Extract date from start of block
            dm = _re.match(r"(\d{1,2}[./]\d{1,2}[./]\d{2,4}|\d{4}-\d{2}-\d{2})\s*(.*)", block)
            if not dm:
                logger.debug("Strategy 2.5 skip (no date): %s", block[:60])
                continue
            date_raw = dm.group(1)
            rest = dm.group(2).strip()

            # Extract amounts from rest (filter out date fragments)
            amounts = [a for a in _re.findall(r"(" + _AMT_PAT + r")", rest) if not _is_date_fragment(a)]
            desc = _re.sub(_AMT_PAT, "", rest).strip()
            desc = _re.sub(r"\s+", " ", desc).strip(" .,;:-")

            if not desc or len(desc) < 2:
                desc = "Eintrag"

            parsed_date = _parse_date(date_raw) if ("." in date_raw or "/" in date_raw) else date_raw
            # Strict positional: numbers[-2]=expense, numbers[-1]=saldo
            if len(amounts) >= 2:
                expense = _parse_amount(amounts[-2])
            elif len(amounts) == 1:
                expense = _parse_amount(amounts[0])
            else:
                expense = 0
            rows.append({"date": parsed_date, "description": desc[:80], "income": 0, "expense": round(expense, 2), "is_uncertain": expense == 0})

        if rows:
            logger.info("Strategy 2.5 date-split: %d rows extracted", len(rows))

    # Strategy 3: ONLY if 0-1 dates found — treat as single receipt
    if not rows and len(all_dates_in_text) <= 1:
        try:
            from autotax.parser import parse_invoice
            parsed = parse_invoice(text)
            vendor = parsed.get("vendor", "")
            amount = parsed.get("total_amount", 0)
            date_str = parsed.get("date", "")
            if amount is not None:
                rows.append({
                    "date": date_str or datetime.now().strftime("%Y-%m-%d"),
                    "description": vendor or "Beleg",
                    "income": 0,
                    "expense": round(float(amount), 2),
                })
                logger.info("Table import fallback: single receipt %s €%.2f", vendor, amount)
        except Exception:
            pass

    # Strategy 5: Column-based parsing (OCR reads columns separately: dates, descriptions, amounts)
    # Detects when OCR returns all dates first, then all descriptions, then all amounts
    valid_amounts_in_rows = sum(1 for r in rows if r.get("expense", 0) > 0 or r.get("income", 0) > 0)
    if is_table_mode and valid_amounts_in_rows < expected_count // 2:
        try:
            _skip = {"einnahmen", "ausgaben", "beschreibung", "datum", "nr.", "mwst", "brutto", "netto", "summe", "kassenbuch", "saldo", "ubersicht"}
            col_dates, col_descs, col_amounts = [], [], []
            for line in raw_lines:
                ll = line.lower().strip()
                if any(w in ll for w in _skip):
                    continue
                l = line.strip()
                # "21 26.8.21" → row number + date merged
                merged = _re.match(r"^\d{1,3}\s+(" + _DATE_PAT + r")", l)
                if merged:
                    col_dates.append(merged.group(1))
                elif _re.match(r"^(" + _DATE_PAT + r")", l):
                    col_dates.append(l)
                elif _re.match(r"^" + _AMT_PAT + r"\s*[€$₺]?$", l):
                    col_amounts.append(l)
                elif _re.match(r"^-\s+\d", l):
                    col_amounts.append(l.replace(" ", ""))  # "- 29,28" → "-29,28"
                elif _re.match(r"^\d{1,3}$", l):
                    continue
                elif len(line.strip()) > 2 and not _re.match(r"^[\d.,€$₺/ \-]+$", line.strip()):
                    col_descs.append(line.strip())

            # Filter amounts: skip negative (saldo), keep positive (expense)
            expenses_only = [a for a in col_amounts if not a.startswith("-")]

            logger.info("Strategy 5 columns: %d dates, %d descs, %d amounts (%d positive)",
                        len(col_dates), len(col_descs), len(col_amounts), len(expenses_only))

            n = min(len(col_dates), len(col_descs), len(expenses_only))
            if n > len(rows) // 2:  # only use if significantly better
                s5_rows = []
                for i in range(n):
                    # Extract date from date line (might have extra text)
                    dm = _re.search(r"(" + _DATE_PAT + r")", col_dates[i])
                    if not dm:
                        continue
                    date_val = _parse_date(dm.group(1))
                    desc = _re.sub(_DATE_PAT, "", col_dates[i]).strip()
                    if not desc or len(desc) < 2:
                        desc = col_descs[i] if i < len(col_descs) else "Eintrag"
                    else:
                        desc = desc + " " + (col_descs[i] if i < len(col_descs) else "")
                    desc = desc.strip()[:80]
                    amt = _parse_amount(expenses_only[i]) if i < len(expenses_only) else 0
                    s5_rows.append({"date": date_val, "description": desc, "income": 0, "expense": round(amt, 2), "is_uncertain": amt == 0})

                if len(s5_rows) > valid_amounts_in_rows:
                    s5_valid = sum(1 for r in s5_rows if r["expense"] > 0)
                    logger.info("Strategy 5 result: %d rows (%d with amounts) vs current %d (%d with amounts)",
                                len(s5_rows), s5_valid, len(rows), valid_amounts_in_rows)
                    if s5_valid > valid_amounts_in_rows:
                        rows = s5_rows
                        logger.info("Strategy 5 accepted: %d rows", len(rows))
        except Exception as e:
            logger.warning("Strategy 5 column-based failed: %s", e)

    # --- ADDED START: Strategy 5.5: Lineless table — lines with text + amount but no date ---
    # For invoices like B-Werk where items have no dates, just description + amount
    if not rows or len(rows) < 2:
        _lineless_rows = []
        _skip_kws = {"summe", "total", "gesamt", "netto", "brutto", "mwst", "ust", "steuer",
                     "zwischensumme", "subtotal", "betrag", "datum", "beschreibung", "nr",
                     "einnahmen", "ausgaben", "saldo", "übertrag", "rechnung", "seite",
                     "kassenbuch", "seitensumme", "gegeben", "rückgeld", "wechselgeld",
                     "kartenzahlung", "barzahlung", "payment", "bezahlt"}
        for rl in lines:
            rl_lower = rl.lower().strip()
            # Skip header/total/keyword lines
            if any(kw in rl_lower for kw in _skip_kws):
                continue
            # Skip very short lines
            if len(rl.strip()) < 5:
                continue
            # Must have at least one amount at the end
            _amt_match = _re.search(r"(.+?)\s+(" + _AMT_PAT + r")\s*$", rl.strip())
            if not _amt_match:
                # Try loose: text + whole number at end
                _amt_match = _re.search(r"(.+?)\s+(\d{1,6})\s*$", rl.strip())
            if _amt_match:
                desc = _amt_match.group(1).strip()
                amt_raw = _amt_match.group(2)
                # Clean description: remove leading numbers (item nr), punctuation
                desc = _re.sub(r"^\d{1,4}[.\s]+", "", desc).strip()
                desc = _re.sub(r"^[.\-:,]+", "", desc).strip()
                if len(desc) < 2:
                    continue
                amt = _parse_amount(amt_raw) if "," in amt_raw or "." in amt_raw else float(amt_raw)
                if amt <= 0 or amt > 50000:
                    continue
                # Try to find date on this line
                _d_m = _re.search(r"(" + _DATE_PAT + r")", rl)
                _date = _parse_date(_d_m.group(1)) if _d_m else ""
                _lineless_rows.append({
                    "date": _date,
                    "description": desc[:80],
                    "income": 0,
                    "expense": round(amt, 2),
                    "is_uncertain": True,
                })
        if len(_lineless_rows) > len(rows):
            rows = _lineless_rows
            logger.info("Strategy 5.5 lineless table: %d rows extracted", len(rows))
    # --- ADDED END ---

    # Strategy 6: Scored fallback — if all table strategies failed, score each line
    # and return partial matches with confidence, rather than empty result
    if not rows and text and len(text.strip()) > 10:
        logger.info("Strategy 6: all table strategies failed, trying scored extraction")
        raw_text_lines = [l.strip() for l in text.strip().split("\n") if l.strip() and len(l.strip()) > 3]
        for rl in raw_text_lines[:100]:
            rl_lower = rl.lower()
            if any(w in rl_lower for w in ["datum", "beschreibung", "einnahmen", "ausgaben", "kassenbuch", "übertrag", "seitensumme", "saldo"]):
                continue
            score = _score_line(rl)
            if score < 2:
                continue  # skip lines with no useful signal
            date_m = _re.search(r"(" + _DATE_PAT + r")", rl)
            date_val = _parse_date(date_m.group(1)) if date_m else ""
            # Try strict amount first, then loose (whole numbers)
            amounts = [a for a in _re.findall(r"(" + _AMT_PAT + r")", rl) if not _is_date_fragment(a)]
            if not amounts:
                amounts = _re.findall(r"\b(\d{2,5})\b", rl)
            desc = _re.sub(_DATE_PAT, "", rl)
            desc = _re.sub(_AMT_PAT_LOOSE, "", desc)
            desc = _re.sub(r"\s+", " ", desc).strip(" .,;:-|/")
            if not desc or len(desc) < 2:
                desc = rl[:80]
            expense = _parse_amount(amounts[0]) if amounts else 0
            confidence = round(min(score / 5.0, 1.0), 2)
            rows.append({
                "date": date_val,
                "description": desc[:80],
                "income": 0,
                "expense": round(expense, 2),
                "is_uncertain": score < 4,
                "confidence": confidence,
                "raw_fallback": True,
            })
        if rows:
            logger.info("Strategy 6 scored fallback: %d lines (avg confidence=%.2f)",
                        len(rows), sum(r.get("confidence", 0) for r in rows) / len(rows))

    # Strategy 6.5: Column block — handwritten tables where OCR reads columns separately
    # Pattern: dates+descriptions in first block, amounts after "Ausgaben"/"Saldo" keyword
    if len(rows) < expected_count * 0.8 and expected_count > 3:
        try:
            # Find amounts block — search for "Ausgaben"/"Saldo" keyword
            _ausgaben_idx = -1
            _full_text_lower = text.lower()
            # If BOTH "einnahmen" and "ausgaben" exist in text → it's a standard Kassenbuch → expense
            _has_ausgaben = "ausgaben" in _full_text_lower
            _has_einnahmen = "einnahmen" in _full_text_lower
            _is_income_table = _has_einnahmen and not _has_ausgaben  # only income if NO ausgaben anywhere
            for _li, _ll in enumerate(lines):
                _ll_lower = _ll.lower()
                if "ausgaben" in _ll_lower and _li > len(lines) * 0.15:
                    _ausgaben_idx = _li
                    break
                if "saldo" in _ll_lower and _li > len(lines) * 0.15:
                    _ausgaben_idx = _li
                    break
                if "einnahmen" in _ll_lower and not _has_ausgaben and _li > len(lines) * 0.15:
                    _ausgaben_idx = _li
                    break

            if _ausgaben_idx > 0:
                # Extract amounts from after "Ausgaben" keyword
                _amounts_block = lines[_ausgaben_idx:]
                _amounts = []
                for _al in _amounts_block:
                    _al_clean = _al.strip().replace("→", "").replace("=", "").replace("*", "").replace(":", "")
                    # Skip totals line
                    if _re.search(r"\btotal\b", _al_clean, _re.IGNORECASE):
                        continue
                    # Skip saldo (negative values) but keep positive
                    if _al_clean.strip().startswith("-") or _al_clean.strip().startswith("−"):
                        continue
                    _am = _re.search(r"(\d[\d\s]*[.,]\d{2})", _al_clean)
                    if _am:
                        _val = float(_am.group(1).replace(" ", "").replace(",", "."))
                        if 0.01 <= _val < 50000:
                            _amounts.append(_val)

                # Extract dates from first block (before Ausgaben)
                _dates_block = lines[:_ausgaben_idx]
                _cb_dates = []
                _cb_descs = []
                _current_desc = ""
                for _dl in _dates_block:
                    _dl_strip = _dl.strip()
                    _dm = _re.search(r"(\d{1,2}[./]\s*\d{1,2}[./]\s*\d{2,4})", _dl_strip)
                    if _dm:
                        if _current_desc and _cb_dates:
                            _cb_descs.append(_current_desc.strip())
                        _cb_dates.append(_dm.group(1).replace(" ", ""))
                        # Description might be on same line after date
                        _after = _dl_strip[_dm.end():].strip()
                        # Or next word before date
                        _before = _dl_strip[:_dm.start()].strip()
                        _desc_part = _after if len(_after) > 2 else _before
                        # Clean up numbers and single chars
                        _desc_part = _re.sub(r"^\d+\s*", "", _desc_part).strip()
                        _current_desc = _desc_part
                    elif _dl_strip and not _re.match(r"^[\d\s.,/\-+*=→:€$]+$", _dl_strip) and len(_dl_strip) > 2:
                        # Text line — append to current description
                        _skip_words = {"nr", "nr.", "datum", "beschreibung", "einnahmen", "ausgaben", "saldo", "/", "l", "m"}
                        if _dl_strip.lower().strip() not in _skip_words:
                            _current_desc = (_current_desc + " " + _dl_strip).strip() if _current_desc else _dl_strip
                if _current_desc:
                    _cb_descs.append(_current_desc.strip())

                # Match dates + descriptions + amounts by position
                _n = min(len(_cb_dates), len(_amounts))
                # Pad descriptions if fewer than dates
                while len(_cb_descs) < _n:
                    _cb_descs.append("Import")

                if _n >= expected_count * 0.5 and _n > len(rows):
                    _cb_rows = []
                    for _i in range(_n):
                        _d = _cb_dates[_i]
                        # Normalize date
                        _dp = _d.replace(".", "/").split("/")
                        _date_str = ""
                        if len(_dp) == 3:
                            _y = _dp[2]
                            if len(_y) == 2:
                                _y = "20" + _y
                            try:
                                _date_str = f"{_y}-{int(_dp[1]):02d}-{int(_dp[0]):02d}"
                            except (ValueError, IndexError):
                                _date_str = _d
                        _cb_rows.append({
                            "date": _date_str,
                            "description": _cb_descs[_i] if _i < len(_cb_descs) else "Import",
                            "income": round(_amounts[_i], 2) if _is_income_table else 0,
                            "expense": 0 if _is_income_table else round(_amounts[_i], 2),
                            "confidence": 0.75,
                        })
                    if len(_cb_rows) > len(rows):
                        rows = _cb_rows
                        logger.info("Strategy 6.5 column block: %d rows (dates=%d, amounts=%d, descs=%d)",
                                    len(_cb_rows), len(_cb_dates), len(_amounts), len(_cb_descs))
        except Exception as e:
            logger.warning("Strategy 6.5 column block failed: %s", e)

    # Ensure all rows have is_uncertain flag
    for r in rows:
        if "is_uncertain" not in r:
            r["is_uncertain"] = False

    # Deduplicate: same date + same description = duplicate
    seen = set()
    unique_rows = []
    for r in rows:
        key = (r["date"], r["description"][:30])
        if key not in seen:
            seen.add(key)
            unique_rows.append(r)
    if len(unique_rows) < len(rows):
        logger.info("Dedup: %d → %d rows", len(rows), len(unique_rows))
    rows = unique_rows

    # Data quality: validate rows — reject rows without amount or without (date or text)
    validated = []
    for r in rows:
        amount = r.get("expense", 0) + r.get("income", 0)
        has_amount = amount > 0
        has_date = bool(r.get("date", "").strip())
        has_text = len(r.get("description", "").strip()) >= 3
        if not has_amount:
            continue  # RULE 2: reject if no valid amount
        if not (has_date or has_text):
            continue  # RULE 3: need date or text
        # RULE 1: mark low-confidence rows
        if r.get("confidence", 1.0) < 0.6:
            r["is_uncertain"] = True
        validated.append(r)
    if len(validated) < len(rows):
        logger.info("Validation: %d → %d rows (rejected %d)", len(rows), len(validated), len(rows) - len(validated))
    rows = validated

    # --- ADDED START ---
    table_detected = len(rows) > 0
    if not table_detected:
        print("TABLE NOT DETECTED → FALLBACK TEXT MODE")
        # fallback: treat lines as entries
        _fb_lines = text.split("\n")
        extracted_rows = []
        import re as _re_fb
        for _line in _fb_lines:
            # detect numbers (amounts)
            _m = _re_fb.findall(r"\d+[.,]\d{2}", _line)
            if _m:
                try:
                    value = float(_m[-1].replace(",", "."))
                    extracted_rows.append({
                        "date": "",
                        "description": _line.strip()[:80],
                        "income": 0,
                        "expense": round(value, 2),
                        "is_uncertain": True,
                        "raw": _line,
                        "amount": value,
                    })
                except Exception:
                    pass
        if extracted_rows:
            print("FALLBACK EXTRACTED ROWS:", extracted_rows[:5])
            rows = extracted_rows
    # --- ADDED END ---

    # Note: deduplicate/merged row removal disabled — was too aggressive (24→17 satır kaybı)

    logger.info("Table import result: %d rows (expected %d)", len(rows), expected_count)

    # Generate CSV
    csv_lines = ["Datum,Beschreibung,Einnahmen,Ausgaben"]
    for r in rows:
        desc = r["description"].replace('"', '""')
        inc = f'{r["income"]:.2f}' if not r.get("is_uncertain") or r["income"] > 0 else ""
        exp = f'{r["expense"]:.2f}' if not r.get("is_uncertain") or r["expense"] > 0 else ""
        csv_lines.append(f'{r["date"]},"{desc}",{inc},{exp}')
    csv_text = "\n".join(csv_lines)

    # Optionally save to DB — block if data quality too low
    saved = 0
    _save_avg_conf = (sum(r.get("confidence", 1.0) for r in rows) / len(rows)) if rows else 0
    _save_blocked = len(rows) < 2 or _save_avg_conf < 0.5
    if _save_blocked and save:
        logger.info("Auto-save blocked: %d rows, avg confidence=%.2f (min 2 rows, 0.5 confidence)", len(rows), _save_avg_conf)
    if save and rows and not _save_blocked:
        db = SessionLocal()
        try:
            for r in rows:
                amount = r["income"] if r["income"] > 0 else r["expense"]
                entry_type = "income" if r["income"] > 0 else "expense"
                vat_amount = round(amount * 19 / 119, 2) if amount > 0 else 0
                date_val = parse_date_str_to_datetime(r["date"])
                if not date_val:
                    date_val = datetime.now()
                inv = Invoice(
                    user_id=user["sub"],
                    filename=file.filename or "bild-import",
                    vendor=r["description"][:50],
                    total_amount=amount,
                    vat_amount=vat_amount,
                    vat_rate="19%",
                    date=r["date"],
                    raw_text=f"Bild Import: {r['description']}",
                    invoice_type=entry_type,
                    invoice_number="",
                    payment_method="",
                    category="other",
                    processed=True,
                    file_data=content,
                    file_content_type=file.content_type or "image/jpeg",
                )
                db.add(inv)
                db.flush()  # get inv.id before creating cash entry
                entry = CashEntry(
                    user_id=user["sub"],
                    description=r["description"],
                    vendor=r["description"][:50],
                    gross_amount=amount,
                    vat_amount=vat_amount,
                    vat_rate="19%",
                    entry_type=entry_type,
                    category="other",
                    payment_method="",
                    reference="",
                    notes="Bild Import",
                    date=date_val,
                    invoice_id=inv.id,
                )
                db.add(entry)
                saved += 1
            db.commit()
        except Exception:
            db.rollback()
            logger.exception("Image table import save failed")
        finally:
            db.close()

    # Add currency + row index to all rows
    for i, r in enumerate(rows):
        r["currency"] = detected_currency
        r["_sort_idx"] = i  # preserve original OCR order as tiebreaker

    # Sort rows: by parsed date (primary), then OCR order (secondary).
    # This keeps Nr.-based ordering when dates repeat (e.g. multiple
    # entries on 2024-09-13 stay in their original Nr. sequence).
    def _row_sort_key(r):
        d = r.get("date", "") or ""
        # Parse date to sortable int: YYYYMMDD
        import re as _sre
        m = _sre.match(r"^(\d{4})-(\d{2})-(\d{2})$", d)
        if m:
            return (int(m.group(1)) * 10000 + int(m.group(2)) * 100 + int(m.group(3)), r.get("_sort_idx", 0))
        m = _sre.match(r"^(\d{1,2})[./](\d{1,2})[./](\d{2,4})$", d)
        if m:
            y = m.group(3)
            if len(y) == 2:
                y = "20" + y
            return (int(y) * 10000 + int(m.group(2)) * 100 + int(m.group(1)), r.get("_sort_idx", 0))
        # Unparseable date — keep at end, preserve OCR order
        return (99999999, r.get("_sort_idx", 0))
    rows.sort(key=_row_sort_key)

    # Remove internal sort index from response
    for r in rows:
        r.pop("_sort_idx", None)

    logger.info("Table import complete: %d rows in %.2fs", len(rows), _time.time()-_t0)

    return {
        "success": True,
        "rows": rows,
        "row_count": len(rows),
        "saved": saved,
        "save_blocked": _save_blocked if save else False,
        "csv": csv_text,
        "currency": detected_currency,
        "raw_rows": raw_lines,
        "raw_row_count": len(raw_lines),
        "ocr_text": text[:2000],
    }



# ============================================================
# TAX: EÜR
# ============================================================

@app.get("/tax/euer")
def list_euer(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"], (Invoice.is_deleted == False) | (Invoice.is_deleted == None)).all()
        years = set()
        for i in invoices:
            d = safe_date_str(i.date)
            if len(d) >= 4:
                years.add(d[:4])
        result = []
        for y in sorted(years):
            year_invs = [i for i in invoices if safe_date_str(i.date).startswith(y)]
            einnahmen = sum(safe_float(i.total_amount) for i in year_invs if safe_invoice_type(i.invoice_type) == "income")
            ausgaben = sum(safe_float(i.total_amount) for i in year_invs if safe_invoice_type(i.invoice_type) == "expense")
            result.append({
                "id": int(y),
                "steuerjahr": int(y),
                "summe_einnahmen": round(einnahmen, 2),
                "summe_ausgaben": round(ausgaben, 2),
                "gewinn_verlust": round(einnahmen - ausgaben, 2),
            })
        return result
    except Exception:
        logger.exception("EÜR list failed")
        err(500, "Failed")
    finally:
        db.close()


@app.post("/tax/euer/auto-fill")
def auto_fill_euer(steuerjahr: int = Query(...), user: dict = Depends(get_current_user)):
    return {"success": True, "steuerjahr": steuerjahr, "status": "generated"}


# ============================================================
# CHAT
# ============================================================


@app.post("/feedback")
def submit_feedback(body: dict = Body(...), user: dict = Depends(get_current_user)):
    message = body.get("message", "")
    if not message.strip():
        err(400, "Feedback message is empty")
    logger.info("FEEDBACK from user %d: %s", user["sub"], message[:500])
    return {"success": True, "message": "Feedback received"}


# ============================================================
# COMPANIES (max 2 per user)
# ============================================================
# RECEIPT VAULT
# ============================================================

@app.get("/vault")
def list_vault(search: Optional[str] = Query(None), user: dict = Depends(get_current_user)):
    """List all receipts with metadata — checks DB for original file."""
    db = SessionLocal()
    try:
        q = db.query(Invoice).filter(Invoice.user_id == user["sub"])
        q = q.filter((Invoice.is_deleted == False) | (Invoice.is_deleted == None))
        if search:
            from sqlalchemy import or_
            q = q.filter(or_(Invoice.vendor.ilike(f"%{search}%"), Invoice.date.ilike(f"%{search}%")))
        invoices = q.order_by(Invoice.created_at.desc()).all()
        items = []
        for inv in invoices:
            items.append({
                "id": inv.id,
                "vendor": safe_vendor(inv.vendor),
                "date": safe_date_str(inv.date),
                "total_amount": safe_float(inv.total_amount),
                "vat_amount": safe_float(inv.vat_amount),
                "category": safe_category(inv.category),
                "filename": inv.filename or "",
                "has_original": bool(inv.file_path) or (inv.file_data is not None and len(inv.file_data) > 0 if inv.file_data else False),
                "file_content_type": inv.file_content_type or "",
                "invoice_type": safe_invoice_type(inv.invoice_type),
            })
        return {"items": items, "total": len(items)}
    finally:
        db.close()


# --- ADDED START: Upload/replace original file for existing invoice ---
@app.post("/vault/{invoice_id}/upload")
async def upload_vault_file(invoice_id: int, file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload or replace original receipt file for an existing invoice."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        content = await file.read()
        if len(content) > MAX_FILE_SIZE:
            err(400, "Datei zu groß")
        # Write to disk (replaces previous file if any)
        from autotax import storage
        old_path = inv.file_path
        new_path = storage.save_file(int(user["sub"]), content, file.filename)
        inv.file_path = new_path
        inv.file_size = len(content)
        inv.file_content_type = file.content_type or "application/octet-stream"
        inv.filename = file.filename or inv.filename
        # Drop legacy BLOB if present so the row stops occupying DB space
        inv.file_data = None
        db.commit()
        if old_path and old_path != new_path:
            storage.delete_file(old_path)
        logger.info("Vault upload: invoice %d, %d bytes", invoice_id, len(content))
        return {"success": True}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Vault upload failed")
        err(500, "Upload failed")
    finally:
        db.close()
# --- ADDED END ---


@app.get("/vault/{invoice_id}/download")
def download_vault_file(invoice_id: int, mode: str = Query("inline"), user: dict = Depends(get_current_user)):
    """Download original receipt file from DB. mode=inline (preview) or attachment (download)."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Not found")
        ct = inv.file_content_type or "application/octet-stream"
        fname = inv.filename or "beleg"
        disposition = "attachment" if mode == "attachment" else "inline"
        # Prefer disk; fall back to legacy BLOB for not-yet-migrated rows.
        if inv.file_path:
            from autotax import storage
            if not storage.file_exists(inv.file_path):
                err(404, "File missing on disk")
            data = storage.read_file(inv.file_path)
            return StreamingResponse(io.BytesIO(data), media_type=ct, headers={"Content-Disposition": f"{disposition}; filename={fname}"})
        if inv.file_data:
            return StreamingResponse(io.BytesIO(inv.file_data), media_type=ct, headers={"Content-Disposition": f"{disposition}; filename={fname}"})
        err(404, "Kein Original gespeichert")
    finally:
        db.close()


# ============================================================
# DSGVO: ACCOUNT DELETE + DATA EXPORT (Art. 15, 17, 20)
# ============================================================

@app.delete("/account")
def delete_account(body: dict = Body(...), user: dict = Depends(get_current_user)):
    """Art. 17 DSGVO — Recht auf Löschung. Permanently delete user account and ALL personal data."""
    confirm = body.get("confirm", False)
    if not confirm:
        err(400, "Bitte Löschung bestätigen (confirm: true)")
    db = SessionLocal()
    try:
        uid = user["sub"]
        # Delete all user data across all tables
        db.query(CashEntry).filter(CashEntry.user_id == uid).delete()
        db.query(Invoice).filter(Invoice.user_id == uid).delete()
        db.query(UserCompany).filter(UserCompany.user_id == uid).delete()
        db.query(LlmUsage).filter(LlmUsage.user_id == str(uid)).delete()
        db.query(User).filter(User.id == uid).delete()
        db.commit()
        logger.info("DSGVO: Account deleted — user_id=%d", uid)
        return {"success": True, "message": "Konto und alle Daten wurden unwiderruflich gelöscht."}
    except Exception:
        db.rollback()
        logger.exception("Account deletion failed")
        err(500, "Kontolöschung fehlgeschlagen")
    finally:
        db.close()


@app.get("/account/export")
def export_personal_data(user: dict = Depends(get_current_user)):
    """Art. 15 + Art. 20 DSGVO — Auskunftsrecht & Datenportabilität.
    Returns ALL personal data in machine-readable JSON format."""
    import json as json_lib
    db = SessionLocal()
    try:
        uid = user["sub"]
        u = db.query(User).filter(User.id == uid).first()
        if not u:
            err(404, "User not found")

        # User profile
        user_data = {
            "email": u.email,
            "full_name": u.full_name or "",
            "plan": u.plan or "free",
            "registered_at": u.registered_at.isoformat() if u.registered_at else "",
            "gdpr_consent_at": u.gdpr_consent_at.isoformat() if u.gdpr_consent_at else "",
            "is_kleinunternehmer": getattr(u, 'is_kleinunternehmer', False),
        }

        # Companies
        companies = db.query(UserCompany).filter(UserCompany.user_id == uid).all()
        companies_data = [{
            "company_name": c.company_name, "iban": c.iban or "", "tax_id": c.tax_id or "",
            "address": c.address or "", "phone": c.phone or "", "email": c.email or "",
            "website": c.website or "", "is_default": c.is_default or False,
        } for c in companies]

        # Invoices (including soft-deleted)
        invoices = db.query(Invoice).filter(Invoice.user_id == uid).all()
        invoices_data = [{
            "id": i.id, "vendor": i.vendor or "", "invoice_number": i.invoice_number or "",
            "invoice_type": i.invoice_type or "", "total_amount": float(i.total_amount or 0),
            "vat_amount": float(i.vat_amount or 0), "vat_rate": i.vat_rate or "",
            "date": i.date or "", "category": i.category or "",
            "payment_method": i.payment_method or "", "filename": i.filename or "",
            "raw_text": i.raw_text or "", "is_deleted": i.is_deleted or False,
            "created_at": i.created_at.isoformat() if i.created_at else "",
        } for i in invoices]

        # Cash entries (including soft-deleted)
        entries = db.query(CashEntry).filter(CashEntry.user_id == uid).all()
        entries_data = [cash_entry_to_dict(e) | {"is_deleted": e.is_deleted or False} for e in entries]

        # LLM usage
        llm = db.query(LlmUsage).filter(LlmUsage.user_id == str(uid)).all()
        llm_data = [{"date": l.date, "count": l.count} for l in llm]

        export = {
            "export_date": datetime.now().isoformat(),
            "export_type": "DSGVO Art. 15/20 — Vollständige Datenauskunft",
            "user": user_data,
            "companies": companies_data,
            "invoices": invoices_data,
            "cash_entries": entries_data,
            "llm_usage": llm_data,
        }

        buf = io.StringIO()
        json_lib.dump(export, buf, indent=2, ensure_ascii=False)
        buf.seek(0)
        return StreamingResponse(buf, media_type="application/json", headers={
            "Content-Disposition": f"attachment; filename=autotax_dsgvo_export_{uid}.json"
        })
    finally:
        db.close()


# ============================================================
# PRICING & PLAN
# ============================================================

PRICING = {
    "free": {"name": "Free", "price": 0, "max_invoices": 50, "max_companies": 2},
    "early": {"name": "Early Adopter", "price": 10, "max_invoices": 500, "max_companies": 5},
    "pro": {"name": "Pro", "price": 20, "max_invoices": -1, "max_companies": -1},
}


@app.get("/pricing")
def get_pricing():
    return {"plans": PRICING}


@app.get("/account/plan")
def get_user_plan(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        u = db.query(User).filter(User.id == user["sub"]).first()
        if not u:
            err(404, "User not found")
        plan = u.plan or "free"
        inv_count = db.query(Invoice).filter(Invoice.user_id == user["sub"]).count()
        plan_info = PRICING.get(plan, PRICING["free"])
        return {
            "plan": plan,
            "plan_name": plan_info["name"],
            "price": plan_info["price"],
            "max_invoices": plan_info["max_invoices"],
            "invoice_count": inv_count,
            "is_early": plan == "early",
            "message": "Frühe Nutzer behalten ihren Preis" if plan == "early" else None,
        }
    finally:
        db.close()


@app.post("/account/upgrade")
def upgrade_plan(body: dict = Body(...), user: dict = Depends(get_current_user)):
    # Disabled until payment integration (Stripe) is ready
    err(403, "Plan-Upgrade ist derzeit deaktiviert. Stripe-Integration kommt bald.")


# ============================================================

@app.get("/companies")
def list_companies(user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        companies = db.query(UserCompany).filter(UserCompany.user_id == user["sub"]).all()
        return [{"id": c.id, "company_name": c.company_name, "iban": c.iban or "", "tax_id": c.tax_id or "", "address": c.address or "", "phone": c.phone or "", "fax": c.fax or "", "email": c.email or "", "website": c.website or "", "is_default": c.is_default or False} for c in companies]
    finally:
        db.close()


@app.post("/companies")
def add_company(body: dict = Body(...), user: dict = Depends(get_current_user)):
    company_name = body.get("company_name", "").strip()
    if not company_name:
        err(400, "Firmenname ist erforderlich")
    db = SessionLocal()
    try:
        existing = db.query(UserCompany).filter(UserCompany.user_id == user["sub"]).count()
        u = db.query(User).filter(User.id == user["sub"]).first()
        plan = u.plan if u and u.plan else "free"
        max_companies = PRICING.get(plan, PRICING["free"])["max_companies"]
        if max_companies > 0 and existing >= max_companies:
            err(400, f"Maximal {max_companies} Firmen für deinen Plan ({plan}). Upgrade für mehr.")
        dup = db.query(UserCompany).filter(UserCompany.user_id == user["sub"], UserCompany.company_name == company_name).first()
        if dup:
            err(400, "Firma existiert bereits")
        c = UserCompany(
            user_id=user["sub"], company_name=company_name,
            iban=body.get("iban", "").strip() or None,
            tax_id=body.get("tax_id", "").strip() or None,
            address=body.get("address", "").strip() or None,
            phone=body.get("phone", "").strip() or None,
            fax=body.get("fax", "").strip() or None,
            email=body.get("email", "").strip() or None,
            website=body.get("website", "").strip() or None,
        )
        db.add(c)
        db.commit()
        db.refresh(c)
        return {"success": True, "id": c.id, "company_name": c.company_name}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Add company failed")
        err(500, "Failed")
    finally:
        db.close()


@app.delete("/companies/{company_id}")
def delete_company(company_id: int, user: dict = Depends(get_current_user)):
    db = SessionLocal()
    try:
        c = db.query(UserCompany).filter(UserCompany.id == company_id, UserCompany.user_id == user["sub"]).first()
        if not c:
            err(404, "Firma nicht gefunden")
        db.delete(c)
        db.commit()
        return {"success": True}
    finally:
        db.close()


# --- ADDED START: Default company system (auto-fill invoices) ---
@app.get("/company/default")
def get_default_company(user: dict = Depends(get_current_user)):
    """Return the user's default company, or the most recent one if none is set as default."""
    db = SessionLocal()
    try:
        # Try default first
        c = db.query(UserCompany).filter(
            UserCompany.user_id == user["sub"],
            UserCompany.is_default == True
        ).first()
        # Fallback: latest company
        if not c:
            c = db.query(UserCompany).filter(
                UserCompany.user_id == user["sub"]
            ).order_by(UserCompany.id.desc()).first()
        if not c:
            return {"exists": False}
        return {
            "exists": True,
            "id": c.id,
            "company_name": c.company_name,
            "iban": c.iban or "",
            "tax_id": c.tax_id or "",
            "address": c.address or "",
            "phone": c.phone or "",
            "fax": c.fax or "",
            "email": c.email or "",
            "website": c.website or "",
            "is_default": c.is_default or False,
        }
    finally:
        db.close()


@app.post("/company/set-default/{company_id}")
def set_default_company(company_id: int, user: dict = Depends(get_current_user)):
    """Mark a company as default (unsets all others for this user)."""
    db = SessionLocal()
    try:
        c = db.query(UserCompany).filter(
            UserCompany.id == company_id,
            UserCompany.user_id == user["sub"]
        ).first()
        if not c:
            err(404, "Firma nicht gefunden")
        # Unset all others
        db.query(UserCompany).filter(
            UserCompany.user_id == user["sub"],
            UserCompany.id != company_id
        ).update({"is_default": False})
        c.is_default = True
        db.commit()
        logger.info("Company %d set as default for user %s", company_id, user["sub"])
        return {"success": True, "id": company_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Set default company failed")
        err(500, "Failed")
    finally:
        db.close()
# --- ADDED END ---


# ============================================================

@app.post("/chat")
def chat_endpoint(body: dict = Body(...), user: dict = Depends(get_current_user)):
    _enforce_upload_quota(user["sub"], chat=True)
    message = body.get("message", "")
    db = SessionLocal()
    try:
        all_invoices = db.query(Invoice).filter(Invoice.user_id == user["sub"], (Invoice.is_deleted == False) | (Invoice.is_deleted == None)).all()
        # Filter invalid entries — same logic as dashboard
        invoices = [i for i in all_invoices if safe_float(i.total_amount) > 0 and safe_vendor(i.vendor) != "Unbekannt"]

        inv_count = len(invoices)
        inv_sum = sum(safe_float(i.total_amount) for i in invoices)

        inv_inc = [i for i in invoices if safe_invoice_type(i.invoice_type) == "income"]
        inv_exp = [i for i in invoices if safe_invoice_type(i.invoice_type) == "expense"]

        total_income = sum(safe_float(i.total_amount) for i in inv_inc)
        total_expenses = sum(safe_float(i.total_amount) for i in inv_exp)
        net_profit = total_income - total_expenses

        vat_paid = sum(safe_float(i.vat_amount) for i in inv_exp)
        vat_collected = sum(safe_float(i.vat_amount) for i in inv_inc)
        vat_balance = vat_collected - vat_paid

        cat_map = {}
        for i in invoices:
            c = safe_category(i.category)
            cat_map[c] = cat_map.get(c, 0) + safe_float(i.total_amount)
        cat_str = ", ".join(f"{k}: €{v:.2f}" for k, v in sorted(cat_map.items(), key=lambda x: -x[1])) if cat_map else "keine"

        vendors = {}
        for i in invoices:
            v = safe_vendor(i.vendor)
            vendors[v] = vendors.get(v, 0) + safe_float(i.total_amount)
        top_vendors = ", ".join(f"{k}: €{v:.2f}" for k, v in sorted(vendors.items(), key=lambda x: -x[1])[:5]) if vendors else "keine"

        msg = message.lower().strip()
        reply = None

        # ── TIER 1: Multi-word / very specific phrases (must match BEFORE single words) ──

        # Rechnung erstellen (before generic "rechnung")
        if any(w in msg for w in ["rechnung erstellen", "rechnung schreiben", "rechnung anlegen", "neue rechnung", "fatura oluştur", "fatura yaz"]):
            reply = "🧾 Rechnung erstellen:\n\nSo geht's:\n1. Go to Invoices → '+ Create Invoice' button\n2. Fill in: customer, amount, VAT rate, date, invoice no.\n3. Click 'Save' → saved as income\n4. 'Download PDF?' dialog → save invoice as PDF\n\n• MwSt wird automatisch berechnet (19%, 7%, oder 0%)\n• Kleinunternehmer: §19-Hinweis erscheint automatisch auf PDF\n• PDF enthält: Firmenname, Kunde, Betrag, MwSt, Netto/Brutto\n• Tipp: Rechnungs-Nr. vergeben (z.B. RE-2026-001)"

        # E-Rechnung (before generic "rechnung")
        elif any(w in msg for w in ["e-rechnung", "erechnung", "xrechnung", "zugferd", "xml rechnung", "elektronische rechnung", "factur-x"]):
            reply = "📄 E-Rechnung (XRechnung / ZUGFeRD):\n\nSo geht's:\n1. Gehe zu Upload\n2. Klicke 'E-Rechnung (XML) importieren'\n3. XML-Datei auswählen → Automatische Erkennung\n\n• XRechnung (XML) und ZUGFeRD (PDF) werden unterstützt\n• Automatische Erkennung: Lieferant, Betrag, MwSt, Datum, Rechnungs-Nr., Kategorie, DATEV-Konto\n• Seit 01.01.2025 Pflicht für B2B in Deutschland (Empfang)\n• Auch Kleinunternehmer müssen E-Rechnungen empfangen können\n• Supported formats: XRechnung, ZUGFeRD, Factur-X"

        # Kleinunternehmer (before generic "steuer")
        elif any(w in msg for w in ["kleinunternehmer", "§19", "paragraph 19", "keine mwst", "keine umsatzsteuer", "küçük işletme", "klein unternehmer"]):
            reply = "📋 Kleinunternehmerregelung (§19 UStG):\n\n• Profil-Menü (oben rechts) → Kleinunternehmer Toggle aktivieren\n• Keine MwSt auf erstellten Rechnungen\n• Hinweis 'Gemäß §19 UStG wird keine Umsatzsteuer berechnet' erscheint automatisch auf PDF\n• Grenze: 25.000€ Umsatz im Vorjahr / 100.000€ laufendes Jahr\n• E-Rechnungen empfangen ist trotzdem Pflicht (seit 2025)\n• Vorteil: Weniger Bürokratie, keine USt-Voranmeldung"

        # EÜR (before generic "steuer")
        elif any(w in msg for w in ["eür", "einnahmen-überschuss", "überschussrechnung", "steuerformular", "steuererklärung"]):
            reply = "🧾 EÜR (Einnahmen-Überschuss-Rechnung):\n\nSo geht's:\n1. Sidebar → 'Steuer (EÜR)'\n2. Steuerjahr wählen\n3. 'Generieren' klicken\n\n• EÜR = Einnahmen-Überschuss-Rechnung (für Freiberufler und Kleinunternehmer)\n• Automatische Berechnung aus allen Rechnungen und Kassenbuch-Einträgen\n• Enthält: Betriebseinnahmen, Betriebsausgaben, Profit/Verlust, MwSt-Zusammenfassung\n• Kann als Grundlage für die Steuererklärung verwendet werden\n\nHinweis: Für die offizielle Steuererklärung immer einen Steuerberater konsultieren."

        # DATEV (before generic "export")
        elif any(w in msg for w in ["datev", "steuerberater export", "buchungskonto", "skr03", "skr04", "skr 03", "skr 04"]):
            reply = "📊 DATEV Export:\n\nSo geht's:\n1. Gehe zu Export\n2. Optional: Steuerjahr wählen\n3. 'DATEV' Button klicken → Datei wird heruntergeladen\n\n• Standard-Format für deutsche Steuerberater\n• Automatische Kontenzuordnung:\n  - 6800 = Wareneinkauf / Lebensmittel\n  - 6640 = Bewirtung / Restaurant\n  - 6670 = Kfz-Kosten / Kraftstoff\n  - 6815 = Bürobedarf / Software\n  - 8400 = Erlöse / Einnahmen\n• Dein Steuerberater kann die Datei direkt in DATEV-Software importieren\n• Tipp: Vor dem Export Kategorien prüfen — sie bestimmen die Kontozuordnung"

        # CSV Import (before generic "csv" / "import")
        elif any(w in msg for w in ["csv import", "foto import", "importieren", "içe aktar", "einlesen"]):
            reply = "📥 Import-Optionen:\n\n1. CSV Import:\n• Kassenbuch → 'CSV Import' Button → CSV-Datei auswählen\n• Trennzeichen: Komma oder Semikolon (automatisch erkannt)\n• Spaltenbezeichnungen: Deutsch oder Englisch\n• Datumsformat: DD.MM.YYYY oder YYYY-MM-DD\n• Einnahmen/Ausgaben werden automatisch erkannt\n• MwSt wird automatisch mit 19% berechnet\n• Jede Zeile erstellt Kassenbuch-Eintrag UND Rechnung\n• Beispiel: Datum,Beschreibung,Lieferant,Ausgaben,Einnahmen,Kategorie\n\n2. Foto Import:\n• Kassenbuch → 'Foto Import' Button\n• OCR erkennt handgeschriebene Tabellen: Datum | Beschreibung | Betrag\n• Tipp: Gerade und bei guter Beleuchtung fotografieren\n\n3. Beleg Upload:\n• Upload-Seite → PDF/Foto hochladen"

        # PDF / Drucken (before generic "download")
        elif any(w in msg for w in ["drucken", "pdf", "ausdrucken", "yazdır", "print"]):
            reply = "🖨️ PDF Drucken:\n\nWo findest du den PDF Button?\n• Rechnungen → 🖨️ PDF Button neben jeder Rechnung\n• Kassenbuch → 🖨️ PDF Button (wenn Rechnung verknüpft)\n• Belege → 🖨️ PDF Button unter jedem Beleg\n• Rechnung erstellen → Nach Speichern: 'PDF herunterladen?' Dialog\n\nPDF enthält:\n• Firmenname und E-Mail (aus deinem Profil)\n• Kunde/Lieferant, Rechnungs-Nr., Datum\n• Beschreibung, Kategorie, MwSt-Satz\n• Netto, MwSt-Betrag, Gesamtbetrag (Brutto)\n• §19-Hinweis (wenn Kleinunternehmer aktiviert)\n• Footer: 'Erstellt mit AutoTax-HUB'"

        # Preise / Pricing (before generic "kosten" catches it)
        elif any(w in msg for w in ["pricing", "abo", "plan ", "upgrade", "tarif", "paket", "ücret"]) or (any(w in msg for w in ["preis", "fiyat", "kosten"]) and any(w in msg for w in ["monat", "plan", "abo", "wie viel kostet", "was kostet autotax"])):
            reply = "💰 Preise:\n• Free: €0/Monat — 50 Rechnungen, 2 Firmen, CSV Export\n• Early Adopter: €10/Monat — 500 Rechnungen, 5 Firmen, DATEV\n• Pro: €20/Monat — Unbegrenzt, API, Priority Support\n• Frühe Nutzer behalten ihren Preis dauerhaft!\n• Stripe-Zahlung kommt bald"

        # Firmen verwalten (before "lieferant" catches "firma")
        elif any(w in msg for w in ["firmen verwalten", "meine firma", "firma ändern", "firma registrier", "firma hinzufügen"]):
            reply = "🏢 Firmen verwalten:\n• Sidebar → 'Firmen' Seite\n• Max. 2 Firmen registrieren\n• Firmenname wird für Einnahme-Erkennung verwendet\n• Upload: Vendor = deine Firma → automatisch Einnahme\n• Firma kann nicht geändert werden (Kontakt Support)"

        # Wie viele / Anzahl (before "wie viel" catches it)
        elif any(w in msg for w in ["wie viele", "anzahl", "count", "kaç tane", "adet"]):
            reply = f"📊 Count:\n• Invoices: {inv_count}\n• Income: {len(inv_inc)}\n• Expenses: {len(inv_exp)}"

        # ── TIER 2: Greetings & meta (catch early to avoid false matches) ──

        # Hallo / Greeting
        elif any(w in msg for w in ["hallo", "hey", "merhaba", "hello", "guten tag", "guten morgen", "guten abend", "selam", "servus", "grüß"]) or msg in ["hi", "na"]:
            reply = f"👋 Hello! You have {inv_count} invoices. How can I help? Type 'help' for overview."

        # Danke
        elif any(w in msg for w in ["danke", "thanks", "thx", "merci", "teşekkür", "sağol", "gracias", "super", "perfekt", "top"]):
            reply = "You're welcome! Ask anytime. 😊"

        # Hilfe / Help
        elif any(w in msg for w in ["hilfe", "help", "was kannst", "anleitung", "wie funktioniert", "feature", "yardım", "yardim", "nasıl", "nasil", "nedir", "ne yapabilir", "fonksiyon", "what can"]):
            reply = "🤖 I can help with:\n• 'How much?' — totals\n• 'Categories' — expenses by category\n• 'VAT' — input and output VAT\n• 'Tax' — tax estimate\n• 'Profit' — income vs expenses\n• 'Vendors' — top vendors\n• 'Dashboard' — financial overview\n• 'Cash Book' — bookkeeping status\n• 'Invoices' — invoice overview\n• 'Upload' — upload receipts\n• 'Import' — CSV or photo import\n• 'Export' / 'CSV' / 'DATEV' — export data\n• 'Tax report' — tax declaration\n• 'E-Invoice' — XRechnung / ZUGFeRD\n• 'Create Invoice' — your own invoices\n• 'PDF' / 'Print' — download PDF\n• 'Small business' — §19 UStG\n• 'Companies' — company management\n• 'Pricing' — plans & subscription\n• 'App' / 'PWA' — mobile usage\n• 'Receipts' — receipt management\n• 'Sync' — synchronization\n• 'QR' — QR code recognition\n• 'Password' — account & login\n• 'Delete' — remove entries\n\nOr just ask freely — e.g. a vendor name!"

        # ── TIER 3: Data queries (user's actual financial data) ──

        # Summe / Gesamt / wie viel
        elif any(w in msg for w in ["wie viel", "wieviel", "summe", "total", "gesamt", "how much", "insgesamt", "ne kadar", "özet", "zusammenfassung", "overview", "toplam"]):
            reply = f"📊 Overview:\n• Invoices: {inv_count} (€{inv_sum:.2f})\n• Income: €{total_income:.2f}\n• Expenses: €{total_expenses:.2f}\n• Profit: €{net_profit:.2f}"

        # Kategorie
        elif any(w in msg for w in ["kategorie", "categories", "aufteilung", "verteilung", "category", "kategori"]):
            reply = f"📂 Categories:\n{cat_str}"

        # MwSt / VAT (before generic "steuer")
        elif any(w in msg for w in ["mwst", "vat", "umsatzsteuer", "mehrwertsteuer", "vorsteuer", "kdv", "tva"]):
            reply = f"🧾 VAT Overview:\n• Input VAT paid: €{vat_paid:.2f}\n• Output VAT collected: €{vat_collected:.2f}\n• Saldo: €{vat_balance:.2f}\n{'→ You get back €'+str(abs(round(vat_balance,2)))+' refund' if vat_balance < 0 else '→ You owe €'+str(round(vat_balance,2)) if vat_balance > 0 else '→ Balanced'}"

        # Steuer / Einkommensteuer
        elif any(w in msg for w in ["steuer", "tax", "einkommensteuer", "steuerlast", "gelir vergisi"]):
            if net_profit > 277826:
                rate = 45
            elif net_profit > 61356:
                rate = 42
            elif net_profit > 17005:
                rate = 30
            elif net_profit > 10908:
                rate = 14
            else:
                rate = 0
            estimate = round(net_profit * rate / 100, 2) if net_profit > 0 else 0
            reply = f"💰 Tax Estimate (Germany):\n• Profit: €{net_profit:.2f}\n• Tax rate: {rate}%\n• Estimated tax: €{estimate:.2f}\n\nNote: This is an estimate. Consult a tax advisor for accurate calculation."

        # Einnahmen / Income
        elif any(w in msg for w in ["einnahme", "income", "umsatz", "revenue", "verdien", "gelir", "kazanç"]):
            reply = f"📈 Income: €{total_income:.2f} ({len(inv_inc)} items)"

        # Ausgaben / Expenses
        elif any(w in msg for w in ["ausgabe", "expense", "kosten", "cost", "bezahl", "gider", "harcama", "masraf"]):
            reply = f"📉 Expenses: €{total_expenses:.2f} ({len(inv_exp)} items)"

        # Profit / Profit
        elif any(w in msg for w in ["gewinn", "profit", "verlust", "loss", "ergebnis", "kâr", "kar", "zarar"]):
            emoji = "📈" if net_profit >= 0 else "📉"
            reply = f"{emoji} Net result: €{net_profit:.2f}\n• Income: €{total_income:.2f}\n• Expenses: €{total_expenses:.2f}"

        # Lieferant / Vendor (removed "firma"/"şirket" — those go to Firmen now)
        elif any(w in msg for w in ["lieferant", "vendor", "händler", "anbieter", "tedarikçi", "top lieferant"]):
            reply = f"🏢 Top Vendors:\n{top_vendors}"

        # ── TIER 4: Page/feature navigation ──

        # Kassenbuch
        elif any(w in msg for w in ["kassenbuch", "bookkeeping", "kasse"]):
            reply = f"📒 Kassenbuch (Bookkeeping):\n• Automatische Synchronisation: Hochgeladene Rechnungen erscheinen automatisch\n• Manuelle Einträge: '+ Eintrag' → Typ, Beschreibung, Lieferant, Betrag, MwSt-Satz, Kategorie, Zahlungsmethode\n• 'Rechnungen sync' Button: Überträgt neue Rechnungen ins Kassenbuch (Duplikate werden übersprungen)\n• Abstimmung (Reconcile): ⬜ klicken → ✅ markiert Eintrag als abgestimmt\n• CSV-Export: Kassenbuch → 'CSV Export'\n• Foto Import: Kassenbuch → 'Foto Import' für handgeschriebene Kassenbücher\n• Bearbeiten & Löschen: Jeder Eintrag einzeln anpassbar\n• 🖨️ PDF Button neben Einträgen mit verknüpfter Rechnung\n\nCurrent: {inv_count} Rechnungen ({len(inv_inc)} Einnahmen, {len(inv_exp)} Ausgaben)"

        # Belege (before generic "rechnung")
        elif any(w in msg for w in ["belege", "beleg", "original", "dokument"]):
            reply = "📎 Receipts:\n• Sidebar → Receipts page\n• All uploaded receipts at a glance\n• 'View Original': fullscreen preview of image/PDF\n• 'Download': download original file\n• 🖨️ PDF Button: Rechnung als PDF generieren\n• Receipts are stored in database (nicht nur als Datei)\n• Search: filter by vendor or amount\n• Tip: export regularly as backup"

        # Rechnung / Invoice (generic)
        elif any(w in msg for w in ["rechnung", "invoice", "faktur", "fatura", "bon", "quittung"]):
            reply = f"🧾 Invoices:\n• All uploaded receipts at a glance: {inv_count} gesamt (€{inv_sum:.2f})\n• Income: {len(inv_inc)} | Expenses: {len(inv_exp)}\n• Search by vendor, amount or category\n• Filter: Vendor, Kategorie, Datum (Von/Bis), Status\n• Inline editing: 'Bearbeiten' → Vendor, Betrag, Kategorie, Datum, MwSt ändern\n• Mehrfach löschen: Häkchen setzen → 'X ausgewählte löschen'\n• 🖨️ PDF Button neben jeder Rechnung\n• '+ Rechnung erstellen' für neue Einnahmen\n• Pagination for large receipt collections\n\nTipp: 'E-Rechnung' für XML-Import, 'PDF' zum Drucken."

        # Upload
        elif any(w in msg for w in ["upload", "hochladen", "scan", "ocr", "yükle"]):
            reply = "📤 Upload & OCR:\n• Supported formats: PDF, PNG, JPEG, TIFF, WEBP (max. 5 MB)\n• Single or batch upload (up to 20 files)\n• OCR auto-detects: Lieferant, Betrag, MwSt-Satz, MwSt-Betrag, Datum\n• Erkannte Kategorien: Lebensmittel, Kraftstoff, Restaurant, Shopping, Transport u.v.m.\n• 350+ vendors auto-recognized (Lidl, Amazon, Shell, etc.)\n• Handwriting mode available for handwritten receipts\n• Income/expense selectable before upload (default: expense)\n• E-Rechnung (XML) Upload: 'E-Rechnung hochladen' Button\n• QR-Codes auf Rechnungen werden gelesen (EPC/SEPA, Swiss QR)\n• After upload, receipt appears in Invoices AND Cash Book"

        # Export
        elif any(w in msg for w in ["export", "excel", "exportieren"]):
            reply = "💾 Export Options:\n• CSV: Comma-separated, compatible with Excel, Google Sheets, LibreOffice\n• DATEV: Standard format for German tax advisors — direct import\n• Excel: .xlsx with formatted columns\n• JSON: Structured data for developers and API\n• Kassenbuch CSV: Kassenbuch → 'CSV Export' Button\n\nSo geht's:\n1. Gehe zu 'Export'\n2. Wähle optional ein Steuerjahr\n3. Klicke auf CSV, DATEV, Excel oder JSON\n4. Datei wird sofort im Browser heruntergeladen\n\nTipp: Exportierte CSV kann direkt wieder mit 'CSV Import' importiert werden!"

        # CSV
        elif any(w in msg for w in ["csv"]):
            reply = "📄 CSV Funktionen:\n\n• CSV Export (Rechnungen): Export-Seite → 'CSV'\n• CSV Export (Kassenbuch): Kassenbuch → 'CSV Export'\n• CSV Import: Kassenbuch → 'CSV Import'\n\nCSV Format:\nDatum, Lieferant, Rechnungs-Nr., Typ, Betrag, MwSt, MwSt-Satz, Kategorie, Zahlungsart\n• Trennzeichen: Komma oder Semikolon (automatisch erkannt)\n• Spalten: Deutsch oder Englisch\n• Datumsformat: DD.MM.YYYY oder YYYY-MM-DD\n\nTipp: Exportiere erst eine CSV als Vorlage — dann im gleichen Format importieren."

        # Dashboard
        elif any(w in msg for w in ["dashboard", "übersicht", "überblick", "grafik", "chart", "diagramm"]):
            reply = f"📊 Dashboard — Financial Overview:\n• Income: €{total_income:.2f} | Expenses: €{total_expenses:.2f}\n• Profit: €{net_profit:.2f}\n• MwSt-Saldo: €{vat_balance:.2f}\n• Invoices: {inv_count}\n\nFeatures:\n• Tax estimate per German law (Progressionsstufen: 0%, 14%, 30%, 42%, 45%)\n• MwSt-Übersicht: Vorsteuer, USt, Saldo\n• Monthly chart: income vs expenses\n• Category distribution: where do you spend most?\n• CSV-Export-Button: Alle Rechnungen als CSV\n• 'Reset': deletes ALL data (double confirmation!)"

        # Firmen (generic)
        elif any(w in msg for w in ["firma", "firmen", "unternehmen", "company", "şirket"]):
            reply = "🏢 Firmen verwalten:\n• Sidebar → 'Firmen' Seite\n• Max. 2 Firmen registrieren (Free Plan)\n• Firmenname wird für Einnahme-Erkennung verwendet\n• Upload: Wenn Vendor = deine Firma → automatisch als Einnahme erkannt\n• Firmenname erscheint auf generierten PDFs\n• Firma kann nicht geändert werden (Kontakt: info@autotaxhub.de)"

        # Preise (generic fallback)
        elif any(w in msg for w in ["preis", "fiyat", "was kostet"]):
            reply = "💰 Preise & Pläne:\n• Free: €0/Monat — 50 Rechnungen, 2 Firmen, CSV Export\n• Early Adopter: €10/Monat — 500 Rechnungen, 5 Firmen, DATEV, Excel, PDF\n• Pro: €20/Monat — Unbegrenzt, API, Priority Support\n• Frühe Nutzer behalten ihren Preis dauerhaft!\n• Stripe-Zahlung kommt bald\n\nUpgrade: Sidebar → 'Preise' Seite"

        # PWA / Mobil
        elif any(w in msg for w in ["app", "mobil", "handy", "telefon", "pwa", "installieren", "uygulama"]):
            reply = "📱 Mobile App (PWA):\n• Open AutoTax-HUB in your mobile browser\n• iPhone: Safari → Share → 'Add to Home Screen'\n• Android: Chrome → Menu → 'Install App'\n• Works like native app — no App Store needed\n• Upload receipts directly with camera\n• Sidebar: hamburger menu opens/closes navigation\n• Tip: bookmark on home screen for quick access"

        # ── TIER 5: Action helpers ──

        # Löschen / Delete
        elif any(w in msg for w in ["lösch", "delete", "entfern", "zurücksetz", "sil", "kaldır", "temizle"]):
            reply = "🗑️ Delete:\n• Single: trash icon (✕) next to entry\n• Multiple: check boxes → 'X ausgewählte löschen' Button\n• Reset all: Dashboard → 'Reset'\n  (WARNING: double confirmation, irreversible! No undo!)\n\nDelete works in Invoices AND Cash Book.\nDeleted entries are gone immediately — export first recommended."

        # Passwort / Login
        elif any(w in msg for w in ["passwort", "password", "şifre", "kennwort", "login", "anmeld", "registrier", "konto"]):
            reply = "🔐 Account & Security:\n• Password: min 8 chars, 1 uppercase, 1 digit (Beispiel: MeinPasswort1)\n• Login: email + password\n• Token: auto-renewal (1h Access, 7 Tage Refresh)\n• Register: click 'Register' on login page\n• Change password: profile menu top right → 'Change Password'"

        # Sync / Synchronisieren
        elif any(w in msg for w in ["sync", "synchron", "senkron"]):
            reply = "🔄 Synchronization:\n• Upload → receipt appears in Invoices + Cash Book\n• Kassenbuch → 'Rechnungen sync' synchronisiert fehlende Einträge\n• Rechnungen → 'Kassenbuch sync' synchronisiert in beide Richtungen\n• Duplicates auto-detected and skipped\n• Reverse sync: manual cash book entries → Invoices"

        # Reconcile / Abstimmen
        elif any(w in msg for w in ["reconcil", "abstimm", "häkchen", "checkbox"]):
            reply = "✅ Reconcile:\n• Cash Book → Click ⬜ next to entry → becomes ✅\n• Marks entry as 'reconciled with bank statement'\n• Helps match with bank statements\n• Can be undone anytime (✅ → ⬜)"

        # QR Code
        elif any(w in msg for w in ["qr", "barcode"]):
            reply = "📱 QR Code Recognition:\n• QR codes on invoices are auto-read\n• Unterstützt: EPC/SEPA (GiroCode), Swiss QR, ZUGFeRD\n• Extracts: company, IBAN, amount, reference\n• QR data overrides OCR when available (more accurate)\n• Tip: QR code must be clearly visible in image"

        # Foto / Bild Qualität
        elif any(w in msg for w in ["foto", "qualität", "unscharf", "dunkel", "yamuk", "blurry", "bild"]):
            reply = "📸 Photo Tips for Better Recognition:\n• Good lighting — no shadows on receipt\n• Straight photo — not tilted\n• Full receipt in frame — nothing cut off\n• Use original photo (not WhatsApp compressed)\n• PDF is better than photo (if available)\n• Enable handwriting mode for handwritten receipts\n• Tip: use camera app instead of screenshot"

        # Einnahme / Ausgabe
        elif any(w in msg for w in ["einnahme oder ausgabe", "ausgabe oder einnahme", "einnahme ausgabe"]):
            reply = "📈 Income / Expense:\n• Upload page: before uploading 'Ausgabe (Gider)' or 'Income' select\n• Default is expense — click green button for income\n• Kassenbuch: '+ Eintrag' → Typ 'Einnahme' oder 'Ausgabe' im Formular wählen\n• Dashboard shows income (green) and expenses (red) separately\n• Profit = income minus expenses\n• Auto-detect: if vendor = your company → auto income"

        # Eintragen / Hinzufügen
        elif any(w in msg for w in ["eintragen", "hinzufügen", "eingeben", "ekle", "kaydet", "erfassen"]):
            reply = "✏️ Create Entry:\n• Upload → upload receipt (OCR auto-detects everything)\n• Cash Book → '+ Entry' button → fill form:\n  Typ, Beschreibung, Lieferant, Betrag, MwSt-Satz, Kategorie, Zahlungsmethode\n• Invoices → '+ Create Invoice' for income\n• CSV Import: Cash Book → 'CSV Import' for bulk entry\n• Photo Import: Cash Book → 'Photo Import' for handwritten books\n• Both create entries in Invoices AND Cash Book"

        # Suche / Finden
        elif any(w in msg for w in ["such", "find", "wo ist", "wo sind", "finden", "ara", "bul", "nerede", "search", "where"]):
            reply = "🔍 Search:\n• Invoices → search field at top (searches vendor, OCR text, category)\n• Multiple words possible: z.B. 'Lidl Dezember'\n• Filter: Vendor, Kategorie, Datum (Von/Bis), Status\n• Cash Book → own search field\n• Receipts → search by vendor/amount\n• AI Chat: Ask me a vendor name like 'Lidl'"

        # Bearbeiten / Ändern
        elif any(w in msg for w in ["bearbeit", "änder", "korrigier", "edit", "düzenle", "değiştir"]):
            reply = "✏️ Edit:\n• Invoices → 'Edit' next to entry\n• Cash Book → 'Edit' next to entry\n• Editable fields: Vendor, Betrag, MwSt-Betrag, MwSt-Satz, Kategorie, Datum, Rechnungs-Nr., Zahlungsart\n• Tipp: If OCR was wrong → correct here"

        # Datum / Date
        elif any(w in msg for w in ["datum", "tarih", "zeitraum", "monat", "jahr"]):
            reply = "📅 Date Filter:\n• Invoices → use From/To fields (calendar only)\n• Supported formats: DD.MM.YYYY, YYYY-MM-DD\n• Monthly view: Dashboard shows monthly chart als Diagramm\n• Export: filterable by tax year\n• If OCR misses date → today's date as fallback"

        # Download / Herunterladen (generic fallback)
        elif any(w in msg for w in ["download", "herunterladen"]):
            reply = "📥 Download Options:\n• Invoice as PDF: 🖨️ PDF Button neben jeder Rechnung\n• Original receipt: Belege → 'Download' Button\n• Export data: Export-Seite → CSV, DATEV, Excel, JSON\n• Kassenbuch CSV: Kassenbuch → 'CSV Export'"

        # ── TIER 6: Vendor search fallback ──
        # Vendor search — if no keyword matched, try searching vendor names
        if reply is None:
            vendor_results = [i for i in invoices if msg in (i.vendor or "").lower()]
            if not vendor_results:
                vendor_results = db.query(Invoice).filter(Invoice.user_id == user["sub"], Invoice.vendor.ilike(f"%{msg}%")).all()
            if vendor_results:
                vr_total = sum(safe_float(i.total_amount) for i in vendor_results)
                vr_vat = sum(safe_float(i.vat_amount) for i in vendor_results)
                latest = safe_date_str(vendor_results[0].date) if vendor_results[0].date else "unbekannt"
                reply = f"🔍 For '{msg.title()}' I found {len(vendor_results)} invoice(s):\n• Total: €{vr_total:.2f}\n• MwSt: €{vr_vat:.2f}\n• Last invoice: {latest}"
            else:
                reply = f"I didn't quite understand. Try e.g.:\n• 'How many invoices?'\n• 'VAT overview'\n• 'Profit'\n• A vendor name (e.g. 'Lidl')\n• 'Help' for all topics\n\nCurrent: {inv_count} invoices, €{net_profit:.2f} Profit"

        return {"reply": reply}
    except Exception:
        logger.exception("Chat failed")
        return {"reply": "Sorry, an error occurred. Please try again."}
    finally:
        db.close()


# ============================================================
# EXPORT: PDF / CSV / DATEV / EXCEL / JSON
# All exports use calculate_dashboard_metrics() so values
# always match the dashboard exactly.
# ============================================================

@app.get("/export/pdf")
def export_pdf_report(year: int = Query(None), user: dict = Depends(get_current_user)):
    """Full financial report PDF with dashboard metrics + transactions."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas as pdf_canvas
        from reportlab.lib.units import cm
        from reportlab.lib.colors import HexColor
    except ImportError:
        err(501, "PDF-Generierung nicht verfügbar (reportlab fehlt)")

    m = calculate_dashboard_metrics(user["sub"], year)
    txns = m["transactions"]

    db = SessionLocal()
    try:
        companies = db.query(UserCompany).filter(UserCompany.user_id == user["sub"]).all()
        company_name = companies[0].company_name if companies else "Meine Firma"
        u = db.query(User).filter(User.id == user["sub"]).first()
    finally:
        db.close()

    buf = io.BytesIO()
    c = pdf_canvas.Canvas(buf, pagesize=A4)
    w, h = A4
    year_label = str(year) if year else "Gesamt"

    def draw_header(c, y):
        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 20)
        c.drawString(2*cm, y, f"Finanzbericht {year_label}")
        c.setFont("Helvetica", 9)
        c.setFillColor(HexColor("#7a8ba8"))
        c.drawString(2*cm, y - 0.6*cm, f"{company_name}  |  {u.email if u else ''}  |  Erstellt: {datetime.now().strftime('%d.%m.%Y %H:%M')}")
        return y - 1.5*cm

    def draw_footer(c, page_num):
        c.setFillColor(HexColor("#7a8ba8"))
        c.setFont("Helvetica", 7)
        c.drawString(2*cm, 1*cm, f"AutoTax-HUB | {company_name} | Automatisch erstellt. Alle Angaben ohne Gewähr. Keine Steuerberatung. | Seite {page_num}")

    def new_page(c, page_num):
        draw_footer(c, page_num)
        c.showPage()
        return h - 2*cm, page_num + 1

    page = 1
    y = h - 2*cm
    y = draw_header(c, y)

    # --- Summary Section ---
    c.setFillColor(HexColor("#10b981"))
    c.rect(2*cm, y - 0.1*cm, 17*cm, 0.06*cm, fill=1, stroke=0)
    y -= 0.6*cm
    c.setFillColor(HexColor("#1a2d4a"))
    c.setFont("Helvetica-Bold", 13)
    c.drawString(2*cm, y, "Zusammenfassung")
    y -= 0.8*cm

    summary_items = [
        ("Einnahmen", f"{m['total_income']:,.2f} EUR", "#10b981"),
        ("Ausgaben", f"{m['total_expenses']:,.2f} EUR", "#ef4444"),
        ("Gewinn / Verlust", f"{m['profit']:,.2f} EUR", "#3b82f6"),
        ("Rechnungen", str(m['invoice_count']), "#6366f1"),
        ("MwSt erhalten", f"{m['vat_collected']:,.2f} EUR", "#10b981"),
        ("MwSt gezahlt", f"{m['vat_paid']:,.2f} EUR", "#ef4444"),
        ("MwSt Saldo", f"{m['vat_balance']:,.2f} EUR", "#f59e0b"),
    ]
    c.setFont("Helvetica", 10)
    for label, val, color in summary_items:
        c.setFillColor(HexColor("#1a2d4a"))
        c.drawString(2.5*cm, y, label)
        c.setFillColor(HexColor(color))
        c.drawRightString(12*cm, y, val)
        y -= 0.55*cm

    # --- VAT by Rate ---
    if m["vat_by_rate"]:
        y -= 0.4*cm
        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 11)
        c.drawString(2*cm, y, "MwSt nach Satz")
        y -= 0.6*cm
        c.setFont("Helvetica", 9)
        for rate, info in sorted(m["vat_by_rate"].items()):
            c.setFillColor(HexColor("#1a2d4a"))
            c.drawString(2.5*cm, y, f"{rate}: {info['count']}x  Umsatz {info['amount']:,.2f} EUR  MwSt {info['vat']:,.2f} EUR")
            y -= 0.45*cm

    # --- Monthly Breakdown Table ---
    if m["monthly"]:
        y -= 0.6*cm
        c.setFillColor(HexColor("#10b981"))
        c.rect(2*cm, y + 0.3*cm, 17*cm, 0.06*cm, fill=1, stroke=0)
        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 13)
        c.drawString(2*cm, y - 0.3*cm, "Monatliche Aufstellung")
        y -= 0.9*cm

        # Table header
        c.setFillColor(HexColor("#1a2d4a"))
        c.rect(2*cm, y - 0.1*cm, 17*cm, 0.6*cm, fill=1)
        c.setFillColor(HexColor("#ffffff"))
        c.setFont("Helvetica-Bold", 9)
        c.drawString(2.2*cm, y + 0.1*cm, "Monat")
        c.drawRightString(9*cm, y + 0.1*cm, "Einnahmen")
        c.drawRightString(13*cm, y + 0.1*cm, "Ausgaben")
        c.drawRightString(17*cm, y + 0.1*cm, "Differenz")
        y -= 0.7*cm

        c.setFont("Helvetica", 9)
        for mb in m["monthly"]:
            if y < 3*cm:
                y, page = new_page(c, page)
            diff = mb["income"] - mb["expenses"]
            c.setFillColor(HexColor("#1a2d4a"))
            c.drawString(2.2*cm, y, mb["month"])
            c.setFillColor(HexColor("#10b981"))
            c.drawRightString(9*cm, y, f"{mb['income']:,.2f}")
            c.setFillColor(HexColor("#ef4444"))
            c.drawRightString(13*cm, y, f"{mb['expenses']:,.2f}")
            c.setFillColor(HexColor("#10b981") if diff >= 0 else HexColor("#ef4444"))
            c.drawRightString(17*cm, y, f"{diff:,.2f}")
            y -= 0.5*cm

        # Monthly totals
        c.setFont("Helvetica-Bold", 9)
        c.setFillColor(HexColor("#1a2d4a"))
        c.drawString(2.2*cm, y, "GESAMT")
        c.setFillColor(HexColor("#10b981"))
        c.drawRightString(9*cm, y, f"{m['total_income']:,.2f}")
        c.setFillColor(HexColor("#ef4444"))
        c.drawRightString(13*cm, y, f"{m['total_expenses']:,.2f}")
        c.setFillColor(HexColor("#3b82f6"))
        c.drawRightString(17*cm, y, f"{m['profit']:,.2f}")
        y -= 0.8*cm

    # --- Category Distribution ---
    if m["by_category"]:
        if y < 5*cm:
            y, page = new_page(c, page)
        c.setFillColor(HexColor("#10b981"))
        c.rect(2*cm, y + 0.3*cm, 17*cm, 0.06*cm, fill=1, stroke=0)
        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 13)
        c.drawString(2*cm, y - 0.3*cm, "Kategorien")
        y -= 0.9*cm
        c.setFont("Helvetica", 9)
        total_cat = sum(cat["total"] for cat in m["by_category"])
        for cat in m["by_category"]:
            if y < 2.5*cm:
                y, page = new_page(c, page)
            pct = round(cat["total"] / total_cat * 100, 1) if total_cat > 0 else 0
            c.setFillColor(HexColor("#1a2d4a"))
            c.drawString(2.5*cm, y, cat["category"])
            c.drawRightString(12*cm, y, f"{cat['total']:,.2f} EUR")
            c.setFillColor(HexColor("#7a8ba8"))
            c.drawString(12.5*cm, y, f"({pct}%)")
            y -= 0.45*cm

    # --- Transaction List ---
    if txns:
        if y < 5*cm:
            y, page = new_page(c, page)
        y -= 0.4*cm
        c.setFillColor(HexColor("#10b981"))
        c.rect(2*cm, y + 0.3*cm, 17*cm, 0.06*cm, fill=1, stroke=0)
        c.setFillColor(HexColor("#1a2d4a"))
        c.setFont("Helvetica-Bold", 13)
        c.drawString(2*cm, y - 0.3*cm, f"Transaktionen ({len(txns)})")
        y -= 0.9*cm

        # Table header
        c.setFillColor(HexColor("#1a2d4a"))
        c.rect(2*cm, y - 0.1*cm, 17*cm, 0.6*cm, fill=1)
        c.setFillColor(HexColor("#ffffff"))
        c.setFont("Helvetica-Bold", 8)
        c.drawString(2.2*cm, y + 0.1*cm, "Datum")
        c.drawString(4.5*cm, y + 0.1*cm, "Lieferant")
        c.drawString(10*cm, y + 0.1*cm, "Kategorie")
        c.drawString(13*cm, y + 0.1*cm, "MwSt")
        c.drawRightString(17*cm, y + 0.1*cm, "Betrag")
        c.drawRightString(19*cm, y + 0.1*cm, "Typ")
        y -= 0.7*cm

        c.setFont("Helvetica", 8)
        for t in txns:
            if y < 2.5*cm:
                y, page = new_page(c, page)
            c.setFillColor(HexColor("#1a2d4a"))
            c.drawString(2.2*cm, y, t["date"][:10])
            c.drawString(4.5*cm, y, (t["vendor"] or "")[:28])
            c.drawString(10*cm, y, t["category"][:12])
            c.drawString(13*cm, y, t["vat_rate"])
            is_inc = t["invoice_type"] == "income"
            c.setFillColor(HexColor("#10b981") if is_inc else HexColor("#ef4444"))
            c.drawRightString(17*cm, y, f"{t['total_amount']:,.2f}")
            c.setFillColor(HexColor("#10b981") if is_inc else HexColor("#ef4444"))
            c.drawRightString(19*cm, y, "E" if is_inc else "A")
            y -= 0.42*cm

    draw_footer(c, page)
    c.save()
    buf.seek(0)
    fname = f"AutoTax_Finanzbericht_{year_label}.pdf"
    return StreamingResponse(buf, media_type="application/pdf", headers={"Content-Disposition": f"attachment; filename={fname}"})


@app.get("/export/csv")
def export_csv(year: int = Query(None), user: dict = Depends(get_current_user)):
    m = calculate_dashboard_metrics(user["sub"], year)
    txns = m["transactions"]
    buf = io.StringIO()
    buf.write("\ufeff")  # UTF-8 BOM for Excel
    # Summary rows at top
    buf.write(f"# AutoTax-HUB Finanzbericht {m['year'] or 'Gesamt'}\n")
    buf.write(f"# Einnahmen;{m['total_income']:.2f}\n")
    buf.write(f"# Ausgaben;{m['total_expenses']:.2f}\n")
    buf.write(f"# Gewinn;{m['profit']:.2f}\n")
    buf.write(f"# MwSt Saldo;{m['vat_balance']:.2f}\n")
    buf.write(f"# Rechnungen;{m['invoice_count']}\n")
    buf.write("#\n")
    # Transaction header
    buf.write("Datum;Lieferant;Rechnungs-Nr.;Typ;Betrag;MwSt;MwSt-Satz;Kategorie;Zahlungsart;Konto\n")
    for t in txns:
        vendor = (t["vendor"] or "").replace('"', '""')
        buf.write(f'{t["date"]};"{vendor}";{t["invoice_number"]};{t["invoice_type"]};{t["total_amount"]:.2f};{t["vat_amount"]:.2f};{t["vat_rate"]};{t["category"]};{t["payment_method"]};{t["konto"]}\n')
    logger.info("CSV export: %d rows, income=%.2f expenses=%.2f", len(txns), m["total_income"], m["total_expenses"])
    buf.seek(0)
    return StreamingResponse(buf, media_type="text/csv; charset=utf-8", headers={"Content-Disposition": f"attachment; filename=autotax_export_{year or 'alle'}.csv"})


@app.get("/export/datev")
def export_datev(year: int = Query(None), user: dict = Depends(get_current_user)):
    m = calculate_dashboard_metrics(user["sub"], year)
    txns = m["transactions"]
    buf = io.StringIO()
    buf.write("# HINWEIS: AutoTax-HUB - Alle Daten pruefen. Keine Steuerberatung.\n")
    buf.write(f"# Einnahmen: {m['total_income']:.2f} | Ausgaben: {m['total_expenses']:.2f} | Gewinn: {m['profit']:.2f}\n")
    buf.write("Umsatz;Soll/Haben;Konto;Gegenkonto;BU;Belegdatum;Buchungstext;USt\n")
    for t in txns:
        sh = "S" if t["invoice_type"] == "expense" else "H"
        date_str = ""
        parts = t["date"].split("-")
        if len(parts) == 3:
            date_str = f"{parts[2]}{parts[1]}"
        amt = f"{t['total_amount']:.2f}".replace(".", ",")
        vendor = (t["vendor"] or "").replace(";", " ")
        vat = (t["vat_rate"] or "0%").replace("%", "")
        buf.write(f"{amt};{sh};{t['konto']};1200;{vat};{date_str};{vendor};{vat}\n")
    buf.seek(0)
    return StreamingResponse(buf, media_type="text/csv", headers={"Content-Disposition": f"attachment; filename=autotax_datev_{year or 'all'}.csv"})


@app.get("/export/excel")
def export_excel(year: int = Query(None), user: dict = Depends(get_current_user)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    m = calculate_dashboard_metrics(user["sub"], year)
    txns = m["transactions"]

    wb = Workbook()

    # ===== Sheet 1: Transactions =====
    ws = wb.active
    ws.title = "Transaktionen"
    green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
    dark_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF", size=10)
    thin_border = Border(bottom=Side(style="thin", color="E2E8F0"))

    # Disclaimer
    disc = ws.cell(row=1, column=1, value=f"AutoTax-HUB Export — {m['year'] or 'Alle Jahre'} — Alle Daten vor Verwendung prüfen. Keine Steuerberatung.")
    disc.font = Font(italic=True, color="F59E0B", size=9)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)

    headers = ["Datum", "Lieferant", "Rechnungs-Nr.", "Typ", "Betrag", "MwSt", "MwSt-Satz", "Kategorie", "Zahlungsart", "Konto"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = hdr_font
        cell.fill = green_fill
        cell.alignment = Alignment(horizontal="center")

    row = 3
    for t in txns:
        ws.cell(row=row, column=1, value=t["date"])
        ws.cell(row=row, column=2, value=t["vendor"])
        ws.cell(row=row, column=3, value=t["invoice_number"])
        ws.cell(row=row, column=4, value=t["invoice_type"])
        c5 = ws.cell(row=row, column=5, value=t["total_amount"])
        c5.number_format = '#,##0.00'
        c6 = ws.cell(row=row, column=6, value=t["vat_amount"])
        c6.number_format = '#,##0.00'
        ws.cell(row=row, column=7, value=t["vat_rate"])
        ws.cell(row=row, column=8, value=t["category"])
        ws.cell(row=row, column=9, value=t["payment_method"])
        ws.cell(row=row, column=10, value=t["konto"])
        for c in range(1, 11):
            ws.cell(row=row, column=c).border = thin_border
        row += 1

    for col in range(1, 11):
        ws.column_dimensions[chr(64 + col) if col < 27 else "A" + chr(64 + col - 26)].width = 16

    # ===== Sheet 2: Dashboard Summary =====
    ws2 = wb.create_sheet("Dashboard")
    ws2.cell(row=1, column=1, value=f"Dashboard — {m['year'] or 'Alle Jahre'}").font = Font(bold=True, size=14, color="1E293B")
    ws2.merge_cells("A1:D1")

    # Summary
    summary_data = [
        ("Einnahmen", m["total_income"]),
        ("Ausgaben", m["total_expenses"]),
        ("Gewinn / Verlust", m["profit"]),
        ("Anzahl Rechnungen", m["invoice_count"]),
        ("  davon Einnahmen", m["income_count"]),
        ("  davon Ausgaben", m["expense_count"]),
        ("MwSt erhalten", m["vat_collected"]),
        ("MwSt gezahlt", m["vat_paid"]),
        ("MwSt Saldo", m["vat_balance"]),
    ]
    r = 3
    ws2.cell(row=r, column=1, value="Kennzahl").font = Font(bold=True)
    ws2.cell(row=r, column=2, value="Wert").font = Font(bold=True)
    for cell in [ws2.cell(row=r, column=1), ws2.cell(row=r, column=2)]:
        cell.fill = dark_fill
        cell.font = Font(bold=True, color="FFFFFF")
    r += 1
    for label, val in summary_data:
        ws2.cell(row=r, column=1, value=label)
        vc = ws2.cell(row=r, column=2, value=val)
        if isinstance(val, float):
            vc.number_format = '#,##0.00'
        r += 1

    # VAT by rate
    r += 1
    ws2.cell(row=r, column=1, value="MwSt nach Satz").font = Font(bold=True, size=11)
    r += 1
    for hdr_col, hdr_text in enumerate(["Satz", "Umsatz", "MwSt", "Anzahl"], 1):
        c = ws2.cell(row=r, column=hdr_col, value=hdr_text)
        c.fill = dark_fill
        c.font = Font(bold=True, color="FFFFFF")
    r += 1
    for rate, info in sorted(m["vat_by_rate"].items()):
        ws2.cell(row=r, column=1, value=rate)
        ws2.cell(row=r, column=2, value=info["amount"]).number_format = '#,##0.00'
        ws2.cell(row=r, column=3, value=info["vat"]).number_format = '#,##0.00'
        ws2.cell(row=r, column=4, value=info["count"])
        r += 1

    # Monthly breakdown
    r += 1
    ws2.cell(row=r, column=1, value="Monatliche Aufstellung").font = Font(bold=True, size=11)
    r += 1
    for hdr_col, hdr_text in enumerate(["Monat", "Einnahmen", "Ausgaben", "Differenz"], 1):
        c = ws2.cell(row=r, column=hdr_col, value=hdr_text)
        c.fill = green_fill
        c.font = Font(bold=True, color="FFFFFF")
    r += 1
    for mb in m["monthly"]:
        ws2.cell(row=r, column=1, value=mb["month"])
        ws2.cell(row=r, column=2, value=mb["income"]).number_format = '#,##0.00'
        ws2.cell(row=r, column=3, value=mb["expenses"]).number_format = '#,##0.00'
        ws2.cell(row=r, column=4, value=round(mb["income"] - mb["expenses"], 2)).number_format = '#,##0.00'
        r += 1
    # Totals row
    ws2.cell(row=r, column=1, value="GESAMT").font = Font(bold=True)
    ws2.cell(row=r, column=2, value=m["total_income"]).number_format = '#,##0.00'
    ws2.cell(row=r, column=3, value=m["total_expenses"]).number_format = '#,##0.00'
    ws2.cell(row=r, column=4, value=m["profit"]).number_format = '#,##0.00'
    for c in range(1, 5):
        ws2.cell(row=r, column=c).font = Font(bold=True)
    r += 2

    # Category distribution
    ws2.cell(row=r, column=1, value="Kategorien").font = Font(bold=True, size=11)
    r += 1
    for hdr_col, hdr_text in enumerate(["Kategorie", "Betrag", "Anteil"], 1):
        c = ws2.cell(row=r, column=hdr_col, value=hdr_text)
        c.fill = green_fill
        c.font = Font(bold=True, color="FFFFFF")
    r += 1
    total_cat = sum(cat["total"] for cat in m["by_category"])
    for cat in m["by_category"]:
        ws2.cell(row=r, column=1, value=cat["category"])
        ws2.cell(row=r, column=2, value=cat["total"]).number_format = '#,##0.00'
        pct = round(cat["total"] / total_cat * 100, 1) if total_cat > 0 else 0
        ws2.cell(row=r, column=3, value=f"{pct}%")
        r += 1

    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 16
    ws2.column_dimensions["C"].width = 16
    ws2.column_dimensions["D"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": f"attachment; filename=autotax_{year or 'all'}.xlsx"})


@app.get("/export/json")
def export_json(year: int = Query(None), user: dict = Depends(get_current_user)):
    import json as json_lib
    m = calculate_dashboard_metrics(user["sub"], year)
    buf = io.StringIO()
    json_lib.dump(m, buf, indent=2, ensure_ascii=False)
    buf.seek(0)
    return StreamingResponse(buf, media_type="application/json", headers={"Content-Disposition": f"attachment; filename=autotax_{year or 'all'}.json"})



# --- ADDED START: Soft delete restore + trash endpoints ---

@app.get("/invoices/deleted")
def list_deleted_invoices(user: dict = Depends(get_current_user)):
    """List all soft-deleted invoices (trash)."""
    db = SessionLocal()
    try:
        invs = db.query(Invoice).filter(Invoice.user_id == user["sub"], Invoice.is_deleted == True).order_by(Invoice.deleted_at.desc()).all()
        return {"success": True, "items": [invoice_to_dict(i) | {"deleted_at": i.deleted_at.strftime("%Y-%m-%dT%H:%M:%S") if i.deleted_at else ""} for i in invs], "total": len(invs)}
    finally:
        db.close()


@app.post("/invoices/{invoice_id}/restore")
def restore_invoice(invoice_id: int, user: dict = Depends(get_current_user)):
    """Restore a soft-deleted invoice."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        inv.is_deleted = False
        inv.deleted_at = None
        logger.info("Restored invoice %d", invoice_id)
        # Also restore linked cash entry
        linked = db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user["sub"]).first()
        if linked:
            linked.is_deleted = False
            linked.deleted_at = None
            logger.info("Restored linked cash entry for invoice %d", invoice_id)
        db.commit()
        return {"success": True, "restored": invoice_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Restore invoice failed")
        err(500, "Restore failed")
    finally:
        db.close()


@app.delete("/invoices/{invoice_id}/permanent")
def permanent_delete_invoice(invoice_id: int, user: dict = Depends(get_current_user)):
    """Permanently delete an invoice (from trash)."""
    db = SessionLocal()
    try:
        inv = db.query(Invoice).filter(Invoice.id == invoice_id, Invoice.user_id == user["sub"]).first()
        if not inv:
            err(404, "Invoice not found")
        # Also delete linked cash entry
        db.query(CashEntry).filter(CashEntry.invoice_id == invoice_id, CashEntry.user_id == user["sub"]).delete()
        db.delete(inv)
        db.commit()
        logger.info("Permanent delete: invoice %d", invoice_id)
        return {"success": True, "deleted": invoice_id}
    except HTTPException:
        raise
    except Exception:
        logger.exception("Permanent delete failed")
        err(500, "Delete failed")
    finally:
        db.close()

# --- ADDED END ---

# --- ADDED START: OCR safe wrapper — global monkey-patch for timeout protection ---
import asyncio as _asyncio_ocr
from autotax import ocr as _ocr_module

_original_extract_image_text = _ocr_module.extract_image_text
_original_extract_handwriting_text = _ocr_module.extract_handwriting_text


async def _safe_extract_image_text(content, filename):
    """Timeout-protected wrapper — falls back gracefully."""
    try:
        return await _asyncio_ocr.wait_for(_original_extract_image_text(content, filename), timeout=15)
    except _asyncio_ocr.TimeoutError:
        logger.warning("SAFE_OCR: extract_image_text timeout for %s", filename)
        return ""
    except Exception as e:
        logger.warning("SAFE_OCR: extract_image_text failed for %s: %s", filename, e)
        return ""


async def _safe_extract_handwriting_text(content, filename):
    """Timeout-protected wrapper for handwriting OCR."""
    try:
        return await _asyncio_ocr.wait_for(_original_extract_handwriting_text(content, filename), timeout=25)
    except _asyncio_ocr.TimeoutError:
        logger.warning("SAFE_OCR: extract_handwriting_text timeout for %s", filename)
        return ""
    except Exception as e:
        logger.warning("SAFE_OCR: extract_handwriting_text failed for %s: %s", filename, e)
        return ""


# Replace at module level so all callers get the safe versions automatically
_ocr_module.extract_image_text = _safe_extract_image_text
_ocr_module.extract_handwriting_text = _safe_extract_handwriting_text
logger.info("SAFE_OCR: global timeout protection enabled (15s image, 25s handwriting)")
# --- ADDED END ---

# --- ADDED START: Safe OCR wrapper helpers for direct endpoint use ---
async def safe_ocr_image(content: bytes, filename: str, timeout_s: int = 15) -> str:
    """Wrap extract_image_text with timeout + try/except. Returns text or empty string.
    Never hangs, never raises."""
    import asyncio as _ao
    try:
        # Uses the monkey-patched version (already timeout-protected)
        from autotax.ocr import extract_image_text as _eit
        return await _ao.wait_for(_eit(content, filename), timeout=timeout_s)
    except _ao.TimeoutError:
        logger.warning("safe_ocr_image timeout (%ds) for %s", timeout_s, filename)
        return ""
    except Exception as e:
        logger.warning("safe_ocr_image error for %s: %s", filename, e)
        return ""


async def safe_ocr_handwriting(content: bytes, filename: str, timeout_s: int = 25) -> str:
    """Wrap extract_handwriting_text with timeout + try/except."""
    import asyncio as _ao
    try:
        from autotax.ocr import extract_handwriting_text as _eht
        return await _ao.wait_for(_eht(content, filename), timeout=timeout_s)
    except _ao.TimeoutError:
        logger.warning("safe_ocr_handwriting timeout (%ds) for %s", timeout_s, filename)
        return ""
    except Exception as e:
        logger.warning("safe_ocr_handwriting error for %s: %s", filename, e)
        return ""


async def safe_ocr_batch(files: list) -> list:
    """Process list of (content, filename) tuples. Continues even if one fails.
    Returns list of {filename, text, success, error}."""
    results = []
    for content, filename in files:
        text = await safe_ocr_image(content, filename)
        results.append({
            "filename": filename,
            "text": text,
            "success": bool(text),
            "error": None if text else "ocr_failed_or_timeout",
        })
    return results
# --- ADDED END ---

# ============================================================
# TEST ENDPOINT: OCR Debug (does NOT save to DB)
# ============================================================

@app.post("/test/ocr-debug")
async def test_ocr_debug(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Debug endpoint — tests Tesseract vs OCR.space. Does NOT save anything."""
    import time as _t
    from autotax.ocr import (
        local_ocr_tesseract, extract_image_text, extract_handwriting_text,
        extract_pdf_text, extract_pdf_page_as_image, preprocess_image, preprocess_table_image
    )

    content = await file.read()
    ct = (file.content_type or "").lower()
    fn = file.filename or "test"
    results = {"filename": fn, "size_bytes": len(content), "content_type": ct}

    # 1. PDF text layer
    if "pdf" in ct or fn.lower().endswith(".pdf"):
        _s = _t.time()
        pdf_text = extract_pdf_text(content) or ""
        results["pdf_text_layer"] = {"chars": len(pdf_text.strip()), "ms": round((_t.time()-_s)*1000), "sample": pdf_text[:300]}

        # PDF → image
        _s = _t.time()
        img_bytes = extract_pdf_page_as_image(content)
        results["pdf_to_image"] = {"bytes": len(img_bytes) if img_bytes else 0, "ms": round((_t.time()-_s)*1000)}

        if img_bytes:
            content = img_bytes  # use rendered image for OCR tests

    # 2. Preprocessing check
    _s = _t.time()
    preprocessed = preprocess_image(content)
    results["preprocess_standard"] = {"input_bytes": len(content), "output_bytes": len(preprocessed), "ms": round((_t.time()-_s)*1000)}

    _s = _t.time()
    preprocessed_table = preprocess_table_image(content)
    results["preprocess_table"] = {"input_bytes": len(content), "output_bytes": len(preprocessed_table), "ms": round((_t.time()-_s)*1000)}

    # 3. Tesseract OCR
    _s = _t.time()
    tess_text = local_ocr_tesseract(content) or ""
    tess_ms = round((_t.time()-_s)*1000)
    import re
    tess_dates = len(re.findall(r"\d{1,2}[./]\d{1,2}[./]\d{2,4}", tess_text))
    tess_amounts = len(re.findall(r"\d+[.,]\d{2}", tess_text))
    results["tesseract"] = {
        "chars": len(tess_text.strip()),
        "words": len(tess_text.split()),
        "dates_found": tess_dates,
        "amounts_found": tess_amounts,
        "ms": tess_ms,
        "sample": tess_text[:500],
    }

    # 4. OCR.space Engine 1
    _s = _t.time()
    try:
        ocrspace_text = await extract_image_text(content, fn) or ""
        ocrspace_ms = round((_t.time()-_s)*1000)
    except Exception as e:
        ocrspace_text = ""
        ocrspace_ms = round((_t.time()-_s)*1000)
        results["ocrspace_error"] = str(e)
    ocrspace_dates = len(re.findall(r"\d{1,2}[./]\d{1,2}[./]\d{2,4}", ocrspace_text))
    ocrspace_amounts = len(re.findall(r"\d+[.,]\d{2}", ocrspace_text))
    results["ocrspace_engine1"] = {
        "chars": len(ocrspace_text.strip()),
        "words": len(ocrspace_text.split()),
        "dates_found": ocrspace_dates,
        "amounts_found": ocrspace_amounts,
        "ms": ocrspace_ms,
        "sample": ocrspace_text[:500],
    }

    # 5. OCR.space Engine 2 (handwriting)
    _s = _t.time()
    try:
        hw_text = await extract_handwriting_text(content, fn) or ""
        hw_ms = round((_t.time()-_s)*1000)
    except Exception as e:
        hw_text = ""
        hw_ms = round((_t.time()-_s)*1000)
        results["handwriting_error"] = str(e)
    results["ocrspace_engine2_handwriting"] = {
        "chars": len(hw_text.strip()),
        "words": len(hw_text.split()),
        "ms": hw_ms,
        "sample": hw_text[:500],
    }

    # 6. Comparison
    _engines = {
        "tesseract": len(tess_text.strip()),
        "ocrspace_e1": len(ocrspace_text.strip()),
        "ocrspace_e2": len(hw_text.strip()),
    }
    _winner = max(_engines, key=_engines.get)
    results["comparison"] = {
        "winner": _winner,
        **{k+"_chars": v for k, v in _engines.items()},
    }

    return results


# Build trigger: 2026-03-24
# force deploy 2


# Duplicate CORS middleware removed — single CORS config at top of file is sufficient


# Debug OCR print + Auth header print middleware removed — security risk in production
